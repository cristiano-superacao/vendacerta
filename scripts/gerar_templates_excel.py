"""
Script para gerar templates Excel para importação de dados
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Criar diretório se não existir
os.makedirs('../static/templates_excel', exist_ok=True)

# ============================================
# TEMPLATE VENDEDORES
# ============================================
print("Criando template de vendedores...")
wb_vendedores = openpyxl.Workbook()
ws_vendedores = wb_vendedores.active
ws_vendedores.title = 'Vendedores'

# Headers
headers_vendedores = ['Nome', 'Email', 'Telefone', 'CPF', 'Supervisor Email', 'Equipe Nome']
ws_vendedores.append(headers_vendedores)

# Estilizar headers
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
for cell in ws_vendedores[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Dados de exemplo
ws_vendedores.append([
    'João Silva',
    'joao@empresa.com',
    '(11) 98765-4321',
    '123.456.789-00',
    'maria@empresa.com',
    'Equipe A'
])
ws_vendedores.append([
    'Maria Santos',
    'maria@empresa.com',
    '(11) 98765-4322',
    '123.456.789-01',
    '',
    'Equipe B'
])

# Ajustar largura das colunas
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_vendedores.column_dimensions[col].width = 22

# Salvar
wb_vendedores.save('../static/templates_excel/template_vendedores.xlsx')
print("✅ Template vendedores criado!")

# ============================================
# TEMPLATE METAS
# ============================================
print("\nCriando template de metas...")
wb_metas = openpyxl.Workbook()
ws_metas = wb_metas.active
ws_metas.title = 'Metas'

# Headers
headers_metas = [
    'Vendedor Email',
    'Mês',
    'Ano',
    'Meta Vendas'
]
ws_metas.append(headers_metas)

# Estilizar headers
for cell in ws_metas[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Dados de exemplo
ws_metas.append([
    'joao@empresa.com',
    1,
    2024,
    50000.00
])
ws_metas.append([
    'maria@empresa.com',
    1,
    2024,
    60000.00
])

# Ajustar largura das colunas
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_metas.column_dimensions[col].width = 20

# Salvar
wb_metas.save('../static/templates_excel/template_metas.xlsx')
print("✅ Template metas criado!")

# ============================================
# TEMPLATE SUPERVISORES
# ============================================
print("\nCriando template de supervisores...")
wb_supervisores = openpyxl.Workbook()
ws_supervisores = wb_supervisores.active
ws_supervisores.title = 'Supervisores'

# Headers (sem telefone - Usuario não tem esse campo)
headers_supervisores = ['Nome', 'Email']
ws_supervisores.append(headers_supervisores)

# Estilizar headers
for cell in ws_supervisores[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Dados de exemplo
ws_supervisores.append([
    'Carlos Supervisor',
    'carlos@empresa.com'
])
ws_supervisores.append([
    'Ana Supervisora',
    'ana@empresa.com'
])

# Ajustar largura das colunas
for col in ['A', 'B']:
    ws_supervisores.column_dimensions[col].width = 30

# Salvar
wb_supervisores.save('../static/templates_excel/template_supervisores.xlsx')
print("✅ Template supervisores criado!")

print("\n" + "="*50)
print("Todos os templates foram criados com sucesso!")
print("="*50)
