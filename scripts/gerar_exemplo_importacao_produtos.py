"""
Script para gerar exemplo de planilha de importação de produtos
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Criar workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Produtos"

# Definir estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Cabeçalhos conforme imagem
headers = [
    'Código', 'Nome', 'Referência', 'Código de barra', 'NCM', 
    'Data Hora', 'Qte entrada', 'Qte saída', 'O.S',
    'Administrativo', 'Técnico', 'Vendedor', 'Supervisor', 'Gerente', 'Status'
]

# Aplicar cabeçalhos
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Ajustar largura das colunas
column_widths = [15, 30, 20, 20, 15, 20, 12, 12, 15, 15, 12, 12, 12, 12, 12]
for col, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Adicionar exemplos
exemplos = [
    ['PROD001', 'Placa Mãe Asus', 'ASUS-B450M', '7891234567890', '84733020', '', '10', '', '', 'Sim', 'Sim', '', '', '', 'Ativo'],
    ['PROD002', 'Memória RAM 8GB DDR4', 'KVR26N19S8', '7899876543210', '84733021', '', '20', '', '', 'Sim', 'Sim', 'Sim', '', '', 'Ativo'],
    ['PROD003', 'SSD 240GB Kingston', 'SA400S37', '7891122334455', '84717050', '', '15', '', '', 'Sim', 'Sim', '', '', '', 'Ativo'],
    ['PROD004', 'Fonte 500W Real', 'FNT-500W', '7896655443322', '85044090', '', '8', '', '', 'Sim', '', '', 'Sim', '', 'Ativo'],
    ['PROD005', 'Cabo HDMI 2m', 'HDMI-2M', '7893344556677', '85444290', '', '50', '', '', '', 'Sim', 'Sim', '', '', 'Ativo'],
]

for row_num, exemplo in enumerate(exemplos, start=2):
    for col, value in enumerate(exemplo, start=1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left' if col <= 5 else 'center', vertical='center')

# Adicionar instruções em nova aba
ws_info = wb.create_sheet("Instruções")
ws_info.column_dimensions['A'].width = 80

instrucoes = [
    "INSTRUÇÕES PARA IMPORTAÇÃO DE PRODUTOS",
    "",
    "1. Preencha os dados na aba 'Produtos'",
    "2. Campos obrigatórios: Código, Nome",
    "3. Campos numéricos: Qte entrada, Qte saída (deixar vazio ou 0 se não aplicável)",
    "4. O.S: Preencher manualmente quando necessário (número da ordem de serviço)",
    "5. Status: Ativo ou Inativo",
    "6. Produtos com código já cadastrado serão ignorados automaticamente",
    "7. Novos produtos receberão ID único automaticamente no sistema",
    "8. Data Hora será preenchida automaticamente no momento da importação",
    "",
    "COLUNAS DE FUNCIONÁRIOS:",
    "- Use 'Sim' ou deixe vazio nas colunas: Administrativo, Técnico, Vendedor, Supervisor, Gerente",
    "- Estas colunas são opcionais e servem para controle de quem pode acessar o produto",
    "",
    "DICAS:",
    "- Código: Use um padrão único para cada produto (ex: PROD001, PROD002...)",
    "- NCM: Código de 8 dígitos da Nomenclatura Comum do Mercosul",
    "- Código de barra: Código EAN-13 (13 dígitos) ou EAN-8 (8 dígitos)",
    "- Referência: Código do fabricante ou referência interna",
]

for row, texto in enumerate(instrucoes, start=1):
    cell = ws_info.cell(row=row, column=1)
    cell.value = texto
    if row == 1:
        cell.font = Font(bold=True, size=14, color="1F4E78")
    elif texto.startswith(('COLUNAS', 'DICAS')):
        cell.font = Font(bold=True, size=12, color="4472C4")

# Salvar arquivo
filename = f'exemplo_importacao_produtos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(filename)

print(f"✅ Arquivo gerado com sucesso: {filename}")
print(f"📋 {len(exemplos)} produtos de exemplo incluídos")
print(f"📖 Aba 'Instruções' com orientações completas")
