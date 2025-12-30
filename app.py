# app.py - Sistema VendaCerta - Gestão Completa de Vendas, Clientes e Metas
import os
import sys
import re
from io import BytesIO
from dotenv import load_dotenv

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

from flask import (
    Flask,
    render_template,
    redirect,
    url_for,
    flash,
    request,
    jsonify,
    send_file,
    send_from_directory,
)
from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    login_required,
    current_user,
)
from werkzeug.middleware.proxy_fix import ProxyFix
from datetime import datetime
from functools import wraps
from werkzeug.utils import secure_filename
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import threading
import time
import secrets
import shutil
import atexit

from config import config
from models import (
    db,
    Usuario,
    Vendedor,
    Meta,
    Equipe,
    Empresa,
    FaixaComissao,
    FaixaComissaoVendedor,
    FaixaComissaoSupervisor,
    Mensagem,
    Cliente,
    CompraCliente,
    Produto,
    EstoqueMovimento,
    Tecnico,
    OrdemServico,
    Configuracao,
)
from forms import (
    LoginForm,
    RegistroForm,
    VendedorForm,
    MetaForm,
    EquipeForm,
    RecuperarSenhaForm,
    RedefinirSenhaForm,
    EmpresaForm,
    ClienteForm,
    CompraClienteForm,
    ProdutoForm,
    EstoqueMovimentoForm,
    OrdemServicoForm,
    OrdemServicoAvaliarForm,
    OrdemServicoAndamentoForm,
    OrdemServicoAvaliacaoForm,
    UsuarioForm,
)
from calculo_projecao import calcular_projecao_mes, formatar_moeda

# Importar helpers reutilizáveis
from helpers import (
    limpar_cpf,
    limpar_cnpj,
    limpar_telefone,
    formatar_cpf,
    formatar_cnpj,
    formatar_telefone,
    flash_sucesso,
    flash_erro,
    flash_aviso,
    flash_info,
    filtrar_vendedores_por_escopo,
    filtrar_clientes_por_escopo,
    paginar_query,
    validar_email,
    gerar_codigo_cliente,
    calcular_porcentagem,
    pode_importar,
    pode_exportar,
    get_cargos_permitidos_importacao,
    get_cargos_permitidos_exportacao,
    validar_arquivo_excel,
    verificar_excel_disponivel,
)
try:
    # Funções de backup (opcional em produção/Railway)
    from backup_helper import criar_backup_db, listar_backups, restaurar_backup, deletar_backup
except Exception as e:
    # Evita falha de inicialização caso o módulo não esteja disponível no ambiente
    print(f"[AVISO] Módulo 'backup_helper' indisponível: {e}")
    print("[INFO] Recursos de backup serão desativados, mantendo o app online.")
    def criar_backup_db(*args, **kwargs):
        return None
    def listar_backups(*args, **kwargs):
        return []
    def restaurar_backup(*args, **kwargs):
        return False
    def deletar_backup(*args, **kwargs):
        return False

ENABLE_EXCEL_CHECK = os.environ.get("ENABLE_EXCEL_CHECK", "0") == "1"
# Imports para Excel (pandas e openpyxl) - somente se habilitado
if ENABLE_EXCEL_CHECK:
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        EXCEL_AVAILABLE = True
        EXCEL_ERROR_MESSAGE = None
        print("✅ Bibliotecas Excel carregadas com sucesso")
    except ImportError as e:
        EXCEL_AVAILABLE = False
        EXCEL_ERROR_MESSAGE = str(e)
        print(f"⚠️  Aviso: Bibliotecas Excel não disponíveis: {e}")
        print("💡 Instale com: pip install pandas openpyxl")
        # Criar placeholders para evitar erros
        pd = None
        Workbook = None
    except Exception as e:
        EXCEL_AVAILABLE = False
        EXCEL_ERROR_MESSAGE = str(e)
        error_str = str(e)
        print(f"⚠️  Erro ao importar bibliotecas Excel: {e}")
        
        # Detectar erro de biblioteca compartilhada
        if "libstdc++" in error_str or ".so" in error_str:
            print("🔧 SOLUÇÃO: Erro de biblioteca do sistema detectado!")
            print("   No Railway/Nixpacks, adicione ao nixpacks.toml:")
            print('   nixPkgs = ["stdenv.cc.cc.lib", "openblas", "libgfortran"]')
        elif "cannot open shared object" in error_str:
            print("🔧 SOLUÇÃO: Falta biblioteca compartilhada do sistema")
            print("   Verifique as dependências nativas no nixpacks.toml")
        
        pd = None
        Workbook = None
else:
    # Desabilitado por padrão em produção para evitar ruído nos logs
    EXCEL_AVAILABLE = False
    EXCEL_ERROR_MESSAGE = None
    pd = None
    Workbook = None

# Tentativa de (re)carregar libs Excel sob demanda
def ensure_excel_available():
    """Garante que pandas/openpyxl estejam carregados em tempo de requisição.

    Retorna True quando disponíveis; em caso de falha mantém mensagem em
    EXCEL_ERROR_MESSAGE e retorna False.
    """
    global EXCEL_AVAILABLE, EXCEL_ERROR_MESSAGE, pd, Workbook, Font, PatternFill, Alignment
    
    if EXCEL_AVAILABLE and pd is not None and Workbook is not None:
        return True
    
    try:
        import pandas
        import openpyxl
        from openpyxl import Workbook as _WB
        from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment
        
        # Atualizar referências globais
        pd = pandas
        Workbook = _WB
        Font = _Font
        PatternFill = _PatternFill
        Alignment = _Alignment
        EXCEL_AVAILABLE = True
        EXCEL_ERROR_MESSAGE = None
        print("✅ Excel libs habilitadas por lazy-load")
        return True
    except Exception as e:
        EXCEL_AVAILABLE = False
        EXCEL_ERROR_MESSAGE = str(e)
        print(f"❌ Falha ao habilitar Excel por lazy-load: {e}")
        
        # Debug avançado para logs do Railway
        import os
        print(f"🔍 DEBUG ENV:")
        print(f"   LD_LIBRARY_PATH: {os.environ.get('LD_LIBRARY_PATH', 'N/A')}")
        print(f"   PATH: {os.environ.get('PATH', 'N/A')}")
        try:
            # Tentar listar bibliotecas disponíveis (apenas Linux)
            if os.path.exists('/usr/lib'):
                print(f"   /usr/lib exists")
            if os.path.exists('/nix/store'):
                print(f"   /nix/store exists")
        except:
            pass
            
        return False

# (endpoints de status movidos para depois da inicialização do app)

# Flag para evitar imports pesados durante inicialização do banco
INIT_DB_ONLY = os.environ.get("INIT_DB_ONLY", "0") == "1"

# Importar compressão se disponível
try:
    from flask_compress import Compress

    COMPRESS_AVAILABLE = True
except ImportError:
    COMPRESS_AVAILABLE = False
    print("Aviso: flask-compress não disponível.")
    print("Instale com: pip install flask-compress")

# Importar cache se disponível
try:
    from flask_caching import Cache

    CACHE_AVAILABLE = True
except ImportError:
    CACHE_AVAILABLE = False
    print("Aviso: flask-caching não disponível.")
    print("Instale com: pip install flask-caching")

# Importar gerador de PDF (exceto durante inicialização do DB)
if not INIT_DB_ONLY:
    try:
        from pdf_generator import gerar_pdf_metas, gerar_pdf_dashboard, gerar_pdf_metas_supervisor
    except ImportError as e:
        print(f"Aviso: Erro ao importar pdf_generator: {e}")

        def gerar_pdf_metas(*args, **kwargs):
            raise RuntimeError("PDF generator não disponível")

        def gerar_pdf_dashboard(*args, **kwargs):
            raise RuntimeError("PDF generator não disponível")

        def gerar_pdf_metas_supervisor(*args, **kwargs):
            raise RuntimeError("PDF generator não disponível")

else:
    # Placeholders quando INIT_DB_ONLY está ativo
    def gerar_pdf_metas(*args, **kwargs):
        raise RuntimeError("PDF generator desabilitado")

    def gerar_pdf_dashboard(*args, **kwargs):
        raise RuntimeError("PDF generator desabilitado")

    def gerar_pdf_metas_supervisor(*args, **kwargs):
        raise RuntimeError("PDF generator desabilitado")

# Inicialização do aplicativo Flask
app = Flask(__name__)

# Carregar configurações
env = os.environ.get("FLASK_ENV", "development")
app.config.from_object(config[env])

# Configurar compressão Gzip (reduz tamanho das respostas em 70-90%)
if COMPRESS_AVAILABLE:
    app.config["COMPRESS_MIMETYPES"] = [
        "text/html",
        "text/css",
        "text/xml",
        "application/json",
        "application/javascript",
        "text/javascript",
    ]
    # Nível de compressão (1-9, 6 é bom balanço)
    app.config["COMPRESS_LEVEL"] = 6
    # Comprimir apenas respostas > 500 bytes
    app.config["COMPRESS_MIN_SIZE"] = 500
    Compress(app)
    print("[OK] Compressao Gzip ativada - Respostas serao 70-90% menores")

# Configurar cache (acelera relatórios e dashboards em 40-60%)
cache = None
if CACHE_AVAILABLE:
    app.config["CACHE_TYPE"] = "SimpleCache"  # Memória local
    app.config["CACHE_DEFAULT_TIMEOUT"] = 300  # 5 minutos
    cache = Cache(app)
    print("[OK] Cache ativado - Relatorios e dashboards 40-60% mais rapidos")

# Configurar ProxyFix para confiar nos headers do Railway
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Inicializar extensões com tratamento de erro
try:
    db.init_app(app)
    print("[OK] Database inicializado com sucesso")
    
    # Testar conexão com o banco
    with app.app_context():
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            print("[OK] Conexão com banco de dados estabelecida")
        except Exception as db_err:
            print(f"[ERRO] Falha ao conectar ao banco: {db_err}")
            print("[AVISO] Sistema continuará, mas pode ter problemas")
except Exception as e:
    print(f"[ERRO] Falha na inicialização do banco: {e}")
    print("[AVISO] Sistema pode não funcionar corretamente")

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Por favor, faça login para acessar esta página."
login_manager.login_message_category = "info"

# ===== SISTEMA DE BACKUP AUTOMÁTICO =====

# Variável global para armazenar configurações de backup
backup_config = {
    "enabled": True,
    "frequency": "daily",  # daily, weekly, monthly
    "time": "02:00",  # Horário (formato HH:MM)
    "keep_last": 7,  # Manter últimos N backups
    "auto_cleanup": True,  # Limpar backups antigos automaticamente
}

def criar_backup_automatico():
    """Cria backup automático do banco de dados"""
    with app.app_context():
        try:
            uri = app.config.get("SQLALCHEMY_DATABASE_URI", "")

            # Pular se for PostgreSQL (Railway gerencia)
            if "postgresql" in uri:
                msg = "Backup automático: PostgreSQL gerenciado"
                app.logger.info(msg)
                return

            # Para SQLite
            backup_dir = os.path.join(app.instance_path, "backups")
            os.makedirs(backup_dir, exist_ok=True)

            # Nome do backup com timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"auto_backup_{timestamp}.db"
            backup_path = os.path.join(backup_dir, backup_name)

            # Copiar banco de dados
            db_path = uri.replace("sqlite:///", "")

            # Se for caminho relativo, usar instance_path
            if not os.path.isabs(db_path):
                db_path = os.path.join(app.instance_path, db_path)

            if os.path.exists(db_path):
                shutil.copy2(db_path, backup_path)
                app.logger.info(f"[OK] Backup automatico criado: {backup_name}")

                # Limpar backups antigos se configurado
                if backup_config["auto_cleanup"]:
                    limpar_backups_antigos(backup_dir)

                # Sincronizar com nuvem
                try:
                    from backup_nuvem import sincronizar_backup_nuvem

                    app.logger.info("[PROC] Iniciando sincronizacao com nuvem...")
                    sincronizar_backup_nuvem()
                    app.logger.info("[OK] Sincronizacao com nuvem concluida")
                except ImportError:
                    msg = "Modulo backup_nuvem nao encontrado"
                    app.logger.warning(f"[AVISO] {msg}")
                except Exception as e:
                    msg = f"Erro na sincronização: {str(e)}"
                    app.logger.error(f"{msg}")
            else:
                app.logger.error(
                    "Banco de dados não encontrado para backup"
                )

        except Exception as e:
            app.logger.error(f"Erro ao criar backup: {str(e)}")

def limpar_backups_antigos(backup_dir):
    """Remove backups antigos mantendo apenas os N mais recentes"""
    try:
        # Listar todos os backups automáticos
        backups = []
        for filename in os.listdir(backup_dir):
            is_backup = filename.startswith(
                "auto_backup_"
            ) and filename.endswith(".db")
            if is_backup:
                filepath = os.path.join(backup_dir, filename)
                backups.append(
                    {"path": filepath, "time": os.path.getmtime(filepath)}
                )

        # Ordenar por data (mais recente primeiro)
        backups.sort(key=lambda x: x["time"], reverse=True)

        # Deletar backups excedentes
        keep_count = backup_config["keep_last"]
        if len(backups) > keep_count:
            for backup in backups[keep_count:]:
                os.remove(backup["path"])
                filename = os.path.basename(backup["path"])
                app.logger.info(f"Backup antigo removido: {filename}")

    except Exception as e:
        app.logger.error(f"Erro ao limpar backups: {str(e)}")

# Inicializar scheduler de backups
scheduler = BackgroundScheduler()

def iniciar_backup_automatico():
    """Inicia o agendamento de backups automáticos"""
    try:
        if not backup_config["enabled"]:
            app.logger.info("Backup automático desabilitado")
            return

        # Converter horário configurado
        hora, minuto = map(int, backup_config["time"].split(":"))

        # Configurar trigger baseado na frequência
        if backup_config["frequency"] == "daily":
            trigger = CronTrigger(hour=hora, minute=minuto)
        elif backup_config["frequency"] == "weekly":
            trigger = CronTrigger(day_of_week="sun", hour=hora, minute=minuto)
        elif backup_config["frequency"] == "monthly":
            trigger = CronTrigger(day=1, hour=hora, minute=minuto)
        else:
            trigger = CronTrigger(hour=hora, minute=minuto)  # Padrão: diário

# (bloco de endpoints de status movido para depois do scheduler)

        # Adicionar job ao scheduler
        scheduler.add_job(
            func=criar_backup_automatico,
            trigger=trigger,
            id="backup_automatico",
            name="Backup Automático do Sistema",
            replace_existing=True,
        )

        # Iniciar scheduler apenas se não estiver rodando
        if not scheduler.running:
            scheduler.start()
            freq = backup_config["frequency"]
            time = backup_config["time"]
            msg = f"Backup automatico iniciado: {freq} as {time}"
            app.logger.info(f"[PROC] {msg}")
    except Exception as e:
        app.logger.error(f"[ERRO] Erro ao iniciar backup: {e}")

# Garantir que o scheduler pare quando o app encerrar
def shutdown_scheduler():
    """Encerra o scheduler de forma segura"""
    try:
        if scheduler.running:
            scheduler.shutdown(wait=False)
            app.logger.info("[STOP] Scheduler encerrado com sucesso")
    except Exception as e:
        app.logger.error(f"[ERRO] Erro ao encerrar scheduler: {e}")

atexit.register(shutdown_scheduler)

# Headers de segurança

@app.before_request
def force_https():
    """Força HTTPS em produção"""
    # Permitir tráfego HTTP para health checks (Railway)
    if request.path in ['/ping', '/health']:
        return None

    if not app.debug and request.url.startswith("http://"):
        url = request.url.replace("http://", "https://", 1)
        return redirect(url, code=301)

@app.after_request
def set_security_headers(response):
    """Adiciona headers de segurança HTTP e cache otimizado"""
    # Headers de segurança
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    
    # HSTS - Força HTTPS por 1 ano
    if not app.debug:
        hsts = "max-age=31536000; includeSubDomains"
        response.headers["Strict-Transport-Security"] = hsts
    
    # Content Security Policy (CSP) - Previne XSS
    # Adicionado suporte para Google Translate e serviços externos comuns
    csp_directives = [
        "default-src 'self'",
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' cdn.jsdelivr.net https://translate.googleapis.com https://translate.google.com",
        "style-src 'self' 'unsafe-inline' cdn.jsdelivr.net fonts.googleapis.com https://www.gstatic.com https://translate.googleapis.com",
        "font-src 'self' fonts.gstatic.com cdn.jsdelivr.net data:",
        "img-src 'self' data: https: blob: https://www.gstatic.com https://translate.googleapis.com",
        "connect-src 'self' https://translate.googleapis.com cdn.jsdelivr.net",
        "frame-ancestors 'self'",
        "base-uri 'self'",
        "form-action 'self'"
    ]
    response.headers["Content-Security-Policy"] = "; ".join(csp_directives)
    
    # Cache headers para performance
    if request.path.startswith('/static/'):
        # Cache de 1 ano para arquivos estáticos (imutáveis)
        response.cache_control.max_age = 31536000
        response.cache_control.public = True
        response.cache_control.immutable = True
    elif request.path.startswith(('/ping', '/health')):
        # Sem cache para health checks
        response.cache_control.no_cache = True
        response.cache_control.no_store = True
        response.cache_control.must_revalidate = True
    elif request.path.startswith('/api/'):
        # Cache curto para APIs (5 minutos)
        response.cache_control.max_age = 300
        response.cache_control.public = True
    
    return response

# Inicializar Rate Limiter para proteção contra ataques
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        default_limits=[
            "1000 per day",
            "200 per hour"
        ],
        storage_uri=os.environ.get('RATELIMIT_STORAGE_URL', 'memory://'),
        strategy="fixed-window"
    )
    RATE_LIMIT_AVAILABLE = True
    app.logger.info("[OK] Rate limiting ativado - Proteção contra brute force")
except ImportError:
    RATE_LIMIT_AVAILABLE = False
    limiter = None
    app.logger.warning("[AVISO] Flask-Limiter não disponível - Instale com: pip install Flask-Limiter")

# Decorador para aplicar rate limit em rotas sensíveis
def rate_limit(limit_string):
    """Decorador de rate limit se disponível"""
    def decorator(f):
        if RATE_LIMIT_AVAILABLE and limiter:
            return limiter.limit(limit_string)(f)
        return f
    return decorator

# Decorator para rotas de super admin
def super_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            msg = "Por favor, faça login para acessar esta página."
            flash(msg, "warning")
            return redirect(url_for("login"))
        if not current_user.is_super_admin:
            msg = "Acesso negado. Apenas super administradores."
            flash(msg, "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function

# Decorator para verificar permissões específicas
def permission_required(permission_name):
    """Decorator para verificar se o usuário tem uma permissão específica"""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                msg = "Por favor, faça login para acessar esta página."
                flash(msg, "warning")
                return redirect(url_for("login"))

            # Super admin sempre tem acesso
            if current_user.is_super_admin:
                return f(*args, **kwargs)

            # Admin e Gerente sempre têm acesso (cargos com permissões totais)
            if current_user.cargo in ["admin", "gerente"]:
                return f(*args, **kwargs)

            # Verificar se o atributo existe (compatibilidade com banco antigo)
            if not hasattr(current_user, permission_name):
                # Se não tem o atributo, permitir acesso para cargos de gestão
                if current_user.cargo in ["supervisor"]:
                    return f(*args, **kwargs)
                # Para vendedores sem atributo, bloquear por segurança
                flash(
                    "Sistema em atualização. Entre em contato com o administrador.",
                    "warning",
                )
                return redirect(url_for("dashboard"))

            # Verificar se tem a permissão específica
            has_perm = getattr(current_user, permission_name, False)
            if not has_perm:
                msg = "Você não tem permissão para acessar este recurso."
                flash(msg, "danger")
                return redirect(url_for("dashboard"))

            return f(*args, **kwargs)

        return decorated_function

    return decorator

# Decorator para admin ou superior

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            msg = "Por favor, faça login para acessar esta página."
            flash(msg, "warning")
            return redirect(url_for("login"))

        if current_user.is_super_admin:
            return f(*args, **kwargs)

        if current_user.cargo not in ["admin", "gerente"]:
            msg = (
                "Acesso negado. "
                "Apenas administradores podem acessar este recurso."
            )
            flash(msg, "danger")
            return redirect(url_for("dashboard"))

        return f(*args, **kwargs)

    return decorated_function

# ===== FUNÇÕES AUXILIARES =====

def atualizar_metas_supervisores():
    """
    Atualiza automaticamente as metas dos supervisores com base
    na soma das metas dos vendedores vinculados
    """
    try:
        # Buscar todos os supervisores
        supervisores = Usuario.query.filter_by(cargo="supervisor").all()

        for supervisor in supervisores:
            # Buscar vendedores deste supervisor
            vendedores = Vendedor.query.filter_by(
                supervisor_id=supervisor.id
            ).all()

            if not vendedores:
                continue

            # Agrupar metas por mês/ano
            metas_agrupadas = {}

            for vendedor in vendedores:
                metas = Meta.query.filter_by(vendedor_id=vendedor.id).all()

                for meta in metas:
                    chave = (meta.mes, meta.ano)
                    if chave not in metas_agrupadas:
                        metas_agrupadas[chave] = {
                            "valor_meta": 0,
                            "receita_alcancada": 0,
                        }

                    var = "valor_meta"
                    rec = "receita_alcancada"
                    metas_agrupadas[chave][var] += meta.valor_meta
                    metas_agrupadas[chave][rec] += meta.receita_alcancada

            # Criar ou atualizar meta do supervisor
            # Usando vendedor_id do primeiro vendedor como referência
            # Nota: Para supervisores, podemos criar um vendedor
            # virtual ou usar outro modelo.
            # Por simplicidade, vamos apenas calcular mas não criar
            # registro ainda (isso pode ser expandido no futuro)

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao atualizar metas de supervisores: {str(e)}")

@login_manager.user_loader
def load_user(user_id):
    """Carrega o usuário pelo ID"""
    return Usuario.query.get(int(user_id))

# ===== ROTAS DE AUTENTICAÇÃO =====

@app.route("/favicon.ico")
def favicon():
    """Serve o favicon"""
    return send_from_directory(
        os.path.join(app.root_path, "static"),
        "favicon.ico",
        mimetype="image/vnd.microsoft.icon",
    )

@app.route("/ping")
@app.route("/health")
def health_check():
    """Health check ultrarrápido sem tocar no banco para evitar timeouts."""
    return jsonify({"status": "ok"}), 200

@app.route("/status")
def public_status():
    """Status público detalhado do sistema (sem autenticação)"""
    status_info = {
        "app": "VendaCerta",
        "status": "online",
        "timestamp": datetime.now().isoformat(),
    }
    
    # Verificar conexão com banco
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        status_info["database"] = {
            "status": "connected",
            "type": "postgresql" if "postgresql" in str(db.engine.url) else "sqlite"
        }
    except Exception as e:
        status_info["database"] = {
            "status": "error",
            "error": str(e)
        }
    
    # Verificar variáveis de ambiente críticas
    env_vars = {
        "DATABASE_URL": bool(os.environ.get("DATABASE_URL")),
        "PGHOST": bool(os.environ.get("PGHOST")),
        "PGDATABASE": bool(os.environ.get("PGDATABASE")),
        "FLASK_SECRET_KEY": bool(os.environ.get("FLASK_SECRET_KEY")),
    }
    status_info["environment"] = env_vars
    
    return jsonify(status_info), 200

# ===== ENDPOINTS DE STATUS (Excel/DB/Env) =====
@app.route("/ping")
def ping():
    """Healthcheck simples para Railway"""
    return "pong", 200

@app.route("/status/excel")
@login_required
def status_excel():
    """Verifica disponibilidade de Excel (pandas/openpyxl) retornando JSON."""
    try:
        from helpers import verificar_excel_disponivel
        disponivel, erro = verificar_excel_disponivel(ensure_func=ensure_excel_available)
        return jsonify({
            "available": bool(disponivel),
            "error": erro,
        }), 200
    except Exception as e:
        return jsonify({
            "available": False,
            "error": str(e),
        }), 200


def _check_db_engine(bind_name=None):
    """Verifica conectividade com o banco (default ou bind específico)."""
    try:
        from sqlalchemy import text
        engine = db.get_engine(app, bind=bind_name) if bind_name else db.engine
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True, None
    except Exception as e:
        return False, str(e)


@app.route("/status/db")
@login_required
def status_db():
    """Endpoint JSON para checar conectividade com DB (default e binds)."""
    if (
        current_user.cargo not in ["admin", "supervisor", "gerente"]
        and not getattr(current_user, "is_super_admin", False)
    ):
        return jsonify({"allowed": False}), 403

    binds = app.config.get("SQLALCHEMY_BINDS") or {}
    result = {}

    ok, err = _check_db_engine(None)
    result["default"] = {"available": ok, "error": err}

    for name in binds.keys():
        ok, err = _check_db_engine(name)
        result[name] = {"available": ok, "error": err}

    return jsonify({"status": result}), 200


@app.route("/status/env")
@login_required
def status_env():
    """Endpoint JSON para auditar variáveis críticas de configuração.
    Não expõe valores sensíveis; apenas presença.
    """
    if (
        current_user.cargo not in ["admin", "supervisor", "gerente"]
        and not getattr(current_user, "is_super_admin", False)
    ):
        return jsonify({"allowed": False}), 403

    expected = [
        "DATABASE_URL",
        "URL_DO_BANCO_DE_DADOS",
        "PGDATABASE",
        "PGHOST",
        "PGUSER",
        "PGPASSWORD",
        "PGPORT",
        "FLASK_SECRET_KEY",
        "SECRET_KEY",
        "INIT_DB_ONLY",
    ]

    presence = {name: bool(os.environ.get(name)) for name in expected}
    return jsonify({"env": presence}), 200

# Função para inicializar o banco de dados automaticamente
def init_database():
    """Inicializa o banco de dados criando todas as tabelas necessárias"""
    with app.app_context():
        try:
            app.logger.info("[PROC] Verificando estrutura do banco de dados...")
            
            # Garantir que o diretório instance existe (para SQLite)
            instance_dir = os.path.join(os.path.dirname(__file__), 'instance')
            if not os.path.exists(instance_dir):
                os.makedirs(instance_dir)
                app.logger.info(f"[OK] Diretorio criado: {instance_dir}")
            
            # Criar todas as tabelas se não existirem
            db.create_all()
            
            # Verificar e garantir acesso de admin (Auto-Recovery)
            admins_to_check = ['admin@vendacerta.com', 'admin@metas.com']
            admin_found = False
            
            for email in admins_to_check:
                admin = Usuario.query.filter_by(email=email).first()
                if admin:
                    # Se existir, garantir que a senha seja admin123 e seja super admin
                    from werkzeug.security import generate_password_hash
                    admin.senha_hash = generate_password_hash('admin123')
                    admin.is_super_admin = True
                    admin.ativo = True
                    admin.cargo = 'admin'
                    db.session.commit()
                    app.logger.info(f"[OK] Admin {email} recuperado/resetado para senha 'admin123'")
                    admin_found = True
            
            # Se nenhum dos dois existir, criar o padrão
            if not admin_found:
                app.logger.info("[PROC] Criando usuario admin padrao...")
                from werkzeug.security import generate_password_hash
                
                # Criar empresa padrão se necessário
                empresa_padrao = Empresa.query.filter_by(cnpj='00000000000000').first()
                if not empresa_padrao:
                    empresa_padrao = Empresa(
                        nome='Empresa Padrão',
                        cnpj='00000000000000',
                        email='contato@empresa.com',
                        plano='enterprise',
                        ativo=True
                    )
                    db.session.add(empresa_padrao)
                    db.session.commit()

                admin = Usuario(
                    nome='Administrador',
                    email='admin@vendacerta.com',
                    senha_hash=generate_password_hash('admin123'),
                    cargo='admin',
                    is_super_admin=True,
                    ativo=True,
                    empresa_id=empresa_padrao.id
                )
                db.session.add(admin)
                db.session.commit()
                app.logger.info("[OK] Usuario admin criado (email: admin@vendacerta.com, senha: admin123)")

            app.logger.info("[OK] Banco de dados inicializado com sucesso!")
            return True
            
        except Exception as e:
            app.logger.error(f"Erro ao inicializar banco de dados: {e}")
            return False

# Inicialização do banco: resiliente e não bloqueante em produção
RUN_DB_INIT_ON_START = os.environ.get("RUN_DB_INIT_ON_START", "0") == "1"

def _init_db_background(max_attempts: int = 3, delay_seconds: int = 5):
    # Pequeno atraso inicial para permitir correção de schema (fix_database_railway)
    time.sleep(10)
    attempts = 0
    while attempts < max_attempts:
        attempts += 1
        try:
            with app.app_context():
                ok = init_database()
                if ok:
                    app.logger.info("[OK] Inicialização de banco concluída (background)")
                    return
        except Exception as e:
            app.logger.error(f"[ERRO] Tentativa {attempts}/{max_attempts} de inicializar banco falhou: {e}")
        time.sleep(delay_seconds)
    app.logger.warning("[AVISO] Inicialização de banco não concluída após tentativas. Continuando em modo degradado.")

if RUN_DB_INIT_ON_START:
    # Ambiente de desenvolvimento: permitir inicialização síncrona
    with app.app_context():
        try:
            init_database()
            iniciar_backup_automatico()
        except Exception as e:
            app.logger.error(f"Erro na inicialização: {e}")
else:
    # Produção: executar em background para evitar atrasos na saúde do serviço
    threading.Thread(target=_init_db_background, daemon=True).start()
    # Backup automático permanece desativado até DB ficar OK; pode ser habilitado após init_db

@app.route("/login", methods=["GET", "POST"])
@rate_limit("10 per minute")  # Máximo 10 tentativas de login por minuto
def login():
    """Página de login com proteção contra brute force"""
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    form = LoginForm()
    if form.validate_on_submit():
        usuario = Usuario.query.filter_by(email=form.email.data).first()
        if usuario and usuario.check_senha(form.senha.data):
            if not usuario.ativo:
                msg = (
                    "Sua conta está inativa. "
                    "Entre em contato com o administrador."
                )
                flash(msg, "warning")
                return redirect(url_for("login"))

            login_user(usuario)
            flash(f"Bem-vindo(a), {usuario.nome}!", "success")

            # Redirecionar para a página solicitada ou dashboard adequado
            next_page = request.args.get("next")
            if next_page:
                return redirect(next_page)
            else:
                # Se for vendedor com vínculo de vendedor, redirecionar para dashboard mobile do vendedor
                if usuario.cargo == "vendedor" and usuario.vendedor_id:
                    return redirect(url_for("vendedor_dashboard"))
                else:
                    return redirect(url_for("dashboard"))
        else:
            flash("Email ou senha incorretos.", "danger")

    return render_template("login.html", form=form)

@app.route("/registro", methods=["GET", "POST"])
def registro():
    """Página de registro de nova empresa e usuário"""
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    form = RegistroForm()
    if form.validate_on_submit():
        try:
            import re

            # Limpar CNPJ
            cnpj_limpo = re.sub(r"\D", "", form.cnpj.data)

            # Criar Empresa
            empresa = Empresa(
                nome=form.nome_empresa.data,
                cnpj=cnpj_limpo,
                email=form.email.data,  # Email do admin como contato inicial
                telefone=form.telefone.data,
                ativo=True,
                plano="basico",  # Plano padrão
            )
            db.session.add(empresa)
            db.session.commit()  # Commit para gerar ID

            # Criar Usuário Admin da Empresa
            usuario = Usuario(
                nome=form.nome.data,
                email=form.email.data,
                cargo="admin",  # Sempre admin ao criar empresa
                empresa_id=empresa.id,
                is_super_admin=False,
                ativo=True,
            )
            usuario.set_senha(form.senha.data)

            db.session.add(usuario)
            db.session.commit()

            msg = (
                "Conta empresarial criada com sucesso! "
                "Faça login para começar."
            )
            flash(msg, "success")
            return redirect(url_for("login"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao criar conta: {str(e)}", "danger")

    return render_template("registro.html", form=form)

@app.route("/logout")
@login_required
def logout():
    """Logout do usuário"""
    logout_user()
    flash("Você saiu da sua conta.", "info")
    return redirect(url_for("login"))



# ===== ROTAS PÚBLICAS =====

@app.route("/recuperar-senha", methods=["GET", "POST"])
def recuperar_senha():
    """Página de recuperação de senha"""
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    form = RecuperarSenhaForm()
    if form.validate_on_submit():
        usuario = Usuario.query.filter_by(email=form.email.data).first()
        if usuario:
            # Em produção, aqui você enviaria um email
            # com link de redefinição. Por enquanto, vamos criar
            # um token temporário e mostrar na tela
            token = secrets.token_urlsafe(32)

            # Armazenar token temporariamente (em produção use Redis ou banco)
            if not hasattr(app, "reset_tokens"):
                app.reset_tokens = {}
            app.reset_tokens[token] = usuario.id

            msg = (
                f"Instruções de recuperação enviadas! "
                f"Use este link: /redefinir-senha/{token}"
            )
            flash(msg, "success")
            return redirect(url_for("login"))
        else:
            msg = (
                "Se o email estiver cadastrado, "
                "você receberá instruções de recuperação."
            )
            flash(msg, "info")

    return render_template("recuperar_senha.html", form=form)

@app.route("/redefinir-senha/<token>", methods=["GET", "POST"])
def redefinir_senha(token):
    """Página de redefinição de senha com token"""
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    # Verificar token
    has_tokens = hasattr(app, "reset_tokens")
    if not has_tokens or token not in app.reset_tokens:
        flash("Link de recuperação inválido ou expirado.", "danger")
        return redirect(url_for("login"))

    form = RedefinirSenhaForm()
    if form.validate_on_submit():
        usuario_id = app.reset_tokens[token]
        usuario = Usuario.query.get(usuario_id)

        if usuario:
            usuario.set_senha(form.nova_senha.data)
            db.session.commit()

            # Remover token usado
            del app.reset_tokens[token]

            msg = (
                "Senha redefinida com sucesso! "
                "Faça login com sua nova senha."
            )
            flash(msg, "success")
            return redirect(url_for("login"))

    return render_template("redefinir_senha.html", form=form, token=token)

@app.route("/ajuda")
@login_required
def ajuda():
    """Central de Ajuda e Suporte"""
    return render_template("ajuda.html")

@app.route("/manual")
@login_required
def manual():
    """Servir o manual do usuário"""
    try:
        manual_path = os.path.join(
            os.path.dirname(__file__), "docs", "guias", "MANUAL_USUARIO.md"
        )
        return send_file(
            manual_path,
            mimetype="text/markdown",
            as_attachment=True,
            download_name="Manual_Usuario.md",
        )
    except Exception:
        msg = "Erro ao carregar o manual. Entre em contato com o suporte."
        flash(msg, "danger")
        return redirect(url_for("ajuda"))

# ===== ROTAS DE SUPER ADMIN - GERENCIAMENTO DE EMPRESAS =====

@app.route("/super-admin/empresas")
@super_admin_required
def super_admin_empresas():
    """Lista todas as empresas cadastradas"""
    empresas = Empresa.query.order_by(Empresa.data_criacao.desc()).all()

    # Adicionar estatísticas
    for empresa in empresas:
        empresa.total_usuarios = Usuario.query.filter_by(
            empresa_id=empresa.id
        ).count()

    return render_template("super_admin/empresas.html", empresas=empresas)

@app.route("/super-admin/empresas/criar", methods=["GET", "POST"])
@super_admin_required
def super_admin_criar_empresa():
    """Criar nova empresa"""
    form = EmpresaForm()

    if form.validate_on_submit():

        cnpj_numbers = re.sub(r"\D", "", form.cnpj.data)

        empresa = Empresa(
            nome=form.nome.data,
            cnpj=cnpj_numbers,
            email=form.email.data,
            telefone=form.telefone.data,
            endereco=form.endereco.data,
            cidade=form.cidade.data,
            estado=form.estado.data.upper() if form.estado.data else None,
            plano=form.plano.data,
            max_usuarios=form.max_usuarios.data,
            max_vendedores=form.max_vendedores.data,
            ativo=True,
            bloqueado=False,
        )

        db.session.add(empresa)
        db.session.commit()

        flash(f"Empresa {empresa.nome} criada com sucesso!", "success")
        return redirect(url_for("super_admin_empresas"))

    return render_template(
        "super_admin/empresa_form.html", form=form, titulo="Criar Nova Empresa"
    )

@app.route("/super-admin/empresas/<int:id>/editar", methods=["GET", "POST"])
@super_admin_required
def super_admin_editar_empresa(id):
    """Editar empresa existente"""
    empresa = Empresa.query.get_or_404(id)
    form = EmpresaForm(empresa_id=id, obj=empresa)

    if form.validate_on_submit():

        cnpj_numbers = re.sub(r"\D", "", form.cnpj.data)

        empresa.nome = form.nome.data
        empresa.cnpj = cnpj_numbers
        empresa.email = form.email.data
        empresa.telefone = form.telefone.data
        empresa.endereco = form.endereco.data
        empresa.cidade = form.cidade.data
        if form.estado.data:
            empresa.estado = form.estado.data.upper()
        else:
            empresa.estado = None
        empresa.plano = form.plano.data
        empresa.max_usuarios = form.max_usuarios.data
        empresa.max_vendedores = form.max_vendedores.data

        db.session.commit()

        flash(f"Empresa {empresa.nome} atualizada com sucesso!", "success")
        return redirect(url_for("super_admin_empresas"))

    titulo = f"Editar Empresa: {empresa.nome}"
    return render_template(
        "super_admin/empresa_form.html",
        form=form,
        titulo=titulo,
        empresa=empresa,
    )

@app.route("/super-admin/empresas/<int:id>/bloquear", methods=["POST"])
@super_admin_required
def super_admin_bloquear_empresa(id):
    """Bloquear/Desbloquear empresa"""
    empresa = Empresa.query.get_or_404(id)

    motivo = request.form.get("motivo", "")

    if empresa.bloqueado:
        empresa.bloqueado = False
        empresa.motivo_bloqueio = None
        msg = f"Empresa {empresa.nome} desbloqueada com sucesso!"
        flash(msg, "success")
    else:
        empresa.bloqueado = True
        if motivo:
            empresa.motivo_bloqueio = motivo
        else:
            empresa.motivo_bloqueio = "Bloqueado pelo administrador"
        flash(f"Empresa {empresa.nome} bloqueada com sucesso!", "warning")

    db.session.commit()
    return redirect(url_for("super_admin_empresas"))

@app.route("/super-admin/empresas/<int:id>/excluir", methods=["POST"])
@super_admin_required
def super_admin_excluir_empresa(id):
    """Excluir empresa (desativa, não remove do banco)"""
    empresa = Empresa.query.get_or_404(id)

    # Desativar empresa ao invés de excluir
    empresa.ativo = False
    empresa.bloqueado = True
    empresa.motivo_bloqueio = "Empresa excluída pelo administrador"

    # Desativar todos os usuários da empresa
    Usuario.query.filter_by(empresa_id=empresa.id).update({"ativo": False})

    db.session.commit()

    flash(f"Empresa {empresa.nome} excluída com sucesso!", "success")
    return redirect(url_for("super_admin_empresas"))

@app.route("/super-admin/empresas/<int:id>/excluir-definitivamente", methods=["POST"])
@super_admin_required
def super_admin_excluir_empresa_definitivamente(id):
    """Excluir empresa e TODOS os seus dados permanentemente"""
    from sqlalchemy import or_
    empresa = Empresa.query.get_or_404(id)
    
    try:
        # 1. Excluir Movimentações de Estoque
        EstoqueMovimento.query.filter_by(empresa_id=empresa.id).delete()
        
        # 2. Excluir Ordens de Serviço
        OrdemServico.query.filter_by(empresa_id=empresa.id).delete()
        
        # 3. Excluir Compras de Clientes
        CompraCliente.query.filter_by(empresa_id=empresa.id).delete()
        
        # 4. Excluir Metas (via Vendedores)
        vendedores = Vendedor.query.filter_by(empresa_id=empresa.id).all()
        for v in vendedores:
            Meta.query.filter_by(vendedor_id=v.id).delete()
            
        # 5. Excluir Clientes
        Cliente.query.filter_by(empresa_id=empresa.id).delete()
        
        # 6. Excluir Vendedores
        Vendedor.query.filter_by(empresa_id=empresa.id).delete()
        
        # 7. Excluir Técnicos
        Tecnico.query.filter_by(empresa_id=empresa.id).delete()
        
        # 8. Excluir Produtos
        Produto.query.filter_by(empresa_id=empresa.id).delete()
        
        # 9. Excluir Equipes
        Equipe.query.filter_by(empresa_id=empresa.id).delete()
        
        # 10. Excluir Faixas de Comissão
        FaixaComissao.query.filter_by(empresa_id=empresa.id).delete()
        FaixaComissaoVendedor.query.filter_by(empresa_id=empresa.id).delete()
        FaixaComissaoSupervisor.query.filter_by(empresa_id=empresa.id).delete()
        
        # 11. Excluir Mensagens (Remetente ou Destinatário na empresa)
        usuarios = Usuario.query.filter_by(empresa_id=empresa.id).all()
        usuario_ids = [u.id for u in usuarios]
        if usuario_ids:
            Mensagem.query.filter(
                or_(
                    Mensagem.remetente_id.in_(usuario_ids),
                    Mensagem.destinatario_id.in_(usuario_ids)
                )
            ).delete(synchronize_session=False)
        
        # 12. Excluir Usuários
        Usuario.query.filter_by(empresa_id=empresa.id).delete()
        
        # 13. Excluir a Empresa
        db.session.delete(empresa)
        
        db.session.commit()
        flash(f"Empresa {empresa.nome} e TODOS os seus dados foram excluídos permanentemente!", "success")
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Erro ao excluir empresa definitivamente: {str(e)}")
        flash(f"Erro ao excluir empresa: {str(e)}", "danger")

    return redirect(url_for("super_admin_empresas"))

@app.route("/super-admin/empresas/<int:id>/visualizar")
@super_admin_required
def super_admin_visualizar_empresa(id):
    """Visualizar detalhes da empresa"""
    empresa = Empresa.query.get_or_404(id)
    usuarios = Usuario.query.filter_by(empresa_id=empresa.id).all()
    vendedores = Vendedor.query.filter_by(empresa_id=empresa.id).all()
    equipes = Equipe.query.filter_by(empresa_id=empresa.id).all()

    stats = {
        "total_usuarios": len(usuarios),
        "total_vendedores": len(vendedores),
        "total_equipes": len(equipes),
    }

    return render_template(
        "super_admin/empresa_detalhes.html",
        empresa=empresa,
        usuarios=usuarios,
        vendedores=vendedores,
        equipes=equipes,
        stats=stats,
    )

# ===== ROTAS DE SUPER ADMIN - GERENCIAMENTO DE USUÁRIOS =====

@app.route("/super-admin/usuarios")
@super_admin_required
def super_admin_usuarios():
    """Lista todos os usuários do sistema"""
    empresa_filtro = request.args.get("empresa_id", type=int)

    if empresa_filtro:
        usuarios = Usuario.query.filter_by(empresa_id=empresa_filtro).all()
    else:
        usuarios = Usuario.query.filter(
            Usuario.is_super_admin.is_(False)
        ).all()

    empresas = Empresa.query.all()

    # Enriquecer dados dos usuários com informações adicionais
    usuarios_enriquecidos = []
    for usuario in usuarios:
        # Contar vendedores vinculados (se for supervisor)
        total_vendedores = 0
        if usuario.cargo == "supervisor":
            total_vendedores = Vendedor.query.filter_by(
                supervisor_id=usuario.id, ativo=True
            ).count()

        # Buscar nome do gerente/administrador responsável
        gerente_nome = None
        if usuario.gerente_id:
            gerente = Usuario.query.get(usuario.gerente_id)
            if gerente:
                gerente_nome = gerente.nome

        usuarios_enriquecidos.append(
            {
                "usuario": usuario,
                "total_vendedores": total_vendedores,
                "gerente_nome": gerente_nome,
            }
        )

    return render_template(
        "super_admin/usuarios.html",
        usuarios_enriquecidos=usuarios_enriquecidos,
        usuarios=usuarios,
        empresas=empresas,
        empresa_filtro=empresa_filtro,
    )

# ===== ROTAS DE SUPERVISORES =====

@app.route("/supervisores/importar", methods=["GET", "POST"])
@login_required
def importar_supervisores():
    """Importar supervisores via planilha Excel - Admin, Supervisor e RH"""
    # Verificar permissões: apenas admin, supervisor e RH podem importar supervisores
    if not pode_importar(current_user, "supervisores"):
        flash(
            "Acesso negado! Apenas Administradores, Supervisores e RH podem importar supervisores.",
            "danger",
        )
        return redirect(url_for("lista_supervisores"))

    if request.method == "POST":
        try:
            # Validar arquivo Excel
            arquivo, erro = validar_arquivo_excel(request)
            if erro:
                flash(erro, "danger")
                return redirect(request.url)

            # Garantir bibliotecas de Excel disponíveis
            if not EXCEL_AVAILABLE and not ensure_excel_available():
                flash("Erro: Bibliotecas Excel não instaladas.", "danger")
                return redirect(request.url)

            # Ler Excel em DataFrame
            df = pd.read_excel(arquivo)

            # Normalizar nomes das colunas
            df.columns = df.columns.str.strip().str.lower()

            # Mapear possíveis nomes de colunas para o padrão
            colunas_map = {
                "nome": ["nome", "nome completo", "supervisor", "nome supervisor"],
                "email": ["email", "e-mail", "e mail"],
                "empresa_cnpj": ["empresa cnpj", "empresa_cnpj", "cnpj empresa"],
            }

            for col_padrao, variantes in colunas_map.items():
                for col in list(df.columns):
                    if col in variantes:
                        df.rename(columns={col: col_padrao}, inplace=True)
                        break

            # Validar colunas obrigatórias
            colunas_obrigatorias = ["nome", "email"]
            if current_user.is_super_admin:
                colunas_obrigatorias.append("empresa_cnpj")

            colunas_faltando = [
                col for col in colunas_obrigatorias if col not in df.columns
            ]

            if colunas_faltando:
                colunas_exibir = [c.capitalize() for c in colunas_faltando]
                msg = f'Colunas obrigatórias faltando: {", ".join(colunas_exibir)}'
                flash(msg, "danger")
                return redirect(request.url)

            erros = []
            sucesso = 0

            for idx, row in df.iterrows():
                try:
                    # Determinar empresa do supervisor
                    if current_user.is_super_admin:
                        empresa_cnpj = str(row.get("empresa_cnpj", "")).strip()
                        if not empresa_cnpj:
                            erros.append(
                                f"Linha {idx + 2}: Coluna 'Empresa CNPJ' obrigatória para Super Admin."
                            )
                            continue

                        empresa = Empresa.query.filter_by(cnpj=empresa_cnpj).first()
                        if not empresa:
                            erros.append(
                                f"Linha {idx + 2}: Empresa CNPJ {empresa_cnpj} não encontrada."
                            )
                            continue
                        empresa_id = empresa.id
                    else:
                        empresa_id = current_user.empresa_id

                    nome = str(row.get("nome", "")).strip()
                    email = str(row.get("email", "")).strip()

                    if not nome or not email:
                        erros.append(
                            f"Linha {idx + 2}: Nome e Email são obrigatórios."
                        )
                        continue

                    if not validar_email(email):
                        erros.append(
                            f"Linha {idx + 2}: Email {email} inválido."
                        )
                        continue

                    # Verificar se email já existe
                    usuario_existente = Usuario.query.filter_by(email=email).first()
                    if usuario_existente:
                        erros.append(
                            f"Linha {idx + 2}: Email {email} já cadastrado no sistema."
                        )
                        continue

                    # Criar supervisor como usuário do sistema
                    supervisor = Usuario(
                        nome=nome,
                        email=email,
                        cargo="supervisor",
                        empresa_id=empresa_id,
                        ativo=True,
                    )
                    supervisor.set_senha("supervisor123")

                    db.session.add(supervisor)
                    sucesso += 1

                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

            if erros:
                db.session.rollback()
                msg_erro = "<br>".join(erros[:10])
                if len(erros) > 10:
                    msg_erro += f"<br>... e mais {len(erros) - 10} erros"
                flash(f"Erros encontrados:<br>{msg_erro}", "warning")
                return render_template("supervisores/importar.html")
            elif sucesso > 0:
                db.session.commit()
                flash(f"{sucesso} supervisor(es) importado(s) com sucesso!", "success")
                return redirect(url_for("lista_supervisores"))
            else:
                flash(
                    "Nenhum supervisor foi importado. Verifique a planilha.",
                    "warning",
                )
                return redirect(request.url)
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao processar arquivo: {str(e)}", "danger")
            return redirect(request.url)

    return render_template("supervisores/importar.html")

@app.route("/supervisores/download-template")
@login_required
def download_template_supervisores():
    """Gera e baixa o template Excel para importação de supervisores"""
    # Verificar permissões
    if not pode_importar(current_user, "supervisores"):
        flash("Acesso negado!", "danger")
        return redirect(url_for("lista_supervisores"))

    if not EXCEL_AVAILABLE and not ensure_excel_available():
        flash("Erro: Bibliotecas Excel não instaladas.", "danger")
        return redirect(url_for("importar_supervisores"))

    try:
        import io
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Supervisores"

        # Definir colunas
        headers = ["Nome", "Email"]
        
        # Adicionar coluna de CNPJ se for Super Admin
        if current_user.is_super_admin:
            headers.append("Empresa CNPJ")

        # Estilos
        header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))

        # Escrever cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
            # Ajustar largura da coluna
            ws.column_dimensions[get_column_letter(col_num)].width = 30

        # Adicionar exemplo na segunda linha
        ws.cell(row=2, column=1, value="João da Silva").border = thin_border
        ws.cell(row=2, column=2, value="joao.silva@email.com").border = thin_border
        
        if current_user.is_super_admin:
            ws.cell(row=2, column=3, value="00.000.000/0001-00").border = thin_border

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="template_supervisores.xlsx"
        )

    except Exception as e:
        app.logger.error(f"Erro ao gerar template: {str(e)}")
        flash(f"Erro ao gerar template: {str(e)}", "danger")
        return redirect(url_for("importar_supervisores"))

@app.route("/supervisores")
@login_required
def lista_supervisores():
    """Lista todos os supervisores"""
    page = request.args.get("page", 1, type=int)

    # Quantidade de registros por página
    per_page = 20

    # Query base por escopo de empresa
    if current_user.is_super_admin:
        base_query = Usuario.query.filter_by(cargo="supervisor", ativo=True)
    else:
        base_query = Usuario.query.filter_by(
            empresa_id=current_user.empresa_id,
            cargo="supervisor",
            ativo=True,
        )

    # Estatísticas globais
    total_supervisores = base_query.count()
    total_vendedores_supervisionados = (
        Vendedor.query.filter(
            Vendedor.ativo.is_(True),
            Vendedor.supervisor_id.in_(
                [s.id for s in base_query.with_entities(Usuario.id).all()]
            ),
        ).count()
        if total_supervisores > 0
        else 0
    )
    media_vendedores = (
        total_vendedores_supervisionados / total_supervisores
        if total_supervisores > 0
        else 0
    )

    stats = {
        "total": total_supervisores,
        "total_vendedores": total_vendedores_supervisionados,
        "media_vendedores": media_vendedores,
    }

    # Paginação ordenada por nome
    pagination = base_query.order_by(Usuario.nome).paginate(
        page=page, per_page=per_page, error_out=False
    )
    supervisores_pagina = pagination.items

    # Preparar dados com contagem de vendedores por supervisor apenas da página
    supervisores_data = []
    for supervisor in supervisores_pagina:
        total_vendedores = Vendedor.query.filter_by(
            supervisor_id=supervisor.id, ativo=True
        ).count()
        supervisores_data.append(
            {"supervisor": supervisor, "total_vendedores": total_vendedores}
        )

    return render_template(
        "supervisores/lista.html",
        supervisores=supervisores_data,
        pagination=pagination,
        stats=stats,
    )

@app.route("/supervisores/novo", methods=["GET", "POST"])
@login_required
def novo_supervisor():
    """Cadastrar novo supervisor"""
    form = UsuarioForm()
    gerentes_disponiveis = _prepare_supervisor_form(form)

    if form.validate_on_submit():
        if not current_user.is_super_admin and form.empresa_id.data != current_user.empresa_id:
            flash("Você só pode criar supervisores na sua empresa!", "danger")
            return redirect(url_for("lista_supervisores"))

        if not request.form.get("gerente_id"):
            flash("É obrigatório vincular o supervisor a um Gerente ou Administrador!", "danger")
        else:
            try:
                supervisor = _save_supervisor_from_form(form)
                flash(f"Supervisor {supervisor.nome} criado com sucesso! Senha padrão: senha123", "success")
                return redirect(url_for("lista_supervisores"))
            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao criar supervisor: {str(e)}", "danger")

    return render_template(
        "supervisores/form.html",
        form=form,
        gerentes_disponiveis=gerentes_disponiveis,
        titulo="Novo Supervisor",
    )

@app.route("/supervisores/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_supervisor(id):
    """Editar supervisor existente"""
    supervisor = Usuario.query.get_or_404(id)

    if supervisor.cargo != "supervisor":
        flash("Este usuário não é um supervisor.", "warning")
        return redirect(url_for("lista_supervisores"))
    if not current_user.is_super_admin and supervisor.empresa_id != current_user.empresa_id:
        flash("Você não tem permissão para editar este supervisor.", "danger")
        return redirect(url_for("lista_supervisores"))

    form = UsuarioForm(usuario_id=id, obj=supervisor)
    gerentes_disponiveis = _prepare_supervisor_form(form, supervisor)

    if form.validate_on_submit():
        if not request.form.get("gerente_id"):
            flash("É obrigatório vincular o supervisor a um Gerente ou Administrador!", "danger")
        else:
            try:
                _save_supervisor_from_form(form, supervisor)
                flash(f"Supervisor {supervisor.nome} atualizado!", "success")
                return redirect(url_for("lista_supervisores"))
            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao atualizar supervisor: {str(e)}", "danger")

    return render_template(
        "supervisores/form.html",
        form=form,
        supervisor=supervisor,
        gerentes_disponiveis=gerentes_disponiveis,
        titulo="Editar Supervisor",
    )

@app.route("/supervisores/<int:id>/deletar", methods=["POST"])
@login_required
def deletar_supervisor(id):
    """Deletar supervisor (soft delete)"""
    supervisor = Usuario.query.get_or_404(id)

    # Verificar permissão
    if not current_user.is_super_admin:
        if supervisor.empresa_id != current_user.empresa_id:
            msg = "Você não tem permissão " "para deletar este supervisor."
            flash(msg, "danger")
            return redirect(url_for("lista_supervisores"))

    # Verificar se é realmente um supervisor
    if supervisor.cargo != "supervisor":
        flash("Este usuário não é um supervisor.", "warning")
        return redirect(url_for("lista_supervisores"))

    # Soft delete
    supervisor.ativo = False
    supervisor.bloqueado = True
    supervisor.motivo_bloqueio = "Supervisor desativado"

    db.session.commit()

    msg = f"Supervisor {supervisor.nome} desativado com sucesso!"
    flash(msg, "success")
    return redirect(url_for("lista_supervisores"))

@app.route("/supervisores/<int:id>/resetar-senha", methods=["POST"])
@login_required
def resetar_senha_supervisor(id):
    """Resetar senha do supervisor para senha padrão"""
    supervisor = Usuario.query.get_or_404(id)

    # Verificar permissão
    if not current_user.is_super_admin:
        if supervisor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para resetar a senha deste supervisor.",
                "danger",
            )
            return redirect(url_for("lista_supervisores"))

        # Apenas admin e gerente podem resetar senhas
        if current_user.cargo not in ["admin", "gerente"]:
            flash(
                "Apenas administradores e gerentes podem resetar senhas.",
                "danger",
            )
            return redirect(url_for("lista_supervisores"))

    # Verificar se é realmente um supervisor
    if supervisor.cargo != "supervisor":
        flash("Este usuário não é um supervisor.", "warning")
        return redirect(url_for("lista_supervisores"))

    # Obter nova senha do formulário ou usar padrão
    nova_senha = request.form.get("nova_senha", "senha123")

    # Resetar senha
    supervisor.set_senha(nova_senha)
    db.session.commit()

    flash(
        f"Senha do supervisor {supervisor.nome} resetada com sucesso! Nova senha: {nova_senha}",
        "success",
    )
    return redirect(url_for("lista_supervisores"))

@app.route("/supervisores/<int:id>/definir-senha", methods=["GET", "POST"])
@login_required
def definir_senha_supervisor(id):
    """Definir/alterar senha do supervisor"""
    supervisor = Usuario.query.get_or_404(id)

    # Verificar permissão
    if not current_user.is_super_admin:
        if supervisor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para alterar a senha deste supervisor.",
                "danger",
            )
            return redirect(url_for("lista_supervisores"))

        # Apenas admin e gerente podem alterar senhas
        if current_user.cargo not in ["admin", "gerente"]:
            flash(
                "Apenas administradores e gerentes podem alterar senhas.",
                "danger",
            )
            return redirect(url_for("lista_supervisores"))

    # Verificar se é realmente um supervisor
    if supervisor.cargo != "supervisor":
        flash("Este usuário não é um supervisor.", "warning")
        return redirect(url_for("lista_supervisores"))

    if request.method == "POST":
        nova_senha = request.form.get("nova_senha", "").strip()
        confirmar_senha = request.form.get("confirmar_senha", "").strip()

        # Validações
        if not nova_senha:
            flash("Por favor, informe a nova senha.", "warning")
            return render_template(
                "supervisores/definir_senha.html", supervisor=supervisor
            )

        if len(nova_senha) < 6:
            flash("A senha deve ter no mínimo 6 caracteres.", "warning")
            return render_template(
                "supervisores/definir_senha.html", supervisor=supervisor
            )

        if nova_senha != confirmar_senha:
            flash("As senhas não conferem!", "danger")
            return render_template(
                "supervisores/definir_senha.html", supervisor=supervisor
            )

        # Definir nova senha
        supervisor.set_senha(nova_senha)
        db.session.commit()

        flash(
            f"Senha do supervisor {supervisor.nome} alterada com sucesso!",
            "success",
        )
        return redirect(url_for("lista_supervisores"))

    return render_template(
        "supervisores/definir_senha.html", supervisor=supervisor
    )

# ===== FUNÇÕES AUXILIARES PARA SUPERVISORES =====

def _prepare_supervisor_form(form, supervisor=None):
    """Prepara o formulário de supervisor.

    - Restringe o cargo para "supervisor".
    - Popula a lista de empresas conforme o escopo do usuário.
    - Retorna a lista de gerentes/administradores disponíveis
      para vincular ao supervisor.
    """

    # Garantir que o cargo esteja fixo como supervisor
    form.cargo.choices = [("supervisor", "Supervisor de Vendas")]
    if not form.cargo.data:
        form.cargo.data = "supervisor"

    # Empresas disponíveis
    if current_user.is_super_admin:
        empresas = Empresa.query.filter_by(ativo=True).order_by(Empresa.nome).all()
    else:
        empresas = (
            Empresa.query.filter_by(id=current_user.empresa_id, ativo=True)
            .order_by(Empresa.nome)
            .all()
        )

    form.empresa_id.choices = [(e.id, e.nome) for e in empresas]

    # Empresa selecionada
    if supervisor and supervisor.empresa_id:
        form.empresa_id.data = supervisor.empresa_id
    elif not supervisor and not current_user.is_super_admin and current_user.empresa_id:
        form.empresa_id.data = current_user.empresa_id

    # Valores padrão de status para novo cadastro
    if not supervisor:
        if form.ativo.data is None:
            form.ativo.data = 1
        if form.bloqueado.data is None:
            form.bloqueado.data = 0

    # Gerentes/administradores disponíveis para vincular
    if current_user.is_super_admin:
        gerentes_query = Usuario.query.filter(
            Usuario.ativo.is_(True),
            Usuario.cargo.in_(["admin", "gerente"]),
        )
    else:
        gerentes_query = Usuario.query.filter(
            Usuario.empresa_id == current_user.empresa_id,
            Usuario.ativo.is_(True),
            Usuario.cargo.in_(["admin", "gerente"]),
        )

    gerentes_disponiveis = gerentes_query.order_by(Usuario.nome).all()
    return gerentes_disponiveis


def _save_supervisor_from_form(form, supervisor=None):
    """Cria ou atualiza um supervisor a partir do formulário.

    Retorna a instância de Usuario correspondente ao supervisor.
    """

    empresa_id = form.empresa_id.data
    empresa = Empresa.query.get(empresa_id)
    if not empresa:
        raise ValueError("Empresa não encontrada.")

    # Garantir que usuário comum só gerencie supervisores da própria empresa
    if not current_user.is_super_admin and empresa_id != current_user.empresa_id:
        raise PermissionError("Você só pode gerenciar supervisores da sua empresa.")

    # Gerente responsável (campo vem do formulário HTML)
    gerente_id_raw = request.form.get("gerente_id")
    if not gerente_id_raw:
        raise ValueError("É obrigatório vincular o supervisor a um Gerente ou Administrador.")

    try:
        gerente_id = int(gerente_id_raw)
    except ValueError:
        raise ValueError("Gerente/Administrador selecionado é inválido.")

    gerente = Usuario.query.get(gerente_id)
    if not gerente or gerente.cargo not in ["admin", "gerente"]:
        raise ValueError("Gerente/Administrador responsável não encontrado ou inválido.")

    if (
        not current_user.is_super_admin
        and gerente.empresa_id != current_user.empresa_id
    ):
        raise PermissionError(
            "Você só pode vincular supervisores a gerentes da sua empresa."
        )

    if supervisor is None:
        # Criação de novo supervisor
        supervisor = Usuario(
            nome=form.nome.data,
            email=form.email.data,
            cargo="supervisor",
            empresa_id=empresa_id,
            ativo=bool(form.ativo.data),
            bloqueado=bool(form.bloqueado.data),
            motivo_bloqueio=form.motivo_bloqueio.data or None,
            gerente_id=gerente.id,
        )
        supervisor.set_senha("senha123")
        db.session.add(supervisor)
    else:
        # Atualização de supervisor existente
        supervisor.nome = form.nome.data
        supervisor.email = form.email.data
        supervisor.empresa_id = empresa_id
        supervisor.ativo = bool(form.ativo.data)
        supervisor.bloqueado = bool(form.bloqueado.data)
        supervisor.motivo_bloqueio = form.motivo_bloqueio.data or None
        supervisor.gerente_id = gerente.id

    db.session.commit()
    return supervisor

# ===== CONTINUAÇÃO - ROTAS DE SUPER ADMIN - USUÁRIOS =====

@app.route("/super-admin/usuarios/criar", methods=["GET", "POST"])
@super_admin_required
def super_admin_criar_usuario():
    """Criar novo usuário"""

    form = UsuarioForm()

    # Preencher choices de empresas
    empresas = Empresa.query.filter_by(ativo=True).all()
    form.empresa_id.choices = [(e.id, e.nome) for e in empresas]

    if form.validate_on_submit():
        usuario = Usuario(
            nome=form.nome.data,
            email=form.email.data,
            cargo=form.cargo.data,
            empresa_id=form.empresa_id.data,
            ativo=bool(form.ativo.data),
            bloqueado=bool(form.bloqueado.data),
        )
        if form.bloqueado.data:
            usuario.motivo_bloqueio = form.motivo_bloqueio.data
        else:
            usuario.motivo_bloqueio = None

        # Senha padrão
        usuario.set_senha("senha123")

        # Permissões padrão específicas para cargo auxiliar
        if usuario.cargo == "auxiliar":
            usuario.pode_enviar_mensagens = True
            usuario.pode_acessar_os = True
            usuario.pode_criar_os = False
            usuario.pode_aprovar_os = False
            usuario.pode_atualizar_os = False
            usuario.pode_cancelar_os = False

            usuario.pode_acessar_clientes = False
            usuario.pode_criar_clientes = False
            usuario.pode_editar_clientes = False
            usuario.pode_excluir_clientes = False
            usuario.pode_importar_clientes = False

            usuario.pode_acessar_estoque = False
            usuario.pode_gerenciar_produtos = False
            usuario.pode_movimentar_estoque = False
            usuario.pode_ver_custos = False
            usuario.pode_ajustar_estoque = False

        db.session.add(usuario)
        db.session.commit()

        msg = f"Usuário {usuario.nome} criado! Senha padrão: senha123"
        flash(msg, "success")
        return redirect(url_for("super_admin_usuarios"))

    return render_template(
        "super_admin/usuario_form.html", form=form, titulo="Criar Novo Usuário"
    )

@app.route("/super-admin/usuarios/<int:id>/editar", methods=["GET", "POST"])
@super_admin_required
def super_admin_editar_usuario(id):
    """Editar usuário existente"""

    try:
        usuario = Usuario.query.get_or_404(id)
        form = UsuarioForm(usuario_id=id, obj=usuario)

        # Preencher choices de empresas
        empresas = Empresa.query.filter_by(ativo=True).all()
        form.empresa_id.choices = [(e.id, e.nome) for e in empresas]

        if request.method == "GET":
            form.ativo.data = 1 if usuario.ativo else 0
            form.bloqueado.data = 1 if usuario.bloqueado else 0

        if form.validate_on_submit():
            try:
                usuario.nome = form.nome.data
                usuario.email = form.email.data
                usuario.cargo = form.cargo.data
                usuario.empresa_id = form.empresa_id.data
                usuario.ativo = bool(form.ativo.data)
                usuario.bloqueado = bool(form.bloqueado.data)
                if form.bloqueado.data:
                    usuario.motivo_bloqueio = form.motivo_bloqueio.data
                else:
                    usuario.motivo_bloqueio = None

                # Vincular supervisor ao gerente/administrador
                gerente_id = request.form.get("gerente_id")
                if gerente_id and gerente_id != "":
                    usuario.gerente_id = (
                        int(gerente_id) if gerente_id != "0" else None
                    )
                else:
                    usuario.gerente_id = None

                # Atualizar permissões
                usuario.pode_ver_dashboard = (
                    "pode_ver_dashboard" in request.form
                )
                usuario.pode_gerenciar_vendedores = (
                    "pode_gerenciar_vendedores" in request.form
                )
                usuario.pode_gerenciar_metas = (
                    "pode_gerenciar_metas" in request.form
                )
                usuario.pode_gerenciar_equipes = (
                    "pode_gerenciar_equipes" in request.form
                )
                usuario.pode_gerenciar_comissoes = (
                    "pode_gerenciar_comissoes" in request.form
                )
                usuario.pode_enviar_mensagens = (
                    "pode_enviar_mensagens" in request.form
                )
                usuario.pode_exportar_dados = (
                    "pode_exportar_dados" in request.form
                )
                usuario.pode_ver_todas_metas = (
                    "pode_ver_todas_metas" in request.form
                )
                usuario.pode_aprovar_comissoes = (
                    "pode_aprovar_comissoes" in request.form
                )

                db.session.commit()
                flash(f"Usuário {usuario.nome} atualizado!", "success")
                return redirect(url_for("super_admin_usuarios"))
            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao atualizar usuário: {str(e)}", "danger")
                return render_template(
                    "super_admin/usuario_form.html",
                    form=form,
                    usuario=usuario,
                    titulo="Editar Usuário",
                )

        # Buscar gerentes e administradores disponíveis
        gerentes_disponiveis = Usuario.query.filter(
            Usuario.cargo.in_(["gerente", "admin"]),
            Usuario.id != usuario.id,
            Usuario.ativo,
        ).all()

        return render_template(
            "super_admin/usuario_form.html",
            form=form,
            usuario=usuario,
            gerentes_disponiveis=gerentes_disponiveis,
            titulo="Editar Usuário",
        )
    except Exception as e:
        flash(f"Erro ao carregar usuário: {str(e)}", "danger")
        return redirect(url_for("super_admin_usuarios"))

@app.route("/super-admin/usuarios/<int:id>/bloquear", methods=["POST"])
@super_admin_required
def super_admin_bloquear_usuario(id):
    """Bloquear/desbloquear usuário"""
    usuario = Usuario.query.get_or_404(id)

    motivo = request.form.get("motivo", "")

    if usuario.bloqueado:
        usuario.bloqueado = False
        usuario.motivo_bloqueio = None
        msg = f"Usuário {usuario.nome} desbloqueado!"
        flash(msg, "success")
    else:
        usuario.bloqueado = True
        usuario.motivo_bloqueio = motivo
        flash(f"Usuário {usuario.nome} bloqueado!", "warning")

    db.session.commit()
    return redirect(url_for("super_admin_usuarios"))

@app.route("/super-admin/usuarios/<int:id>/deletar", methods=["POST"])
@super_admin_required
def super_admin_deletar_usuario(id):
    """Deletar usuário"""
    usuario = Usuario.query.get_or_404(id)

    if usuario.is_super_admin:
        msg = "Não é possível deletar um super administrador!"
        flash(msg, "danger")
        return redirect(url_for("super_admin_usuarios"))

    nome = usuario.nome
    db.session.delete(usuario)
    db.session.commit()

    flash(f"Usuário {nome} deletado!", "success")
    return redirect(url_for("super_admin_usuarios"))

# ===== ROTA DE MIGRAÇÃO DE FAIXAS DE COMISSÃO =====

@app.route("/migrar-faixas-comissao-agora")
def migrar_faixas_comissao_agora():
    """
    Rota temporária para criar tabela FaixaComissao e popular dados
    Acesse via: https://vendacerta.up.railway.app/migrar-faixas-comissao-agora
    """
    try:
        from sqlalchemy import text, inspect

        # Verifica se a tabela já existe
        inspector = inspect(db.engine)
        existing_tables = inspector.get_table_names()

        if "faixa_comissao" in existing_tables:
            # Verifica se tem dados
            count = db.session.execute(
                text("SELECT COUNT(*) FROM faixa_comissao")
            ).scalar()

            if count > 0:
                return f"""
                <html>
                <head>
                    <title>Faixas Já Criadas</title>
                    <meta name="viewport" content="width=device-width, initial-scale=1">
                    <style>
                        body {{
                            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            display: flex;
                            justify-content: center;
                            align-items: center;
                            min-height: 100vh;
                            margin: 0;
                            padding: 20px;
                        }}
                        .container {{
                            background: white;
                            padding: 40px;
                            border-radius: 15px;
                            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                            max-width: 600px;
                            width: 100%;
                        }}
                        h1 {{ color: #667eea; margin-bottom: 20px; text-align: center; }}
                        .success {{ color: #28a745; font-weight: bold; text-align: center; font-size: 1.2em; }}
                        p {{ color: #666; margin: 15px 0; }}
                        .btn {{
                            display: block;
                            margin: 20px auto 0;
                            padding: 15px 40px;
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: white;
                            text-decoration: none;
                            border-radius: 8px;
                            font-weight: 600;
                            text-align: center;
                            transition: all 0.3s;
                        }}
                        .btn:hover {{
                            transform: translateY(-2px);
                            box-shadow: 0 5px 15px rgba(102,126,234,0.4);
                        }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <h1>Sistema Já Configurado!</h1>
                        <p class="success">A tabela 'faixa_comissao' já existe com {count} registros.</p>
                        <p style="text-align: center;">O sistema está pronto para uso!</p>
                        <a href="/configuracoes/comissoes" class="btn">Ver Faixas de Comissão</a>
                    </div>
                </body>
                </html>
                """

        # Cria a tabela
        db.create_all()

        # Popula com faixas padrão
        faixas_padrao = [
            {
                "ordem": 1,
                "alcance_min": 0.0,
                "alcance_max": 50.0,
                "taxa_comissao": 0.01,
                "cor": "danger",
                "ativa": True,
            },
            {
                "ordem": 2,
                "alcance_min": 51.0,
                "alcance_max": 75.0,
                "taxa_comissao": 0.02,
                "cor": "warning",
                "ativa": True,
            },
            {
                "ordem": 3,
                "alcance_min": 76.0,
                "alcance_max": 100.0,
                "taxa_comissao": 0.03,
                "cor": "info",
                "ativa": True,
            },
            {
                "ordem": 4,
                "alcance_min": 101.0,
                "alcance_max": 125.0,
                "taxa_comissao": 0.04,
                "cor": "primary",
                "ativa": True,
            },
            {
                "ordem": 5,
                "alcance_min": 125.1,
                "alcance_max": 10000.0,
                "taxa_comissao": 0.05,
                "cor": "success",
                "ativa": True,
            },
        ]

        for dados in faixas_padrao:
            faixa = FaixaComissao(empresa_id=None, **dados)
            db.session.add(faixa)

        db.session.commit()

        return """
        <html>
        <head>
            <title>Migração Concluída</title>
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                    padding: 20px;
                }
                .container {
                    background: white;
                    padding: 40px;
                    border-radius: 15px;
                    box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                    max-width: 700px;
                    width: 100%;
                }
                h1 { color: #28a745; margin-bottom: 20px; text-align: center; }
                .success { color: #28a745; font-weight: bold; text-align: center; font-size: 1.3em; margin: 20px 0; }
                .faixa {
                    padding: 15px;
                    margin: 10px 0;
                    border-radius: 8px;
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                }
                .faixa.danger { background: #f8d7da; color: #721c24; }
                .faixa.warning { background: #fff3cd; color: #856404; }
                .faixa.info { background: #d1ecf1; color: #0c5460; }
                .faixa.primary { background: #cfe2ff; color: #084298; }
                .faixa.success { background: #d1e7dd; color: #0f5132; }
                .btn {
                    display: block;
                    margin: 30px auto 0;
                    padding: 15px 40px;
                    background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
                    color: white;
                    text-decoration: none;
                    border-radius: 8px;
                    font-weight: 600;
                    text-align: center;
                    transition: all 0.3s;
                }
                .btn:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(40,167,69,0.4);
                }
                p { color: #666; text-align: center; margin: 10px 0; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Migração Concluída!</h1>
                <p class="success">Tabela 'faixa_comissao' criada e populada com sucesso!</p>

                <h3 style="color: #666; text-align: center; margin: 30px 0 20px;">Faixas Criadas:</h3>

                <div class="faixa danger">
                    <span><strong>0% - 50%</strong></span>
                    <span>Taxa: 1%</span>
                </div>
                <div class="faixa warning">
                    <span><strong>51% - 75%</strong></span>
                    <span>Taxa: 2%</span>
                </div>
                <div class="faixa info">
                    <span><strong>76% - 100%</strong></span>
                    <span>Taxa: 3%</span>
                </div>
                <div class="faixa primary">
                    <span><strong>101% - 125%</strong></span>
                    <span>Taxa: 4%</span>
                </div>
                <div class="faixa success">
                    <span><strong>Acima de 125%</strong></span>
                    <span>Taxa: 5%</span>
                </div>

                <p style="margin-top: 30px;">Agora você pode criar e editar faixas de comissão!</p>
                <a href="/configuracoes/comissoes" class="btn">Ir para Configurações</a>
            </div>
        </body>
        </html>
        """

    except Exception as e:
        db.session.rollback()
        return f"""
        <html>
        <head>
            <title>Erro na Migração</title>
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #dc3545 0%, #c82333 100%);
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                    padding: 20px;
                }}
                .container {{
                    background: white;
                    padding: 40px;
                    border-radius: 15px;
                    box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                    max-width: 600px;
                    width: 100%;
                }}
                h1 {{ color: #dc3545; margin-bottom: 20px; text-align: center; }}
                .error {{ color: #dc3545; font-weight: bold; padding: 20px; background: #f8d7da; border-radius: 8px; margin: 20px 0; }}
                pre {{ background: #f1f1f1; padding: 15px; border-radius: 5px; overflow-x: auto; font-size: 12px; }}
                .btn {{
                    display: block;
                    margin: 20px auto 0;
                    padding: 15px 40px;
                    background: #dc3545;
                    color: white;
                    text-decoration: none;
                    border-radius: 8px;
                    font-weight: 600;
                    text-align: center;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Erro na Migração</h1>
                <div class="error">
                    Ocorreu um erro ao criar a tabela faixa_comissao.
                </div>
                <pre>{str(e)}</pre>
                <p style="text-align: center; color: #666;">Entre em contato com o suporte técnico.</p>
                <a href="/dashboard" class="btn">Voltar ao Dashboard</a>
            </div>
        </body>
        </html>
        """

# ===== ROTA DE INICIALIZAÇÃO (Setup inicial do sistema) =====

@app.route("/setup-inicial-sistema")
def setup_inicial_sistema():
    """
    Cria estrutura inicial do sistema (empresa padrão e super admin)
    Esta rota pode ser acessada uma única vez para configurar o sistema
    """
    try:
        # Verificar se já existe super admin
        # (sistema já configurado)
        super_admin = Usuario.query.filter_by(
            email="superadmin@suameta.com"
        ).first()
        if super_admin:
            return """
            <html>
            <head>
                <title>Sistema Já Configurado</title>
                <style>
                    body {
                        font-family: 'Segoe UI', Tahoma, Geneva,
                            Verdana, sans-serif;
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        height: 100vh;
                        margin: 0;
                    }
                    .container {
                        background: white;
                        padding: 40px;
                        border-radius: 15px;
                        box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                        text-align: center;
                        max-width: 500px;
                    }
                    h1 { color: #667eea; margin-bottom: 20px; }
                    p { color: #666; margin: 15px 0; }
                    .success { color: #28a745; font-weight: bold; }
                    a {
                        display: inline-block;
                        margin-top: 20px;
                        padding: 12px 30px;
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        color: white;
                        text-decoration: none;
                        border-radius: 8px;
                        font-weight: 600;
                    }
                    a:hover {
                        transform: translateY(-2px);
                        box-shadow: 0 5px 15px rgba(102,126,234,0.4);
                    }
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>Sistema Já Configurado!</h1>
                    <p class="success">
                        O super administrador já existe no sistema.
                    </p>
                    <p>Faça login com as credenciais:</p>
                    <p><strong>Email:</strong> superadmin@suameta.com</p>
                    <p><strong>Senha:</strong> 18042016</p>
                    <a href="/login">Ir para Login</a>
                </div>
            </body>
            </html>
            """

        # Criar empresa padrão se não existir
        empresa = Empresa.query.filter_by(nome="Empresa Padrão").first()
        if not empresa:
            empresa = Empresa(
                nome="Empresa Padrão",
                cnpj="00.000.000/0000-00",
                email="contato@empresapadrao.com",
                telefone="(00) 0000-0000",
                ativo=True,
                bloqueado=False,
            )
            db.session.add(empresa)
            db.session.commit()

        # Criar super admin
        super_admin = Usuario(
            nome="Super Administrador",
            email="admin@suameta.com.br",
            cargo="admin",
            is_super_admin=True,
            empresa_id=None,
            ativo=True,
            bloqueado=False,
        )
        super_admin.set_senha("Admin@2025!")
        db.session.add(super_admin)
        db.session.commit()

        # Criar admin da empresa
        admin = Usuario.query.filter_by(email="gerente@suameta.com.br").first()
        if not admin:
            admin = Usuario(
                nome="Gerente Empresa",
                email="gerente@suameta.com.br",
                cargo="admin",
                is_super_admin=False,
                empresa_id=empresa.id,
                ativo=True,
                bloqueado=False,
            )
            admin.set_senha("Gerente@2025!")
            db.session.add(admin)
            db.session.commit()

        return """
        <html>
        <head>
            <title>Sistema Configurado</title>
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                    padding: 20px;
                }
                .container {
                    background: white;
                    padding: 40px;
                    border-radius: 15px;
                    box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                    text-align: center;
                    max-width: 600px;
                    width: 100%;
                }
                h1 { color: #667eea; margin-bottom: 20px; font-size: 2rem; }
                .success {
                    background: #d4edda;
                    border: 1px solid #c3e6cb;
                    color: #155724;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 20px 0;
                }
                .credentials {
                    background: #f8f9fa;
                    padding: 20px;
                    border-radius: 8px;
                    margin: 20px 0;
                    text-align: left;
                }
                .credentials h3 {
                    color: #667eea;
                    margin-top: 0;
                }
                .credentials p {
                    margin: 10px 0;
                    font-family: 'Courier New', monospace;
                }
                .credentials strong {
                    display: inline-block;
                    width: 100px;
                }
                a {
                    display: inline-block;
                    margin-top: 20px;
                    padding: 12px 30px;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    text-decoration: none;
                    border-radius: 8px;
                    font-weight: 600;
                    transition: all 0.3s ease;
                }
                a:hover {
                    transform: translateY(-2px);
                    box-shadow: 0 5px 15px rgba(102,126,234,0.4);
                }
                .warning {
                    background: #fff3cd;
                    border: 1px solid #ffc107;
                    color: #856404;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 20px 0;
                }
                @media (max-width: 768px) {
                    .container { padding: 20px; }
                    h1 { font-size: 1.5rem; }
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Sistema Configurado com Sucesso!</h1>

                <div class="success">
                    <strong>Configuração Completa</strong><br>
                    O sistema foi inicializado e está pronto para uso!
                </div>

                <div class="credentials">
                    <h3>👑 Super Administrador</h3>
                    <p><strong>Email:</strong> admin@suameta.com.br</p>
                    <p><strong>Senha:</strong> Admin@2025!</p>
                </div>

                <div class="credentials">
                    <h3>🔑 Gerente da Empresa</h3>
                    <p><strong>Email:</strong> gerente@suameta.com.br</p>
                    <p><strong>Senha:</strong> Gerente@2025!</p>
                </div>

                <div class="warning">
                    ⚠️ <strong>Importante:</strong> Altere as senhas após o primeiro login!
                </div>

                <a href="/login">Acessar o Sistema</a>
            </div>
        </body>
        </html>
        """

    except Exception as e:
        return f"""
        <html>
        <head>
            <title>Erro na Configuração</title>
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    min-height: 100vh;
                    margin: 0;
                    padding: 20px;
                }}
                .container {{
                    background: white;
                    padding: 40px;
                    border-radius: 15px;
                    box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                    text-align: center;
                    max-width: 600px;
                    width: 100%;
                }}
                h1 {{ color: #dc3545; margin-bottom: 20px; }}
                .error {{
                    background: #f8d7da;
                    border: 1px solid #f5c6cb;
                    color: #721c24;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 20px 0;
                }}
                code {{
                    display: block;
                    background: #f8f9fa;
                    padding: 10px;
                    border-radius: 5px;
                    margin-top: 10px;
                    text-align: left;
                    overflow-x: auto;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Erro na Configuração</h1>
                <div class="error">
                    <strong>Erro:</strong><br>
                    <code>{str(e)}</code>
                </div>
                <p>Entre em contato com o suporte.</p>
            </div>
        </body>
        </html>
        """

# ===== ROTAS DE BACKUP E RESTAURAÇÃO =====

@app.route("/super-admin/backups")
@super_admin_required
def super_admin_backups():
    """Gerenciar backups do sistema"""
    # Verificar tipo de banco de dados
    uri = app.config.get("SQLALCHEMY_DATABASE_URI", "")
    is_postgresql = "postgresql" in uri

    # Para PostgreSQL, backups devem ser gerenciados pelo Railway/Render
    if is_postgresql:
        db_info = {
            "tipo": "PostgreSQL",
            "mensagem": "Backups automáticos gerenciados pelo Railway",
            "tamanho": 0,
            "ultima_modificacao": datetime.now(),
        }
        return render_template(
            "super_admin/backups.html",
            backups=[],
            db_info=db_info,
            is_postgresql=True,
        )

    # Para SQLite (ambiente local)
    # Criar pasta de backups se não existir
    backup_dir = os.path.join(app.instance_path, "backups")
    os.makedirs(backup_dir, exist_ok=True)

    # Listar backups existentes
    backups = []
    if os.path.exists(backup_dir):
        for filename in os.listdir(backup_dir):
            if filename.endswith(".db"):
                filepath = os.path.join(backup_dir, filename)
                try:
                    stat = os.stat(filepath)
                    # Usar timestamp real do arquivo
                    data_modificacao = datetime.fromtimestamp(stat.st_mtime)
                    backups.append(
                        {
                            "nome": filename,
                            "tamanho": round(stat.st_size / 1024, 2),  # KB
                            # MB
                            "tamanho_mb": round(
                                stat.st_size / (1024 * 1024), 2
                            ),
                            "data": data_modificacao,
                            "data_formatada": data_modificacao.strftime(
                                "%d/%m/%Y às %H:%M:%S"
                            ),
                            "path": filepath,
                            "tipo": (
                                "auto"
                                if "auto_backup" in filename
                                else "manual"
                            ),
                        }
                    )
                except Exception as e:
                    app.logger.error(f"Erro ao ler backup {filename}: {e}")

    # Ordenar por data (mais recente primeiro)
    backups.sort(key=lambda x: x["data"], reverse=True)

    # Informações do banco atual
    db_path = uri.replace("sqlite:///", "")
    db_info = None
    if os.path.exists(db_path):
        stat = os.stat(db_path)
        db_info = {
            "tipo": "SQLite",
            "tamanho": round(stat.st_size / 1024, 2),
            "ultima_modificacao": datetime.fromtimestamp(stat.st_mtime),
        }

    return render_template(
        "super_admin/backups.html",
        backups=backups,
        db_info=db_info,
        is_postgresql=False,
    )

@app.route("/super-admin/backups/criar", methods=["POST"])
@super_admin_required
def criar_backup():
    """Criar novo backup do banco de dados"""
    try:
        # Verificar tipo de banco
        uri = app.config.get("SQLALCHEMY_DATABASE_URI", "")

        if "postgresql" in uri:
            flash(
                "Backups do PostgreSQL são gerenciados automaticamente pelo Railway. Acesse o painel do Railway para gerenciar backups.",
                "info",
            )
            return redirect(url_for("super_admin_backups"))

        # Para SQLite
        # Criar pasta de backups
        backup_dir = os.path.join(app.instance_path, "backups")
        os.makedirs(backup_dir, exist_ok=True)

        # Nome do backup com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"backup_{timestamp}.db"
        backup_path = os.path.join(backup_dir, backup_name)

        # Copiar banco de dados atual
        db_path = uri.replace("sqlite:///", "")

        if not os.path.exists(db_path):
            flash("Banco de dados não encontrado!", "danger")
            return redirect(url_for("super_admin_backups"))

        shutil.copy2(db_path, backup_path)

        # Sincronizar com nuvem (se módulo opcional estiver disponível)
        try:
            from backup_nuvem import sincronizar_backup_nuvem

            app.logger.info(
                "[PROC] Iniciando sincronizacao com nuvem (backup manual)..."
            )
            sincronizar_backup_nuvem()
            app.logger.info(
                "[OK] Sincronizacao com nuvem concluida (backup manual)"
            )
        except ImportError:
            app.logger.warning(
                "[AVISO] Modulo backup_nuvem nao encontrado para backup manual"
            )
        except Exception as e:
            app.logger.error(
                f"Erro na sincronizacao de backup manual: {str(e)}"
            )

        flash(f"Backup criado com sucesso: {backup_name}", "success")
    except Exception as e:
        flash(f"Erro ao criar backup: {str(e)}", "danger")

    return redirect(url_for("super_admin_backups"))

@app.route("/super-admin/backups/download/<nome>")
@super_admin_required
def download_backup(nome):
    """Fazer download de um backup"""
    try:
        backup_dir = os.path.join(app.instance_path, "backups")
        backup_path = os.path.join(backup_dir, secure_filename(nome))

        if not os.path.exists(backup_path):
            flash("Backup não encontrado!", "danger")
            return redirect(url_for("super_admin_backups"))

        return send_file(backup_path, as_attachment=True, download_name=nome)
    except Exception as e:
        flash(f"Erro ao fazer download: {str(e)}", "danger")
        return redirect(url_for("super_admin_backups"))

@app.route("/super-admin/backups/restaurar/<nome>", methods=["POST"])
@super_admin_required
def restaurar_backup(nome):
    """Restaurar backup do banco de dados"""
    try:
        backup_dir = os.path.join(app.instance_path, "backups")
        backup_path = os.path.join(backup_dir, secure_filename(nome))

        if not os.path.exists(backup_path):
            flash("Backup não encontrado!", "danger")
            return redirect(url_for("super_admin_backups"))

        # Criar backup do banco atual antes de restaurar
        uri = app.config.get("SQLALCHEMY_DATABASE_URI", "")
        db_path = uri.replace("sqlite:///", "")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safety_backup = os.path.join(backup_dir, f"pre_restore_{timestamp}.db")

        if os.path.exists(db_path):
            shutil.copy2(db_path, safety_backup)

        # Restaurar backup
        shutil.copy2(backup_path, db_path)

        flash(f"Backup {nome} restaurado com sucesso!", "success")
        msg_info = f"Backup de segurança criado: pre_restore_{timestamp}.db"
        flash(msg_info, "info")
    except Exception as e:
        flash(f"Erro ao restaurar backup: {str(e)}", "danger")

    return redirect(url_for("super_admin_backups"))

@app.route("/super-admin/backups/deletar/<nome>", methods=["POST"])
@super_admin_required
def deletar_backup(nome):
    """Deletar um backup"""
    try:
        backup_dir = os.path.join(app.instance_path, "backups")
        backup_path = os.path.join(backup_dir, secure_filename(nome))

        if not os.path.exists(backup_path):
            flash("Backup não encontrado!", "danger")
            return redirect(url_for("super_admin_backups"))

        os.remove(backup_path)
        flash(f"Backup {nome} deletado!", "success")
    except Exception as e:
        flash(f"Erro ao deletar backup: {str(e)}", "danger")

    return redirect(url_for("super_admin_backups"))

@app.route("/super-admin/backups/upload", methods=["POST"])
@super_admin_required
def upload_backup():
    """Upload de backup externo"""
    try:
        if "backup_file" not in request.files:
            flash("Nenhum arquivo selecionado!", "danger")
            return redirect(url_for("super_admin_backups"))

        file = request.files["backup_file"]

        if file.filename == "":
            flash("Nenhum arquivo selecionado!", "danger")
            return redirect(url_for("super_admin_backups"))

        if not file.filename.endswith(".db"):
            flash("Apenas arquivos .db são permitidos!", "danger")
            return redirect(url_for("super_admin_backups"))

        # Salvar arquivo
        backup_dir = os.path.join(app.instance_path, "backups")
        os.makedirs(backup_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"upload_{timestamp}.db"
        filepath = os.path.join(backup_dir, filename)

        file.save(filepath)

        flash(f"Backup enviado com sucesso: {filename}", "success")
    except Exception as e:
        flash(f"Erro ao enviar backup: {str(e)}", "danger")

    return redirect(url_for("super_admin_backups"))

# ===== ROTA PRINCIPAL - DASHBOARD =====

@app.route("/")
def index():
    """Rota inicial - redireciona para login ou dashboard"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route("/dashboard")
@login_required
def dashboard():
    """Dashboard principal com métricas e ranking"""
    
    # CACHE: Usar cache key baseado em usuário, mês e ano (5 minutos)
    hoje = datetime.now()
    mes_atual = request.args.get("mes", hoje.month, type=int)
    ano_atual = request.args.get("ano", hoje.year, type=int)
    
    cache_key = f"dashboard_{current_user.id}_{mes_atual}_{ano_atual}"
    
    # Tentar obter do cache se disponível
    if cache:
        cached_result = cache.get(cache_key)
        if cached_result:
            return cached_result
    
    try:
        # Obter o mês e ano atuais
        hoje = datetime.now()
        mes_atual = request.args.get("mes", hoje.month, type=int)
        ano_atual = request.args.get("ano", hoje.year, type=int)

        # Buscar todas as metas do período (filtrando por empresa e cargo)
        # OTIMIZAÇÃO: Usar joinedload para evitar N+1 queries
        from sqlalchemy.orm import joinedload

        query = (
            Meta.query.options(
                joinedload(Meta.vendedor).joinedload(Vendedor.equipe_obj),
                joinedload(Meta.vendedor).joinedload(Vendedor.supervisor_obj),
            )
            .filter_by(mes=mes_atual, ano=ano_atual)
            .join(Vendedor)
        )

        if current_user.is_super_admin:
            # Super admin vê tudo
            pass
        elif current_user.cargo == "supervisor":
            # Supervisor vê apenas metas dos seus vendedores
            query = query.filter(Vendedor.supervisor_id == current_user.id)
        elif current_user.cargo == "vendedor" and current_user.vendedor_id:
            # Vendedor vê apenas suas metas e dos colegas da mesma equipe
            vendedor_atual = Vendedor.query.get(current_user.vendedor_id)
            if vendedor_atual and vendedor_atual.equipe_id:
                query = query.filter(
                    Vendedor.equipe_id == vendedor_atual.equipe_id
                )
            else:
                # Se não tem equipe, vê apenas suas próprias metas
                query = query.filter(Vendedor.id == current_user.vendedor_id)
        else:
            # Outros usuários veem apenas da sua empresa
            query = query.filter(
                Vendedor.empresa_id == current_user.empresa_id
            )

        metas = query.all()

        # Processar dados para o dashboard
        vendedores_data = []
        for meta in metas:
            # Verificar se o vendedor existe
            if not meta.vendedor:
                continue

            # Calcular projeção para este vendedor
            try:
                projecao = calcular_projecao_mes(
                    receita_atual=meta.receita_alcancada or 0.0,
                    meta_mes=meta.valor_meta or 0.0,
                    ano=ano_atual,
                    mes=mes_atual,
                    dia_atual=(
                        hoje.day
                        if (mes_atual == hoje.month and ano_atual == hoje.year)
                        else None
                    ),
                )
            except Exception as e:
                app.logger.error(f"Erro ao calcular projeção para meta {meta.id}: {e}")
                # Usar valores padrão se houver erro
                projecao = {
                    "media_diaria": 0.0,
                    "projecao_mes": 0.0,
                    "percentual_projecao": 0.0,
                    "dias_uteis_total": 0,
                    "dias_uteis_trabalhados": 0,
                    "dias_uteis_restantes": 0,
                    "meta_diaria_necessaria": 0.0,
                    "receita_faltante": 0.0,
                    "status_projecao": "Sem dados",
                }

            # Buscar nome do supervisor (já carregado via joinedload)
            supervisor_nome = "Sem supervisor"
            supervisor_obj = None
            try:
                # Usar relacionamento já carregado
                if meta.vendedor.supervisor_obj:
                    supervisor_obj = meta.vendedor.supervisor_obj
                    supervisor_nome = supervisor_obj.nome
                elif meta.vendedor.supervisor_nome:
                    supervisor_nome = meta.vendedor.supervisor_nome
            except Exception as e:
                app.logger.error(f"Erro ao buscar supervisor para vendedor {meta.vendedor.id}: {e}")

            # Buscar nome da equipe (já carregada via joinedload)
            equipe_nome = "Sem Equipe"
            try:
                if meta.vendedor.equipe_obj:
                    equipe_nome = meta.vendedor.equipe_obj.nome
            except Exception as e:
                app.logger.error(f"Erro ao buscar equipe para vendedor {meta.vendedor.id}: {e}")

            vendedores_data.append(
                {
                    "id": meta.id,
                    "nome": meta.vendedor.nome,
                    "supervisor": supervisor_nome,
                    "supervisor_obj": supervisor_obj,
                    "equipe_nome": equipe_nome,
                    "meta": meta.valor_meta or 0.0,
                    "receita_alcancada": meta.receita_alcancada or 0.0,
                    "percentual_alcance": meta.percentual_alcance or 0.0,
                    "percentual_alcance_formatado": (
                        f"{(meta.percentual_alcance or 0.0):.2f}%"
                    ),
                    "comissao_total": meta.comissao_total or 0.0,
                    "comissao_total_formatada": (
                        f"R$ {(meta.comissao_total or 0.0):,.2f}".replace(
                            ",", "X"
                        )
                        .replace(".", ",")
                        .replace("X", ".")
                    ),
                    "status_comissao": meta.status_comissao or "Pendente",
                    # Dados de projeção
                    "projecao": {
                        "media_diaria": projecao["media_diaria"],
                        "media_diaria_formatada": formatar_moeda(
                            projecao["media_diaria"]
                        ),
                        "projecao_mes": projecao["projecao_mes"],
                        "projecao_mes_formatada": formatar_moeda(
                            projecao["projecao_mes"]
                        ),
                        "percentual_projecao": projecao["percentual_projecao"],
                        "dias_uteis_total": projecao["dias_uteis_total"],
                        "dias_uteis_trabalhados": projecao[
                            "dias_uteis_trabalhados"
                        ],
                        "dias_uteis_restantes": projecao[
                            "dias_uteis_restantes"
                        ],
                        "meta_diaria_necessaria": projecao[
                            "meta_diaria_necessaria"
                        ],
                        "meta_diaria_necessaria_formatada": formatar_moeda(
                            projecao["meta_diaria_necessaria"]
                        ),
                        "receita_faltante": projecao["receita_faltante"],
                        "receita_faltante_formatada": formatar_moeda(
                            projecao["receita_faltante"]
                        ),
                        "status_projecao": projecao["status_projecao"],
                    },
                }
            )

        # Ordenar por percentual de alcance (ranking)
        vendedores_data.sort(
            key=lambda x: x["percentual_alcance"], reverse=True
        )

        # Calcular resumo global
        total_receita = sum(v["receita_alcancada"] for v in vendedores_data)
        total_meta = sum(v["meta"] for v in vendedores_data)
        total_comissao = sum(v["comissao_total"] for v in vendedores_data)

        # Calcular projeção global da equipe
        projecao_global = calcular_projecao_mes(
            receita_atual=total_receita,
            meta_mes=total_meta,
            ano=ano_atual,
            mes=mes_atual,
            dia_atual=(
                hoje.day
                if (mes_atual == hoje.month and ano_atual == hoje.year)
                else None
            ),
        )

        # ===== AGRUPAMENTO POR EQUIPE/MESA =====
        projecoes_por_equipe = {}
        for v in vendedores_data:
            # Buscar vendedor e equipe
            vendedor_obj = Vendedor.query.get(
                Meta.query.get(v["id"]).vendedor_id
            )
            if vendedor_obj and vendedor_obj.equipe_id:
                equipe = Equipe.query.get(vendedor_obj.equipe_id)
                equipe_nome = (
                    equipe.nome
                    if equipe
                    else f"Equipe {vendedor_obj.equipe_id}"
                )
            else:
                equipe_nome = "Sem Equipe"

            if equipe_nome not in projecoes_por_equipe:
                projecoes_por_equipe[equipe_nome] = {
                    "nome": equipe_nome,
                    "vendedores": 0,
                    "receita_total": 0.0,
                    "meta_total": 0.0,
                    "comissao_total": 0.0,
                }

            projecoes_por_equipe[equipe_nome]["vendedores"] += 1
            projecoes_por_equipe[equipe_nome]["receita_total"] += v[
                "receita_alcancada"
            ]
            projecoes_por_equipe[equipe_nome]["meta_total"] += v["meta"]
            projecoes_por_equipe[equipe_nome]["comissao_total"] += v[
                "comissao_total"
            ]

        # Calcular projeções para cada equipe
        for equipe_nome, dados in projecoes_por_equipe.items():
            projecao_equipe = calcular_projecao_mes(
                receita_atual=dados["receita_total"],
                meta_mes=dados["meta_total"],
                ano=ano_atual,
                mes=mes_atual,
                dia_atual=(
                    hoje.day
                    if (mes_atual == hoje.month and ano_atual == hoje.year)
                    else None
                ),
            )
            # Adicionar formatação dos valores
            dados["projecao"] = {
                "media_diaria": projecao_equipe["media_diaria"],
                "media_diaria_formatada": formatar_moeda(
                    projecao_equipe["media_diaria"]
                ),
                "projecao_mes": projecao_equipe["projecao_mes"],
                "projecao_mes_formatada": formatar_moeda(
                    projecao_equipe["projecao_mes"]
                ),
                "percentual_projecao": projecao_equipe["percentual_projecao"],
                "status_projecao": projecao_equipe["status_projecao"],
            }
            dados["percentual_alcance"] = (
                (dados["receita_total"] / dados["meta_total"] * 100)
                if dados["meta_total"] > 0
                else 0
            )

        # ===== AGRUPAMENTO POR SUPERVISÃO =====
        projecoes_por_supervisor = {}
        for v in vendedores_data:
            supervisor_nome = v["supervisor"]
            # Objeto Usuario do supervisor
            supervisor_obj = v.get("supervisor_obj")

            if supervisor_nome not in projecoes_por_supervisor:
                # Calcular meta do supervisor (soma das metas dos vendedores)
                meta_supervisionada = 0.0
                gerente_nome = None

                if supervisor_obj:
                    meta_supervisionada = (
                        supervisor_obj.calcular_meta_supervisionada(
                            mes_atual, ano_atual
                        )
                    )
                    # Buscar o gerente do supervisor
                    if supervisor_obj.gerente:
                        gerente_nome = supervisor_obj.gerente.nome

                projecoes_por_supervisor[supervisor_nome] = {
                    "nome": supervisor_nome,
                    "gerente": gerente_nome or "N/A",
                    "vendedores": 0,
                    "receita_total": 0.0,
                    "meta_total": 0.0,
                    "meta_supervisionada": meta_supervisionada,
                    "comissao_total": 0.0,
                }

            projecoes_por_supervisor[supervisor_nome]["vendedores"] += 1
            projecoes_por_supervisor[supervisor_nome]["receita_total"] += v[
                "receita_alcancada"
            ]
            projecoes_por_supervisor[supervisor_nome]["meta_total"] += v[
                "meta"
            ]
            projecoes_por_supervisor[supervisor_nome]["comissao_total"] += v[
                "comissao_total"
            ]

        # Calcular projeções para cada supervisor
        for supervisor_nome, dados in projecoes_por_supervisor.items():
            projecao_supervisor = calcular_projecao_mes(
                receita_atual=dados["receita_total"],
                meta_mes=dados["meta_total"],
                ano=ano_atual,
                mes=mes_atual,
                dia_atual=(
                    hoje.day
                    if (mes_atual == hoje.month and ano_atual == hoje.year)
                    else None
                ),
            )
            # Adicionar formatação dos valores
            dados["projecao"] = {
                "media_diaria": projecao_supervisor["media_diaria"],
                "media_diaria_formatada": formatar_moeda(
                    projecao_supervisor["media_diaria"]
                ),
                "projecao_mes": projecao_supervisor["projecao_mes"],
                "projecao_mes_formatada": formatar_moeda(
                    projecao_supervisor["projecao_mes"]
                ),
                "percentual_projecao": projecao_supervisor[
                    "percentual_projecao"
                ],
                "status_projecao": projecao_supervisor["status_projecao"],
            }
            dados["percentual_alcance"] = (
                (dados["receita_total"] / dados["meta_total"] * 100)
                if dados["meta_total"] > 0
                else 0
            )

            # Comissão do supervisor baseada nas faixas de supervisor
            taxa_sup = _obter_taxa_por_alcance(
                "supervisor",
                current_user.empresa_id if current_user.cargo != "super_admin" else None,
                dados["percentual_alcance"],
            )
            dados["taxa_supervisor"] = taxa_sup
            dados["comissao_supervisor"] = dados["receita_total"] * taxa_sup

        # Converter para listas ordenadas
        projecoes_por_equipe = sorted(
            projecoes_por_equipe.values(),
            key=lambda x: x["percentual_alcance"],
            reverse=True,
        )
        projecoes_por_supervisor = sorted(
            projecoes_por_supervisor.values(),
            key=lambda x: x["percentual_alcance"],
            reverse=True,
        )

        resumo_global = {
            "total_vendedores": len(vendedores_data),
            "total_receita": total_receita,
            "total_receita_formatada": (
                f"R$ {total_receita:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            ),
            "total_meta": total_meta,
            "total_meta_formatada": (
                f"R$ {total_meta:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            ),
            "total_comissao_formatada": (
                f"R$ {total_comissao:,.2f}".replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            ),
            "alcance_equipe": (
                (total_receita / total_meta * 100) if total_meta > 0 else 0
            ),
            "mes": mes_atual,
            "ano": ano_atual,
            # Projeção global
            "projecao_global": {
                "media_diaria": projecao_global["media_diaria"],
                "media_diaria_formatada": formatar_moeda(
                    projecao_global["media_diaria"]
                ),
                "projecao_mes": projecao_global["projecao_mes"],
                "projecao_mes_formatada": formatar_moeda(
                    projecao_global["projecao_mes"]
                ),
                "percentual_projecao": projecao_global["percentual_projecao"],
                "dias_uteis_total": projecao_global["dias_uteis_total"],
                "dias_uteis_trabalhados": projecao_global[
                    "dias_uteis_trabalhados"
                ],
                "dias_uteis_restantes": projecao_global[
                    "dias_uteis_restantes"
                ],
                "status_projecao": projecao_global["status_projecao"],
            },
        }

        result = render_template(
            "dashboard.html",
            vendedores=vendedores_data,
            resumo_global=resumo_global,
            projecoes_por_equipe=projecoes_por_equipe,
            projecoes_por_supervisor=projecoes_por_supervisor,
        )
        
        # Armazenar no cache (5 minutos)
        if cache:
            cache.set(cache_key, result, timeout=300)
        
        return result

    except Exception as e:
        app.logger.error(f"Erro no dashboard: {e}")
        import traceback

        app.logger.error(traceback.format_exc())
        flash(f"Erro ao carregar dashboard: {str(e)}", "danger")
        return render_template(
            "dashboard.html",
            vendedores=[],
            resumo_global={
                "total_vendedores": 0,
                "total_receita": 0.0,
                "total_receita_formatada": "R$ 0,00",
                "total_meta": 0.0,
                "total_meta_formatada": "R$ 0,00",
                "total_comissao_formatada": "R$ 0,00",
                "alcance_equipe": 0.0,
                "mes": datetime.now().month,
                "ano": datetime.now().year,
                "projecao_global": {
                    "media_diaria": 0.0,
                    "media_diaria_formatada": "R$ 0,00",
                    "projecao_mes": 0.0,
                    "projecao_mes_formatada": "R$ 0,00",
                    "percentual_projecao": 0.0,
                    "dias_uteis_total": 0,
                    "dias_uteis_trabalhados": 0,
                    "dias_uteis_restantes": 0,
                    "status_projecao": "Erro",
                },
            },
        )

@app.route("/supervisor/dashboard")
@login_required
def supervisor_dashboard():
    """Dashboard específico para supervisores com projeções consolidadas"""
    # Verificar se o usuário é supervisor
    if current_user.cargo != "supervisor":
        flash(
            "Acesso negado. Esta área é exclusiva para supervisores.", "danger"
        )
        return redirect(url_for("dashboard"))

    # Obter o mês e ano atuais ou filtrados
    hoje = datetime.now()
    mes_atual = request.args.get("mes", hoje.month, type=int)
    ano_atual = request.args.get("ano", hoje.year, type=int)

    # Buscar vendedores vinculados a este supervisor
    vendedores = Vendedor.query.filter_by(
        supervisor_id=current_user.id, ativo=True
    ).all()

    if not vendedores:
        # Faixas dinâmicas para legenda do supervisor
        faixas_supervisor = (
            FaixaComissaoSupervisor.query.filter_by(
                empresa_id=current_user.empresa_id
            ).order_by(FaixaComissaoSupervisor.ordem).all()
        )
        if not faixas_supervisor:
            faixas_supervisor = (
                FaixaComissaoSupervisor.query.filter(
                    FaixaComissaoSupervisor.empresa_id.is_(None)
                ).order_by(FaixaComissaoSupervisor.ordem).all()
            )

        return render_template(
            "supervisores/dashboard.html",
            vendedores_data=[],
            resumo_supervisor={},
            mes_atual=mes_atual,
            ano_atual=ano_atual,
            faixas_supervisor=faixas_supervisor,
        )

    # Processar dados dos vendedores
    vendedores_data = []
    total_meta = 0
    total_receita = 0
    total_comissao = 0

    for vendedor in vendedores:
        # Buscar meta do vendedor no período
        meta = Meta.query.filter_by(
            vendedor_id=vendedor.id, mes=mes_atual, ano=ano_atual
        ).first()

        if meta:
            # Calcular projeção
            projecao = calcular_projecao_mes(
                receita_atual=meta.receita_alcancada,
                meta_mes=meta.valor_meta,
                ano=ano_atual,
                mes=mes_atual,
                dia_atual=(
                    hoje.day
                    if (mes_atual == hoje.month and ano_atual == hoje.year)
                    else None
                ),
            )

            vendedores_data.append(
                {
                    "id": vendedor.id,
                    "nome": vendedor.nome,
                    "email": vendedor.email,
                    "meta": meta.valor_meta,
                    "receita_alcancada": meta.receita_alcancada,
                    "percentual_alcance": meta.percentual_alcance,
                    "comissao_total": meta.comissao_total,
                    "status_comissao": meta.status_comissao,
                    "projecao": projecao,
                }
            )

            total_meta += meta.valor_meta
            total_receita += meta.receita_alcancada
            total_comissao += meta.comissao_total

    # Ordenar por percentual de alcance
    vendedores_data.sort(key=lambda x: x["percentual_alcance"], reverse=True)

    # Calcular projeção consolidada do supervisor
    projecao_supervisor = calcular_projecao_mes(
        receita_atual=total_receita,
        meta_mes=total_meta,
        ano=ano_atual,
        mes=mes_atual,
        dia_atual=(
            hoje.day
            if (mes_atual == hoje.month and ano_atual == hoje.year)
            else None
        ),
    )

    # Resumo do supervisor
    resumo_supervisor = {
        "total_vendedores": len(vendedores_data),
        "total_meta": total_meta,
        "total_meta_formatada": formatar_moeda(total_meta),
        "total_receita": total_receita,
        "total_receita_formatada": formatar_moeda(total_receita),
        "total_comissao": total_comissao,
        "total_comissao_formatada": formatar_moeda(total_comissao),
        "percentual_alcance": (
            (total_receita / total_meta * 100) if total_meta > 0 else 0
        ),
        "projecao": projecao_supervisor,
        "mes": mes_atual,
        "ano": ano_atual,
    }

    # Comissão específica do supervisor pelas faixas de supervisor
    try:
        taxa_sup = _obter_taxa_por_alcance(
            "supervisor",
            current_user.empresa_id if current_user.cargo != "super_admin" else None,
            resumo_supervisor["percentual_alcance"],
        )
        resumo_supervisor["taxa_supervisor"] = taxa_sup
        resumo_supervisor["comissao_supervisor"] = total_receita * taxa_sup
        resumo_supervisor["comissao_supervisor_formatada"] = formatar_moeda(
            resumo_supervisor["comissao_supervisor"]
        )
    except Exception as e:
        app.logger.error(f"Erro ao calcular comissão do supervisor: {e}")

    # Faixas dinâmicas para legenda do supervisor
    faixas_supervisor = (
        FaixaComissaoSupervisor.query.filter_by(
            empresa_id=current_user.empresa_id
        ).order_by(FaixaComissaoSupervisor.ordem).all()
    )
    if not faixas_supervisor:
        faixas_supervisor = (
            FaixaComissaoSupervisor.query.filter(
                FaixaComissaoSupervisor.empresa_id.is_(None)
            ).order_by(FaixaComissaoSupervisor.ordem).all()
        )

    return render_template(
        "supervisores/dashboard.html",
        vendedores_data=vendedores_data,
        resumo_supervisor=resumo_supervisor,
        mes_atual=mes_atual,
        ano_atual=ano_atual,
        faixas_supervisor=faixas_supervisor,
    )

@app.route("/vendedor/dashboard")
@login_required
def vendedor_dashboard():
    """Dashboard mobile para vendedores"""
    # Verificar se o usuário é um vendedor
    if not current_user.vendedor_id:
        flash(
            "Acesso negado. Esta área é exclusiva para vendedores.", "danger"
        )
        return redirect(url_for("dashboard"))

    vendedor = current_user.vendedor

    # Obter mês e ano atuais
    hoje = datetime.now()
    mes_atual = hoje.month
    ano_atual = hoje.year

    # Buscar meta do mês atual
    meta_atual = Meta.query.filter_by(
        vendedor_id=vendedor.id, mes=mes_atual, ano=ano_atual
    ).first()

    # Se não houver meta, mostrar mensagem
    if not meta_atual:
        return render_template(
            "vendedor/dashboard.html",
            vendedor=vendedor,
            meta_atual=None,
            projecao=None,
            ranking_equipe=[],
            historico=[],
        )

    # Garantir que percentual de alcance e comissão estejam atualizados
    try:
        meta_atual.calcular_comissao()
        db.session.commit()
    except Exception as e:
        app.logger.error(
            f"Erro ao recalcular comissão da meta {meta_atual.id} no dashboard do vendedor: {e}"
        )

    # Calcular projeção
    projecao = calcular_projecao_mes(
        receita_atual=meta_atual.receita_alcancada,
        meta_mes=meta_atual.valor_meta,
        ano=ano_atual,
        mes=mes_atual,
        dia_atual=hoje.day,
    )

    # Buscar ranking da equipe (se o vendedor tiver equipe)
    ranking_equipe = []
    if vendedor.equipe_id:
        # Buscar todas as metas da equipe no mês atual
        metas_equipe = (
            Meta.query.join(Vendedor)
            .filter(
                Vendedor.equipe_id == vendedor.equipe_id,
                Meta.mes == mes_atual,
                Meta.ano == ano_atual,
                Vendedor.ativo,
            )
            .all()
        )

        # Processar dados do ranking
        for meta in metas_equipe:
            ranking_equipe.append(
                {
                    "id": meta.vendedor.id,
                    "nome": meta.vendedor.nome,
                    "meta": meta.valor_meta,
                    "receita": meta.receita_alcancada,
                    "percentual": meta.percentual_alcance,
                    "is_current_user": meta.vendedor_id == vendedor.id,
                }
            )

        # Ordenar por percentual de alcance
        ranking_equipe.sort(key=lambda x: x["percentual"], reverse=True)

        # Adicionar posição
        for idx, item in enumerate(ranking_equipe, 1):
            item["posicao"] = idx

    # Buscar histórico dos últimos 3 meses
    historico = []
    for i in range(3):
        # Calcular mês e ano
        mes_hist = mes_atual - i
        ano_hist = ano_atual

        if mes_hist <= 0:
            mes_hist += 12
            ano_hist -= 1

        # Buscar meta do mês
        meta_hist = Meta.query.filter_by(
            vendedor_id=vendedor.id, mes=mes_hist, ano=ano_hist
        ).first()

        if meta_hist:
            historico.append(
                {
                    "mes": mes_hist,
                    "ano": ano_hist,
                    "mes_nome": [
                        "Jan",
                        "Fev",
                        "Mar",
                        "Abr",
                        "Mai",
                        "Jun",
                        "Jul",
                        "Ago",
                        "Set",
                        "Out",
                        "Nov",
                        "Dez",
                    ][mes_hist - 1],
                    "meta": meta_hist.valor_meta,
                    "receita": meta_hist.receita_alcancada,
                    "percentual": meta_hist.percentual_alcance,
                }
            )

    return render_template(
        "vendedor/dashboard.html",
        vendedor=vendedor,
        meta_atual=meta_atual,
        projecao=projecao,
        ranking_equipe=ranking_equipe,
        historico=historico,
        hoje=hoje,
    )

# ===== ROTAS DE VENDEDORES =====

@app.route("/vendedores")
@login_required
def lista_vendedores():
    """Lista todos os vendedores"""
    page = request.args.get("page", 1, type=int)

    # Quantidade de registros por página (lista operacional, costuma ser menor que clientes)
    per_page = 20

    # Construir query base respeitando permissões
    if current_user.is_super_admin:
        base_query = Vendedor.query.filter_by(ativo=True)
    elif current_user.cargo == "supervisor":
        # Supervisor vê apenas seus vendedores
        base_query = Vendedor.query.filter_by(
            supervisor_id=current_user.id, ativo=True
        )
    elif current_user.cargo == "vendedor" and current_user.vendedor_id:
        # Vendedor vê apenas ele mesmo e vendedores da mesma equipe (ou só ele, se sem equipe)
        vendedor_atual = Vendedor.query.get(current_user.vendedor_id)
        if vendedor_atual and vendedor_atual.equipe_id:
            base_query = Vendedor.query.filter_by(
                equipe_id=vendedor_atual.equipe_id, ativo=True
            )
        else:
            base_query = Vendedor.query.filter_by(
                id=current_user.vendedor_id, ativo=True
            )
    else:
        base_query = Vendedor.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        )

    # Estatísticas globais (baseadas no conjunto filtrado)
    total_vendedores = base_query.count()
    total_com_supervisor = base_query.filter(
        Vendedor.supervisor_id.isnot(None)
    ).count()
    total_em_equipes = base_query.filter(
        Vendedor.equipe_id.isnot(None)
    ).count()

    stats = {
        "total": total_vendedores,
        "com_supervisor": total_com_supervisor,
        "em_equipes": total_em_equipes,
    }

    # Paginação ordenada por nome
    pagination = base_query.order_by(Vendedor.nome).paginate(
        page=page, per_page=per_page, error_out=False
    )
    vendedores = pagination.items

    return render_template(
        "vendedores/lista.html",
        vendedores=vendedores,
        pagination=pagination,
        stats=stats,
    )

@app.route("/vendedores/novo", methods=["GET", "POST"])
@login_required
def novo_vendedor():
    """Cadastrar novo vendedor"""
    form = VendedorForm()

    # Preencher choices de supervisores da mesma empresa
    if current_user.is_super_admin:
        supervisores = Usuario.query.filter_by(
            cargo="supervisor", ativo=True
        ).all()
    else:
        supervisores = Usuario.query.filter_by(
            empresa_id=current_user.empresa_id, cargo="supervisor", ativo=True
        ).all()

    form.supervisor_id.choices = [(0, "Selecione um supervisor")] + [
        (s.id, s.nome) for s in supervisores
    ]

    # Preencher choices de equipes da mesma empresa
    if current_user.is_super_admin:
        equipes = Equipe.query.filter_by(ativa=True).all()
    else:
        equipes = Equipe.query.filter_by(
            empresa_id=current_user.empresa_id, ativa=True
        ).all()

    form.equipe_id.choices = [(0, "Selecione uma equipe")] + [
        (e.id, e.nome) for e in equipes
    ]

    if form.validate_on_submit():
        # Processar supervisor_id - converter 0 para None
        supervisor_id = None
        if form.supervisor_id.data and form.supervisor_id.data != 0:
            supervisor_id = form.supervisor_id.data

        supervisor = None
        if supervisor_id:
            supervisor = Usuario.query.get(supervisor_id)

        # Processar equipe_id - converter 0 para None
        equipe_id = None
        if form.equipe_id.data and form.equipe_id.data != 0:
            equipe_id = form.equipe_id.data

        vendedor = Vendedor(
            nome=form.nome.data,
            email=form.email.data,
            telefone=form.telefone.data or None,
            cpf=form.cpf.data or None,
            empresa_id=current_user.empresa_id,
            supervisor_id=supervisor_id,
            supervisor_nome=supervisor.nome if supervisor else None,
            equipe_id=equipe_id,
        )

        db.session.add(vendedor)
        db.session.commit()

        flash(f"Vendedor {vendedor.nome} cadastrado com sucesso!", "success")
        return redirect(url_for("lista_vendedores"))

    return render_template(
        "vendedores/form.html", form=form, titulo="Novo Vendedor"
    )

@app.route("/vendedores/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_vendedor(id):
    """Editar vendedor existente"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    # (exceto super admin)
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            msg = "Você não tem permissão " "para editar este vendedor."
            flash(msg, "danger")
            return redirect(url_for("lista_vendedores"))

    form = VendedorForm(vendedor_id=id, obj=vendedor)

    # Preencher choices de supervisores da mesma empresa
    if current_user.is_super_admin:
        supervisores = Usuario.query.filter_by(
            cargo="supervisor", ativo=True
        ).all()
    else:
        supervisores = Usuario.query.filter_by(
            empresa_id=current_user.empresa_id, cargo="supervisor", ativo=True
        ).all()
    form.supervisor_id.choices = [(0, "Selecione um supervisor")] + [
        (s.id, s.nome) for s in supervisores
    ]

    # Preencher choices de equipes da mesma empresa
    if current_user.is_super_admin:
        equipes = Equipe.query.filter_by(ativa=True).all()
    else:
        equipes = Equipe.query.filter_by(
            empresa_id=current_user.empresa_id, ativa=True
        ).all()
    form.equipe_id.choices = [(0, "Selecione uma equipe")] + [
        (e.id, e.nome) for e in equipes
    ]

    # Pré-preencher o formulário com dados do vendedor
    if request.method == "GET":
        if vendedor.supervisor_id:
            form.supervisor_id.data = vendedor.supervisor_id
        else:
            form.supervisor_id.data = 0
        if vendedor.equipe_id:
            form.equipe_id.data = vendedor.equipe_id
        else:
            form.equipe_id.data = 0

    if form.validate_on_submit():
        # Processar supervisor_id - converter 0 para None
        if form.supervisor_id.data and form.supervisor_id.data != 0:
            supervisor_id = form.supervisor_id.data
        else:
            supervisor_id = None
        supervisor = None
        if supervisor_id:
            supervisor = Usuario.query.get(supervisor_id)

        # Processar equipe_id - converter 0 para None
        if form.equipe_id.data and form.equipe_id.data != 0:
            equipe_id = form.equipe_id.data
        else:
            equipe_id = None

        vendedor.nome = form.nome.data
        vendedor.email = form.email.data
        vendedor.telefone = form.telefone.data or None
        vendedor.cpf = form.cpf.data or None
        vendedor.supervisor_id = supervisor_id
        vendedor.supervisor_nome = supervisor.nome if supervisor else None
        vendedor.equipe_id = equipe_id
        # Nota: empresa_id não deve ser alterado na edição

        db.session.commit()
        flash(f"Vendedor {vendedor.nome} atualizado com sucesso!", "success")
        return redirect(url_for("lista_vendedores"))

    return render_template(
        "vendedores/form.html",
        form=form,
        titulo="Editar Vendedor",
        vendedor=vendedor,
    )

@app.route("/vendedores/<int:id>/deletar", methods=["POST"])
@login_required
def deletar_vendedor(id):
    """Desativar vendedor"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    # (exceto super admin)
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            msg = "Você não tem permissão " "para deletar este vendedor."
            flash(msg, "danger")
            return redirect(url_for("lista_vendedores"))

    vendedor.ativo = False
    db.session.commit()
    msg = f"Vendedor {vendedor.nome} desativado com sucesso!"
    flash(msg, "success")
    return redirect(url_for("lista_vendedores"))

@app.route("/vendedores/<int:id>/criar-login", methods=["GET", "POST"])
@login_required
@permission_required("pode_gerenciar_vendedores")
def criar_login_vendedor(id):
    """Criar login de acesso para o vendedor"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para acessar este vendedor.", "danger"
            )
            return redirect(url_for("lista_vendedores"))

    # Verificar se já tem login
    usuario_existente = Usuario.query.filter_by(
        vendedor_id=vendedor.id
    ).first()
    if usuario_existente:
        flash(f"Vendedor {vendedor.nome} já possui login cadastrado!", "warning")
        return redirect(url_for("lista_vendedores"))

    if request.method == "POST":
        email = request.form.get("email", "").strip()
        senha = request.form.get("senha")
        confirmar_senha = request.form.get("confirmar_senha")

        if not email or not senha or not confirmar_senha:
            flash("Preencha todos os campos obrigatórios!", "danger")
            return render_template(
                "vendedores/criar_login.html", vendedor=vendedor
            )

        if senha != confirmar_senha:
            flash("As senhas não coincidem!", "danger")
            return render_template(
                "vendedores/criar_login.html", vendedor=vendedor
            )

        if len(senha) < 6:
            flash("A senha deve ter no mínimo 6 caracteres!", "danger")
            return render_template(
                "vendedores/criar_login.html", vendedor=vendedor
            )

        # Verificar se email já está em uso
        email_existe = Usuario.query.filter_by(email=email).first()
        if email_existe:
            flash(
                f"O email {email} já está cadastrado no sistema! Use um email diferente.",
                "danger",
            )
            return render_template(
                "vendedores/criar_login.html", vendedor=vendedor
            )

        try:
            # Criar usuário do tipo vendedor
            usuario = Usuario(
                nome=vendedor.nome,
                email=email,
                cargo="vendedor",
                empresa_id=vendedor.empresa_id,
                vendedor_id=vendedor.id,
                is_super_admin=False,
                ativo=True,
                # Permissões padrão do vendedor
                pode_ver_dashboard=True,
                pode_gerenciar_vendedores=False,
                pode_gerenciar_metas=False,
                pode_gerenciar_equipes=False,
                pode_gerenciar_comissoes=False,
                pode_enviar_mensagens=True,
                pode_exportar_dados=False,
                pode_ver_todas_metas=False,
                pode_aprovar_comissoes=False,
                # Clientes: vendedor pode cadastrar e gerenciar sua carteira
                pode_acessar_clientes=True,
                pode_criar_clientes=True,
                pode_editar_clientes=True,
                pode_excluir_clientes=False,
                pode_importar_clientes=False,
                # Estoque: acesso somente leitura ao saldo/níveis
                pode_acessar_estoque=True,
                pode_gerenciar_produtos=False,
                pode_movimentar_estoque=False,
                pode_ver_custos=False,
                pode_ajustar_estoque=False,
            )
            usuario.set_senha(senha)

            db.session.add(usuario)
            db.session.commit()

            flash(f"Login criado com sucesso para {vendedor.nome}!", "success")
            return redirect(url_for("lista_vendedores"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao criar login: {str(e)}", "danger")

    return render_template("vendedores/criar_login.html", vendedor=vendedor)

@app.route("/vendedores/<int:id>/editar-login", methods=["GET", "POST"])
@login_required
@permission_required("pode_gerenciar_vendedores")
def editar_login_vendedor(id):
    """Editar informações de login do vendedor"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para acessar este vendedor.", "danger"
            )
            return redirect(url_for("lista_vendedores"))

    # Buscar usuário do vendedor
    usuario = Usuario.query.filter_by(vendedor_id=vendedor.id).first()
    if not usuario:
        flash(f"Vendedor {vendedor.nome} não possui login cadastrado!", "warning")
        return redirect(url_for("lista_vendedores"))

    if request.method == "POST":
        email = request.form.get("email")
        nome = request.form.get("nome")
        nova_senha = request.form.get("nova_senha")
        confirmar_senha = request.form.get("confirmar_senha")

        if not email or not nome:
            flash("Preencha todos os campos obrigatórios!", "danger")
            return render_template(
                "vendedores/editar_login.html",
                vendedor=vendedor,
                usuario=usuario,
            )

        # Se informou senha, validar
        if nova_senha or confirmar_senha:
            if nova_senha != confirmar_senha:
                flash("As senhas não coincidem!", "danger")
                return render_template(
                    "vendedores/editar_login.html",
                    vendedor=vendedor,
                    usuario=usuario,
                )

            if len(nova_senha) < 6:
                flash("A senha deve ter no mínimo 6 caracteres!", "danger")
                return render_template(
                    "vendedores/editar_login.html",
                    vendedor=vendedor,
                    usuario=usuario,
                )

        try:
            # Atualizar dados do usuário
            usuario.email = email
            usuario.nome = nome

            # Atualizar senha se fornecida
            if nova_senha:
                usuario.set_senha(nova_senha)

            db.session.commit()

            flash(f"Login de {vendedor.nome} atualizado com sucesso!", "success")
            return redirect(url_for("lista_vendedores"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar login: {str(e)}", "danger")

    return render_template(
        "vendedores/editar_login.html", vendedor=vendedor, usuario=usuario
    )

@app.route("/vendedores/<int:id>/excluir-login", methods=["POST"])
@login_required
@permission_required("pode_gerenciar_vendedores")
def excluir_login_vendedor(id):
    """Excluir login de acesso do vendedor"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para acessar este vendedor.", "danger"
            )
            return redirect(url_for("lista_vendedores"))

    # Buscar usuário do vendedor
    usuario = Usuario.query.filter_by(vendedor_id=vendedor.id).first()
    if not usuario:
        flash(f"Vendedor {vendedor.nome} não possui login cadastrado!", "warning")
        return redirect(url_for("lista_vendedores"))

    try:
        # Remover vinculo e desativar usuário
        usuario.vendedor_id = None
        usuario.ativo = False
        usuario.bloqueado = True
        usuario.motivo_bloqueio = "Login excluído pelo administrador"

        db.session.commit()

        flash(f"Login de {vendedor.nome} excluído com sucesso!", "success")
        return redirect(url_for("lista_vendedores"))

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir login: {str(e)}", "danger")
        return redirect(url_for("lista_vendedores"))

@app.route("/vendedores/<int:id>/resetar-senha", methods=["GET", "POST"])
@login_required
@permission_required("pode_gerenciar_vendedores")
def resetar_senha_vendedor(id):
    """Resetar senha do login do vendedor"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para acessar este vendedor.", "danger"
            )
            return redirect(url_for("lista_vendedores"))

    # Buscar usuário do vendedor
    usuario = Usuario.query.filter_by(vendedor_id=vendedor.id).first()
    if not usuario:
        flash(f"Vendedor {vendedor.nome} não possui login cadastrado!", "warning")
        return redirect(url_for("lista_vendedores"))

    if request.method == "POST":
        nova_senha = request.form.get("nova_senha")
        confirmar_senha = request.form.get("confirmar_senha")

        if not nova_senha or not confirmar_senha:
            flash("Preencha todos os campos!", "danger")
            return render_template(
                "vendedores/resetar_senha.html", vendedor=vendedor
            )

        if nova_senha != confirmar_senha:
            flash("As senhas não coincidem!", "danger")
            return render_template(
                "vendedores/resetar_senha.html", vendedor=vendedor
            )

        if len(nova_senha) < 6:
            flash("A senha deve ter no mínimo 6 caracteres!", "danger")
            return render_template(
                "vendedores/resetar_senha.html", vendedor=vendedor
            )

        try:
            usuario.set_senha(nova_senha)
            db.session.commit()

            flash(f"Senha resetada com sucesso para {vendedor.nome}!", "success")
            return redirect(url_for("lista_vendedores"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao resetar senha: {str(e)}", "danger")

    return render_template("vendedores/resetar_senha.html", vendedor=vendedor)

@app.route("/vendedores/<int:id>/ativar", methods=["POST"])
@login_required
@permission_required("pode_gerenciar_vendedores")
def ativar_vendedor(id):
    """Ativar vendedor desativado"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para acessar este vendedor.", "danger"
            )
            return redirect(url_for("lista_vendedores"))

    vendedor.ativo = True

    # Ativar também o usuário se existir
    usuario = Usuario.query.filter_by(vendedor_id=vendedor.id).first()
    if usuario:
        usuario.ativo = True

    db.session.commit()
    flash(f"Vendedor {vendedor.nome} ativado com sucesso!", "success")
    return redirect(url_for("lista_vendedores"))

@app.route("/vendedores/<int:id>/desativar", methods=["POST"])
@login_required
@permission_required("pode_gerenciar_vendedores")
def desativar_vendedor(id):
    """Desativar vendedor"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para acessar este vendedor.", "danger"
            )
            return redirect(url_for("lista_vendedores"))

    vendedor.ativo = False

    # Desativar também o usuário se existir
    usuario = Usuario.query.filter_by(vendedor_id=vendedor.id).first()
    if usuario:
        usuario.ativo = False
        usuario.bloqueado = True
        usuario.motivo_bloqueio = "Vendedor desativado"

    db.session.commit()
    flash(f"Vendedor {vendedor.nome} desativado com sucesso!", "success")
    return redirect(url_for("lista_vendedores"))

@app.route("/vendedores/<int:id>/permissoes", methods=["GET", "POST"])
@login_required
@admin_required
def gerenciar_permissoes_vendedor(id):
    """Gerenciar permissões específicas do vendedor"""
    vendedor = Vendedor.query.get_or_404(id)

    # Verificar se o vendedor pertence à empresa do usuário
    if not current_user.is_super_admin:
        if vendedor.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para acessar este vendedor.", "danger"
            )
            return redirect(url_for("lista_vendedores"))

    # Buscar usuário do vendedor
    usuario = Usuario.query.filter_by(vendedor_id=vendedor.id).first()
    if not usuario:
        flash(f"Vendedor {vendedor.nome} não possui login cadastrado!", "warning")
        return redirect(url_for("lista_vendedores"))

    if request.method == "POST":
        # Atualizar permissões
        usuario.pode_ver_dashboard = "pode_ver_dashboard" in request.form
        usuario.pode_gerenciar_vendedores = (
            "pode_gerenciar_vendedores" in request.form
        )
        usuario.pode_gerenciar_metas = "pode_gerenciar_metas" in request.form
        usuario.pode_gerenciar_equipes = (
            "pode_gerenciar_equipes" in request.form
        )
        usuario.pode_gerenciar_comissoes = (
            "pode_gerenciar_comissoes" in request.form
        )
        usuario.pode_enviar_mensagens = "pode_enviar_mensagens" in request.form
        usuario.pode_exportar_dados = "pode_exportar_dados" in request.form
        usuario.pode_ver_todas_metas = "pode_ver_todas_metas" in request.form
        usuario.pode_aprovar_comissoes = (
            "pode_aprovar_comissoes" in request.form
        )

        # Permissões de Clientes
        usuario.pode_acessar_clientes = "pode_acessar_clientes" in request.form
        usuario.pode_criar_clientes = "pode_criar_clientes" in request.form
        usuario.pode_editar_clientes = "pode_editar_clientes" in request.form
        usuario.pode_excluir_clientes = "pode_excluir_clientes" in request.form
        usuario.pode_importar_clientes = "pode_importar_clientes" in request.form

        # Permissões de Estoque
        usuario.pode_acessar_estoque = "pode_acessar_estoque" in request.form
        usuario.pode_gerenciar_produtos = "pode_gerenciar_produtos" in request.form
        usuario.pode_movimentar_estoque = "pode_movimentar_estoque" in request.form
        usuario.pode_ver_custos = "pode_ver_custos" in request.form
        usuario.pode_ajustar_estoque = "pode_ajustar_estoque" in request.form

        db.session.commit()
        flash(f"Permissões de {vendedor.nome} atualizadas com sucesso!", "success")
        return redirect(url_for("lista_vendedores"))

    return render_template(
        "vendedores/permissoes.html", vendedor=vendedor, usuario=usuario
    )

@app.route("/vendedores/importar", methods=["GET", "POST"])
@login_required
def importar_vendedores():
    """Importar vendedores via planilha Excel - Admin, Supervisor, RH e Gerente"""
    # Verificar permissões: admin, supervisor, RH e gerente podem importar vendedores
    if not pode_importar(current_user, "vendedores"):
        flash(
            "Acesso negado! Apenas Administradores, Supervisores, RH e Gerentes podem importar vendedores.",
            "danger",
        )
        return redirect(url_for("lista_vendedores"))
    
    if request.method == "POST":
        # Validar arquivo Excel
        arquivo, erro = validar_arquivo_excel(request)
        if erro:
            flash(erro, "danger")
            return redirect(request.url)

        # Verificar disponibilidade do Excel
        if not EXCEL_AVAILABLE and not ensure_excel_available():
            flash("Erro: Bibliotecas Excel não instaladas. Contate o administrador.", "danger")
            return redirect(request.url)

        try:
            # Ler Excel
            df = pd.read_excel(arquivo)

            # Normalizar nomes das colunas (remover espaços, lowercase)
            df.columns = df.columns.str.strip().str.lower()

            # Mapear colunas com nomes variados
            colunas_map = {
                "nome": ["nome", "nome completo", "vendedor"],
                "email": ["email", "e-mail", "e mail"],
                "telefone": ["telefone", "fone", "celular", "tel"],
                "cpf": ["cpf", "documento"],
            }

            # Renomear colunas para padrão
            for col_padrao, variantes in colunas_map.items():
                for col in df.columns:
                    if col in variantes:
                        df.rename(columns={col: col_padrao}, inplace=True)
                        break

            # Validar colunas obrigatórias
            colunas_obrigatorias = ["nome", "email", "telefone", "cpf"]
            colunas_faltando = [
                col for col in colunas_obrigatorias if col not in df.columns
            ]

            if colunas_faltando:
                # Formatar para exibição
                colunas_exibir = [c.capitalize() for c in colunas_faltando]
                msg = f'Colunas obrigatórias faltando: {", ".join(colunas_exibir)}'
                flash(msg, "danger")
                return redirect(request.url)

            # Determinar empresa
            if current_user.is_super_admin:
                # Super admin deve especificar empresa no Excel
                if (
                    "empresa_cnpj" not in df.columns and
                    "empresa cnpj" not in df.columns
                ):
                    # Renomear se existir com espaço
                    if "empresa cnpj" in df.columns:
                        df.rename(
                            columns={"empresa cnpj": "empresa_cnpj"},
                            inplace=True,
                        )
                    else:
                        msg = (
                            'Super admin deve incluir coluna "Empresa CNPJ" '
                            "no Excel"
                        )
                        flash(msg, "danger")
                        return redirect(request.url)
            else:
                empresa_id = current_user.empresa_id

            erros = []
            sucesso = 0

            for idx, row in df.iterrows():
                try:
                    # Buscar empresa (se super admin)
                    if current_user.is_super_admin:
                        empresa = Empresa.query.filter_by(
                            cnpj=str(row["empresa_cnpj"])
                        ).first()
                        if not empresa:
                            erros.append(
                                f"Linha {idx + 2}: Empresa CNPJ "
                                f"{row['empresa_cnpj']} não encontrada"
                            )
                            continue
                        empresa_id = empresa.id

                    # Verificar duplicatas
                    vendedor_existe = Vendedor.query.filter_by(
                        email=row["email"]
                    ).first()
                    if vendedor_existe:
                        erros.append(f"Linha {idx + 2}: Email {row['email']} já cadastrado")
                        continue

                    # Buscar supervisor (opcional)
                    supervisor_id = None
                    col_supervisor = next(
                        (
                            c
                            for c in df.columns
                            if "supervisor" in c and "email" in c
                        ),
                        None,
                    )
                    if col_supervisor and pd.notna(row[col_supervisor]):
                        supervisor = Usuario.query.filter_by(
                            email=row[col_supervisor], empresa_id=empresa_id
                        ).first()
                        if supervisor:
                            supervisor_id = supervisor.id

                    # Buscar equipe (opcional)
                    equipe_id = None
                    col_equipe = next(
                        (
                            c
                            for c in df.columns
                            if "equipe" in c and "nome" in c
                        ),
                        None,
                    )
                    if col_equipe and pd.notna(row[col_equipe]):
                        equipe = Equipe.query.filter_by(
                            nome=row[col_equipe], empresa_id=empresa_id
                        ).first()
                        if equipe:
                            equipe_id = equipe.id

                    # Criar vendedor
                    vendedor = Vendedor(
                        nome=row["nome"],
                        email=row["email"],
                        telefone=str(row["telefone"]),
                        cpf=str(row["cpf"]).replace(".", "").replace("-", ""),
                        empresa_id=empresa_id,
                        supervisor_id=supervisor_id,
                        equipe_id=equipe_id,
                        ativo=True,
                    )

                    db.session.add(vendedor)
                    db.session.flush()  # Para obter o ID do vendedor

                    # Verificar se há colunas de meta (opcionais)
                    col_mes = next(
                        (
                            c
                            for c in df.columns
                            if "mes" in c.lower() or "mês" in c.lower()
                        ),
                        None,
                    )
                    col_ano = next(
                        (c for c in df.columns if "ano" in c.lower()), None
                    )
                    col_meta = next(
                        (
                            c
                            for c in df.columns
                            if "meta" in c.lower() and "valor" in c.lower()
                        ),
                        None,
                    )

                    # Se todas as colunas de meta existirem, criar meta
                    if col_mes and col_ano and col_meta:
                        if (
                            pd.notna(row[col_mes]) and
                            pd.notna(row[col_ano]) and
                            pd.notna(row[col_meta])
                        ):
                            try:
                                mes_meta = int(row[col_mes])
                                ano_meta = int(row[col_ano])
                                valor_meta = float(row[col_meta])

                                # Verificar se meta já existe
                                meta_existe = Meta.query.filter_by(
                                    vendedor_id=vendedor.id,
                                    mes=mes_meta,
                                    ano=ano_meta,
                                ).first()

                                if not meta_existe:
                                    meta = Meta(
                                        vendedor_id=vendedor.id,
                                        mes=mes_meta,
                                        ano=ano_meta,
                                        valor_meta=valor_meta,
                                    )
                                    db.session.add(meta)
                            except (ValueError, TypeError):
                                # Ignorar erro de conversão de meta
                                pass

                    sucesso += 1

                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

            if erros:
                db.session.rollback()
                msg_erro = "<br>".join(erros[:10])  # Mostrar até 10 erros
                if len(erros) > 10:
                    msg_erro += f"<br>... e mais {len(erros) - 10} erros"
                flash(f"Erros encontrados:<br>{msg_erro}", "warning")
                # Retornar para página de importação sem redirecionar
                return render_template("vendedores/importar.html")
            elif sucesso > 0:
                db.session.commit()

                # Atualizar metas dos supervisores automaticamente
                atualizar_metas_supervisores()

                msg = f"{sucesso} vendedor(es) importado(s) com sucesso!"
                flash(msg, "success")
                return redirect(url_for("lista_vendedores"))
            else:
                flash(
                    "Nenhum vendedor foi importado. Verifique a planilha.",
                    "warning",
                )
                return redirect(request.url)

        except Exception as e:
            flash(f"Erro ao processar arquivo: {str(e)}", "danger")

    return render_template("vendedores/importar.html")

@app.route("/vendedores/download-template")
@login_required
def download_template_vendedores():
    """Gera e baixa o template Excel para importação de vendedores"""
    # Verificar permissões
    if not pode_importar(current_user, "vendedores"):
        flash("Acesso negado!", "danger")
        return redirect(url_for("lista_vendedores"))

    if not EXCEL_AVAILABLE and not ensure_excel_available():
        flash("Erro: Bibliotecas Excel não instaladas.", "danger")
        return redirect(url_for("importar_vendedores"))

    try:
        import io
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendedores"

        # Definir colunas
        headers = ["Nome", "Email", "Telefone", "CPF"]
        
        # Adicionar coluna de CNPJ se for Super Admin
        if current_user.is_super_admin:
            headers.append("Empresa CNPJ")

        # Estilos
        header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))

        # Escrever cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            
            # Ajustar largura da coluna
            ws.column_dimensions[get_column_letter(col_num)].width = 25

        # Adicionar exemplo
        ws.cell(row=2, column=1, value="Maria Oliveira").border = thin_border
        ws.cell(row=2, column=2, value="maria@email.com").border = thin_border
        ws.cell(row=2, column=3, value="11999999999").border = thin_border
        ws.cell(row=2, column=4, value="123.456.789-00").border = thin_border
        
        if current_user.is_super_admin:
            ws.cell(row=2, column=5, value="00.000.000/0001-00").border = thin_border

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="template_vendedores.xlsx"
        )

    except Exception as e:
        app.logger.error(f"Erro ao gerar template: {str(e)}")
        flash(f"Erro ao gerar template: {str(e)}", "danger")
        return redirect(url_for("importar_vendedores"))

# ===== ROTAS DE FUNCIONÁRIOS (CRUD ADMINISTRATIVO) =====

@app.route("/funcionarios")
@login_required
def lista_funcionarios():
    """Lista todos os usuários/funcionários da empresa (somente Admin)."""
    # Apenas Admin pode gerenciar funcionários
    if current_user.cargo != "admin":
        flash(
            "Acesso negado! Apenas Administradores podem gerenciar funcionários.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    # Super admin enxerga todos os usuários (exceto super admins)
    if current_user.is_super_admin:
        funcionarios = (
            Usuario.query.filter(Usuario.is_super_admin.is_(False))
            .order_by(Usuario.nome)
            .all()
        )
    else:
        # Admin comum vê todos os usuários da própria empresa (exceto super admins)
        funcionarios = (
            Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.is_super_admin.is_(False),
            )
            .order_by(Usuario.nome)
            .all()
        )

    return render_template(
        "funcionarios/lista.html", funcionarios=funcionarios
    )

@app.route("/funcionarios/criar", methods=["GET", "POST"])
@login_required
def criar_funcionario():
    """Criar novo funcionário - Apenas Admin"""
    # Apenas Admin pode criar funcionários
    if current_user.cargo != "admin":
        flash(
            "Acesso negado! Apenas Administradores podem criar funcionários.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    # Buscar hierarquia disponível na empresa para o formulário
    if current_user.is_super_admin:
        admins_disponiveis = (
            Usuario.query.filter(
                Usuario.cargo == "admin", Usuario.ativo.is_(True)
            )
            .order_by(Usuario.nome)
            .all()
        )
        gerentes_disponiveis = (
            Usuario.query.filter(
                Usuario.cargo.in_(["gerente", "gerente_manutencao"]),
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )
        supervisores_disponiveis = (
            Usuario.query.filter(
                Usuario.cargo.in_(["supervisor", "supervisor_manutencao"]),
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )
    else:
        admins_disponiveis = (
            Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.cargo == "admin",
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )
        gerentes_disponiveis = (
            Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.cargo.in_(["gerente", "gerente_manutencao"]),
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )
        supervisores_disponiveis = (
            Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.cargo.in_(["supervisor", "supervisor_manutencao"]),
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )

    if request.method == "POST":
        try:
            nome = request.form.get("nome", "").strip()
            email = request.form.get("email", "").strip().lower()
            cargo = request.form.get("cargo", "").strip()
            departamento = request.form.get("departamento", "").strip()
            senha = request.form.get("senha", "").strip()
            confirmar_senha = request.form.get("confirmar_senha", "").strip()

            # Vínculos hierárquicos (opcionais conforme cargo)
            admin_id_raw = request.form.get("admin_id", "").strip()
            gerente_id_raw = request.form.get("gerente_id", "").strip()
            supervisor_id_raw = request.form.get("supervisor_id", "").strip()

            # Validações
            if not all([nome, email, cargo, senha]):
                flash(
                    "⚠️ Por favor, preencha todos os campos obrigatórios!",
                    "warning",
                )
                return render_template("funcionarios/form.html")

            if len(senha) < 6:
                flash("⚠️ A senha deve ter no mínimo 6 caracteres!", "warning")
                return render_template("funcionarios/form.html")

            if senha != confirmar_senha:
                flash("⚠️ As senhas não conferem!", "warning")
                return render_template("funcionarios/form.html")

            # Verificar se cargo é válido
            cargos_permitidos = [
                "admin",
                "gerente",
                "gerente_manutencao",
                "supervisor",
                "supervisor_manutencao",
                "auxiliar",
                "vendedor",
                "tecnico",
                "financeiro",
                "rh",
                "usuario",
            ]
            if cargo not in cargos_permitidos:
                flash("⚠️ Cargo inválido!", "warning")
                return render_template("funcionarios/form.html")

            # Verificar se email já existe
            if Usuario.query.filter_by(email=email).first():
                flash("⚠️ Este e-mail já está cadastrado!", "warning")
                return render_template(
                    "funcionarios/form.html",
                    funcionario=None,
                    admins_disponiveis=admins_disponiveis,
                    gerentes_disponiveis=gerentes_disponiveis,
                    supervisores_disponiveis=supervisores_disponiveis,
                )

            # Regras de hierarquia
            gerente_id = None
            supervisor_id = None

            # Todo GERENTE (incluindo gerente de manutenção) deve estar vinculado a um ADMIN
            if cargo in ["gerente", "gerente_manutencao"]:
                if not admins_disponiveis:
                    flash(
                        "⚠️ Nenhum administrador disponível para vincular o gerente. Cadastre um administrador primeiro.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                if not admin_id_raw:
                    flash(
                        "⚠️ Selecione o Administrador responsável por este gerente.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                try:
                    admin_id = int(admin_id_raw)
                except ValueError:
                    flash("⚠️ Administrador selecionado inválido.", "warning")
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                admin = Usuario.query.get(admin_id)
                if not admin or admin.cargo != "admin":
                    flash(
                        "⚠️ Administrador responsável não encontrado ou inválido.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                gerente_id = admin.id

            # Todo SUPERVISOR (incluindo supervisor de manutenção) deve estar vinculado a um GERENTE
            if cargo in ["supervisor", "supervisor_manutencao"]:
                if not gerentes_disponiveis:
                    flash(
                        "⚠️ Nenhum gerente disponível para vincular o supervisor. Cadastre um gerente primeiro.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                if not gerente_id_raw:
                    flash(
                        "⚠️ Selecione o Gerente responsável por este supervisor.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                try:
                    gerente_id_val = int(gerente_id_raw)
                except ValueError:
                    flash("⚠️ Gerente selecionado inválido.", "warning")
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                gerente_responsavel = Usuario.query.get(gerente_id_val)
                if not gerente_responsavel or gerente_responsavel.cargo not in [
                    "gerente",
                    "gerente_manutencao",
                ]:
                    flash(
                        "⚠️ Gerente responsável não encontrado ou inválido.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                gerente_id = gerente_responsavel.id

            # Todos VENDEDOR, TÉCNICO, AUXILIAR e USUÁRIO devem estar vinculados a um SUPERVISOR
            cargos_que_precisam_supervisor = [
                "vendedor",
                "tecnico",
                "auxiliar",
                "usuario",
            ]
            if cargo in cargos_que_precisam_supervisor:
                if not supervisores_disponiveis:
                    flash(
                        "⚠️ Nenhum supervisor disponível para vincular este funcionário. Cadastre um supervisor primeiro.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                if not supervisor_id_raw:
                    flash(
                        "⚠️ Selecione o Supervisor responsável por este funcionário.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                try:
                    supervisor_id_val = int(supervisor_id_raw)
                except ValueError:
                    flash("⚠️ Supervisor selecionado inválido.", "warning")
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                supervisor_responsavel = Usuario.query.get(supervisor_id_val)
                if not supervisor_responsavel or supervisor_responsavel.cargo not in [
                    "supervisor",
                    "supervisor_manutencao",
                ]:
                    flash(
                        "⚠️ Supervisor responsável não encontrado ou inválido.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=None,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                supervisor_id = supervisor_responsavel.id

            # Criar funcionário
            funcionario = Usuario(
                nome=nome,
                email=email,
                cargo=cargo,
                departamento=departamento if departamento else None,
                empresa_id=(
                    current_user.empresa_id
                    if not current_user.is_super_admin
                    else None
                ),
                ativo=True,
                gerente_id=gerente_id,
                supervisor_id=supervisor_id,
            )
            funcionario.set_senha(senha)

            db.session.add(funcionario)
            db.session.commit()

            flash(f"Funcionário {nome} cadastrado com sucesso!", "success")
            return redirect(url_for("lista_funcionarios"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar funcionário: {str(e)}", "danger")

    return render_template("funcionarios/form.html", funcionario=None)

@app.route("/funcionarios/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_funcionario(id):
    """Editar funcionário - Apenas Admin"""
    # Apenas Admin pode editar funcionários
    if current_user.cargo != "admin":
        flash(
            "Acesso negado! Apenas Administradores podem editar funcionários.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    funcionario = Usuario.query.get_or_404(id)

    # Verificar se pode editar (mesma empresa)
    if not current_user.is_super_admin:
        if funcionario.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para editar este funcionário!",
                "danger",
            )
            return redirect(url_for("lista_funcionarios"))

    # Buscar hierarquia disponível na empresa para o formulário
    if current_user.is_super_admin:
        admins_disponiveis = (
            Usuario.query.filter(
                Usuario.cargo == "admin", Usuario.ativo.is_(True)
            )
            .order_by(Usuario.nome)
            .all()
        )
        gerentes_disponiveis = (
            Usuario.query.filter(
                Usuario.cargo.in_(["gerente", "gerente_manutencao"]),
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )
        supervisores_disponiveis = (
            Usuario.query.filter(
                Usuario.cargo.in_(["supervisor", "supervisor_manutencao"]),
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )
    else:
        admins_disponiveis = (
            Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.cargo == "admin",
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )
        gerentes_disponiveis = (
            Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.cargo.in_(["gerente", "gerente_manutencao"]),
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )
        supervisores_disponiveis = (
            Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.cargo.in_(["supervisor", "supervisor_manutencao"]),
                Usuario.ativo.is_(True),
            )
            .order_by(Usuario.nome)
            .all()
        )

    if request.method == "POST":
        try:
            nome = request.form.get("nome", "").strip()
            email = request.form.get("email", "").strip().lower()
            cargo = request.form.get("cargo", "").strip()
            departamento = request.form.get("departamento", "").strip() or None

            admin_id_raw = request.form.get("admin_id", "").strip()
            gerente_id_raw = request.form.get("gerente_id", "").strip()
            supervisor_id_raw = request.form.get("supervisor_id", "").strip()

            # Validações
            if not all([nome, email, cargo]):
                flash(
                    "⚠️ Por favor, preencha todos os campos obrigatórios!",
                    "warning",
                )
                return render_template(
                    "funcionarios/form.html",
                    funcionario=funcionario,
                    admins_disponiveis=admins_disponiveis,
                    gerentes_disponiveis=gerentes_disponiveis,
                    supervisores_disponiveis=supervisores_disponiveis,
                )

            # Verificar se cargo é válido
            cargos_permitidos = [
                "admin",
                "gerente",
                "gerente_manutencao",
                "supervisor",
                "supervisor_manutencao",
                "auxiliar",
                "vendedor",
                "tecnico",
                "financeiro",
                "rh",
                "usuario",
            ]
            if cargo not in cargos_permitidos:
                flash("⚠️ Cargo inválido!", "warning")
                return render_template(
                    "funcionarios/form.html",
                    funcionario=funcionario,
                    admins_disponiveis=admins_disponiveis,
                    gerentes_disponiveis=gerentes_disponiveis,
                    supervisores_disponiveis=supervisores_disponiveis,
                )

            # Regras de hierarquia
            gerente_id = funcionario.gerente_id
            supervisor_id = funcionario.supervisor_id

            # Todo GERENTE (incluindo gerente de manutenção) deve estar vinculado a um ADMIN
            if cargo in ["gerente", "gerente_manutencao"]:
                if not admins_disponiveis:
                    flash(
                        "⚠️ Nenhum administrador disponível para vincular o gerente. Cadastre um administrador primeiro.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                if not admin_id_raw:
                    flash(
                        "⚠️ Selecione o Administrador responsável por este gerente.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                try:
                    admin_id = int(admin_id_raw)
                except ValueError:
                    flash("⚠️ Administrador selecionado inválido.", "warning")
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                admin = Usuario.query.get(admin_id)
                if not admin or admin.cargo != "admin":
                    flash(
                        "⚠️ Administrador responsável não encontrado ou inválido.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                gerente_id = admin.id

            # Todo SUPERVISOR (incluindo supervisor de manutenção) deve estar vinculado a um GERENTE
            if cargo in ["supervisor", "supervisor_manutencao"]:
                if not gerentes_disponiveis:
                    flash(
                        "⚠️ Nenhum gerente disponível para vincular o supervisor. Cadastre um gerente primeiro.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                if not gerente_id_raw:
                    flash(
                        "⚠️ Selecione o Gerente responsável por este supervisor.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                try:
                    gerente_id_val = int(gerente_id_raw)
                except ValueError:
                    flash("⚠️ Gerente selecionado inválido.", "warning")
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                gerente_responsavel = Usuario.query.get(gerente_id_val)
                if not gerente_responsavel or gerente_responsavel.cargo not in [
                    "gerente",
                    "gerente_manutencao",
                ]:
                    flash(
                        "⚠️ Gerente responsável não encontrado ou inválido.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                gerente_id = gerente_responsavel.id

            # Todos VENDEDOR, TÉCNICO, AUXILIAR e USUÁRIO devem estar vinculados a um SUPERVISOR
            cargos_que_precisam_supervisor = [
                "vendedor",
                "tecnico",
                "auxiliar",
                "usuario",
            ]
            if cargo in cargos_que_precisam_supervisor:
                if not supervisores_disponiveis:
                    flash(
                        "⚠️ Nenhum supervisor disponível para vincular este funcionário. Cadastre um supervisor primeiro.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                if not supervisor_id_raw:
                    flash(
                        "⚠️ Selecione o Supervisor responsável por este funcionário.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                try:
                    supervisor_id_val = int(supervisor_id_raw)
                except ValueError:
                    flash("⚠️ Supervisor selecionado inválido.", "warning")
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                supervisor_responsavel = Usuario.query.get(supervisor_id_val)
                if not supervisor_responsavel or supervisor_responsavel.cargo not in [
                    "supervisor",
                    "supervisor_manutencao",
                ]:
                    flash(
                        "⚠️ Supervisor responsável não encontrado ou inválido.",
                        "warning",
                    )
                    return render_template(
                        "funcionarios/form.html",
                        funcionario=funcionario,
                        admins_disponiveis=admins_disponiveis,
                        gerentes_disponiveis=gerentes_disponiveis,
                        supervisores_disponiveis=supervisores_disponiveis,
                    )

                supervisor_id = supervisor_responsavel.id
            elif cargo not in ["gerente", "supervisor"]:
                # Para outros cargos, vínculo com supervisor é opcional
                supervisor_id = supervisor_id or None

            # Atualizar dados
            funcionario.nome = nome
            funcionario.email = email
            funcionario.cargo = cargo
            funcionario.departamento = departamento
            funcionario.gerente_id = gerente_id
            funcionario.supervisor_id = supervisor_id

            # Verificar se quer alterar senha
            nova_senha = request.form.get("senha", "").strip()
            if nova_senha:
                confirmar_senha = request.form.get(
                    "confirmar_senha", ""
                ).strip()
                if len(nova_senha) < 6:
                    flash(
                        "⚠️ A senha deve ter no mínimo 6 caracteres!", "warning"
                    )
                    return render_template(
                        "funcionarios/form.html", funcionario=funcionario
                    )
                if nova_senha != confirmar_senha:
                    flash("⚠️ As senhas não conferem!", "warning")
                    return render_template(
                        "funcionarios/form.html", funcionario=funcionario
                    )
                funcionario.set_senha(nova_senha)

            db.session.commit()
            flash(f"Funcionário {funcionario.nome} atualizado com sucesso!", "success")
            return redirect(url_for("lista_funcionarios"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar funcionário: {str(e)}", "danger")

    return render_template(
        "funcionarios/form.html",
        funcionario=funcionario,
        admins_disponiveis=admins_disponiveis,
        gerentes_disponiveis=gerentes_disponiveis,
        supervisores_disponiveis=supervisores_disponiveis,
    )

@app.route("/funcionarios/<int:id>/deletar", methods=["POST"])
@login_required
def deletar_funcionario(id):
    """Deletar funcionário - Apenas Admin"""
    # Apenas Admin pode deletar funcionários
    if current_user.cargo != "admin":
        flash(
            "Acesso negado! Apenas Administradores podem deletar funcionários.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    funcionario = Usuario.query.get_or_404(id)

    # Verificar se pode deletar (mesma empresa)
    if not current_user.is_super_admin:
        if funcionario.empresa_id != current_user.empresa_id:
            flash(
                "Você não tem permissão para deletar este funcionário!",
                "danger",
            )
            return redirect(url_for("lista_funcionarios"))

    # Não pode deletar a si mesmo
    if funcionario.id == current_user.id:
        flash("⚠️ Você não pode deletar seu próprio usuário!", "warning")
        return redirect(url_for("lista_funcionarios"))

    try:
        nome = funcionario.nome
        db.session.delete(funcionario)
        db.session.commit()
        flash(f"Funcionário {nome} removido com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao remover funcionário: {str(e)}", "danger")

    return redirect(url_for("lista_funcionarios"))

@app.route("/funcionarios/<int:id>/ativar-desativar", methods=["POST"])
@login_required
def ativar_desativar_funcionario(id):
    """Ativar/Desativar funcionário - Apenas Admin"""
    # Apenas Admin pode ativar/desativar funcionários
    if current_user.cargo != "admin":
        flash("Acesso negado!", "danger")
        return redirect(url_for("dashboard"))

    funcionario = Usuario.query.get_or_404(id)

    # Verificar permissões
    if not current_user.is_super_admin:
        if funcionario.empresa_id != current_user.empresa_id:
            flash("Você não tem permissão!", "danger")
            return redirect(url_for("lista_funcionarios"))

    try:
        funcionario.ativo = not funcionario.ativo
        db.session.commit()

        status = "ativado" if funcionario.ativo else "desativado"
        flash(f"Funcionário {funcionario.nome} {status} com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro: {str(e)}", "danger")

    return redirect(url_for("lista_funcionarios"))

# ===== ROTAS DE CLIENTES =====

@app.route("/clientes")
@login_required
def lista_clientes():
    """Lista todos os clientes com paginação"""
    # Parâmetros de paginação
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get(
        "per_page", 10, type=int
    )  # 10 clientes por página (melhor desempenho e leitura)

    # Obter filtros da query string
    status_filtro = request.args.get("status", "todos")
    cidade_filtro = request.args.get("cidade", "")
    bairro_filtro = request.args.get("bairro", "")
    dia_visita_filtro = request.args.get("dia_visita", "")
    busca = request.args.get("busca", "")

    # Base query com lazy loading otimizado
    if current_user.cargo == "vendedor" and current_user.vendedor_id:
        # Vendedor vê apenas seus clientes
        query = Cliente.query.filter_by(
            vendedor_id=current_user.vendedor_id, ativo=True
        )
    elif current_user.cargo == "supervisor":
        # Supervisor vê clientes de seus vendedores
        vendedores_ids = [v.id for v in current_user.vendedores]
        query = Cliente.query.filter(
            Cliente.vendedor_id.in_(vendedores_ids), Cliente.ativo
        )
    elif current_user.is_super_admin:
        query = Cliente.query.filter_by(ativo=True)
    else:
        # Admin da empresa vê todos clientes da empresa
        query = Cliente.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        )

    # Aplicar busca por nome, CPF ou CNPJ
    if busca:
        busca_clean = busca.strip()
        query = query.filter(
            db.or_(
                Cliente.nome.ilike(f"%{busca_clean}%"),
                Cliente.cpf.ilike(f"%{busca_clean}%"),
                Cliente.cnpj.ilike(f"%{busca_clean}%"),
                Cliente.telefone.ilike(f"%{busca_clean}%"),
            )
        )

    # Aplicar filtros
    if cidade_filtro:
        query = query.filter(Cliente.cidade.ilike(f"%{cidade_filtro}%"))

    if bairro_filtro:
        query = query.filter(Cliente.bairro.ilike(f"%{bairro_filtro}%"))

    if dia_visita_filtro:
        query = query.filter_by(dia_visita=dia_visita_filtro)

    # Ordenar por nome
    query = query.order_by(Cliente.nome)

    # Paginar resultados
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    clientes = pagination.items

    # Filtrar por status (cores) apenas nos resultados paginados
    if status_filtro != "todos":
        clientes = [c for c in clientes if c.get_status_cor() == status_filtro]

    # Estatísticas (otimizadas - contar no banco)
    total_clientes = pagination.total

    # Para estatísticas de status, usamos a página atual
    clientes_verde = len(
        [c for c in clientes if c.get_status_cor() == "verde"]
    )
    clientes_amarelo = len(
        [c for c in clientes if c.get_status_cor() == "amarelo"]
    )
    clientes_vermelho = len(
        [c for c in clientes if c.get_status_cor() == "vermelho"]
    )

    return render_template(
        "clientes/lista.html",
        clientes=clientes,
        pagination=pagination,
        total_clientes=total_clientes,
        clientes_verde=clientes_verde,
        clientes_amarelo=clientes_amarelo,
        clientes_vermelho=clientes_vermelho,
        status_filtro=status_filtro,
        cidade_filtro=cidade_filtro,
        bairro_filtro=bairro_filtro,
        dia_visita_filtro=dia_visita_filtro,
        busca=busca,
    )

@app.route("/clientes/novo", methods=["GET", "POST"])
@login_required
@permission_required("pode_criar_cliente")
def novo_cliente():
    """Cadastrar novo cliente"""
    try:
        # Criar formulário com empresa_id para popular dropdowns automaticamente
        form = ClienteForm(empresa_id=current_user.empresa_id)
        
        # Se o usuário for vendedor, pré-selecionar seu ID
        if request.method == "GET" and current_user.cargo == "vendedor":
            vendedor_atual = Vendedor.query.filter_by(usuario_id=current_user.id).first()
            if vendedor_atual:
                form.vendedor_id.data = vendedor_atual.id
                
    except Exception as e:
        app.logger.error(f"Erro ao inicializar formulário de cliente: {str(e)}")
        flash("Erro ao carregar formulário. Tente novamente.", "danger")
        return redirect(url_for("lista_clientes"))

    if form.validate_on_submit():
        try:

            # Determinar vendedor_id
            vendedor_id = form.vendedor_id.data

            # Limpar CPF, CNPJ e telefones
            cpf_limpo = (
                re.sub(r"\D", "", form.cpf.data) if form.cpf.data else None
            )
            cnpj_limpo = (
                re.sub(r"\D", "", form.cnpj.data) if form.cnpj.data else None
            )
            telefone_limpo = (
                re.sub(r"\D", "", form.telefone.data)
                if form.telefone.data
                else None
            )
            telefone2_limpo = (
                re.sub(r"\D", "", form.telefone2.data)
                if form.telefone2.data
                else None
            )
            celular_limpo = (
                re.sub(r"\D", "", form.celular.data)
                if form.celular.data
                else None
            )
            cep_limpo = (
                re.sub(r"\D", "", form.cep.data) if form.cep.data else None
            )

            # Gerar código único do cliente
            municipio = (
                form.municipio.data.strip()
                if form.municipio.data
                else (
                    form.cidade.data.strip()
                    if form.cidade.data
                    else "SEM_CIDADE"
                )
            )
            # Gerar código único do cliente com verificação preventiva
            codigo_cliente = Cliente.gerar_codigo_cliente(
                municipio, current_user.empresa_id
            )
            # Evitar colisões por concorrência ou dados anteriores
            try_count = 0
            while True:
                existente = Cliente.query.filter_by(
                    empresa_id=current_user.empresa_id,
                    codigo_cliente=codigo_cliente,
                ).first()
                if not existente:
                    break
                try_count += 1
                if try_count > 5:
                    raise Exception(
                        "Falha ao gerar código único para cliente após múltiplas tentativas"
                    )
                # Recalcular usando estado atual do banco
                codigo_cliente = Cliente.gerar_codigo_cliente(
                    municipio, current_user.empresa_id
                )

            # Criar cliente
            cliente = Cliente(
                vendedor_id=vendedor_id,
                codigo_cliente=codigo_cliente,
                nome=form.nome.data.strip() if form.nome.data else None,
                razao_social=(
                    form.razao_social.data.strip()
                    if form.razao_social.data
                    else None
                ),
                sigla=form.sigla.data.strip() if form.sigla.data else None,
                cpf=cpf_limpo,
                cnpj=cnpj_limpo,
                inscricao_estadual=(
                    form.inscricao_estadual.data.strip()
                    if form.inscricao_estadual.data
                    else None
                ),
                codigo_bp=(
                    form.codigo_bp.data.strip()
                    if form.codigo_bp.data
                    else None
                ),
                logradouro=(
                    form.logradouro.data.strip()
                    if form.logradouro.data
                    else None
                ),
                municipio=municipio,
                cep=cep_limpo,
                bairro=form.bairro.data.strip() if form.bairro.data else None,
                cidade=municipio,  # Mantém compatibilidade
                ponto_referencia=(
                    form.ponto_referencia.data.strip()
                    if form.ponto_referencia.data
                    else None
                ),
                coordenada_x=(
                    form.coordenada_x.data.strip()
                    if form.coordenada_x.data
                    else None
                ),
                coordenada_y=(
                    form.coordenada_y.data.strip()
                    if form.coordenada_y.data
                    else None
                ),
                telefone=telefone_limpo,
                telefone2=telefone2_limpo,
                celular=celular_limpo,
                email=(
                    form.email.data.strip().lower()
                    if form.email.data
                    else None
                ),
                dia_visita=(
                    form.dia_visita.data if form.dia_visita.data else None
                ),
                observacoes=(
                    form.observacoes.data.strip()
                    if form.observacoes.data
                    else None
                ),
                empresa_id=current_user.empresa_id,
                ativo=form.ativo.data if form.ativo.data is not None else True,
            )

            # Configurar formas de pagamento
            if form.formas_pagamento.data:
                cliente.set_formas_pagamento_list(form.formas_pagamento.data)
            else:
                cliente.set_formas_pagamento_list([])

            db.session.add(cliente)
            try:
                db.session.commit()
            except Exception as e:
                # Em caso de violação de unicidade, tentar novamente uma vez
                from sqlalchemy.exc import IntegrityError
                db.session.rollback()
                if isinstance(e, IntegrityError):
                    codigo_cliente = Cliente.gerar_codigo_cliente(
                        municipio, current_user.empresa_id
                    )
                    cliente.codigo_cliente = codigo_cliente
                    db.session.add(cliente)
                    db.session.commit()
                else:
                    raise

            flash("Cliente cadastrado com sucesso!", "success")
            return redirect(url_for("lista_clientes"))

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Erro ao cadastrar cliente: {str(e)}")
            flash(f"Erro ao cadastrar cliente: {str(e)}", "danger")

    return render_template(
        "clientes/form.html", form=form, titulo="Novo Cliente", cliente=None
    )

@app.route("/clientes/<int:id>")
@login_required
def ver_cliente(id):
    """Ver detalhes do cliente"""
    cliente = Cliente.query.get_or_404(id)

    # Verificar permissão
    if not verificar_acesso_cliente(cliente):
        flash("Você não tem permissão para acessar este cliente.", "danger")
        return redirect(url_for("lista_clientes"))

    # Buscar histórico de compras
    compras = (
        cliente.compras.order_by(CompraCliente.data_compra.desc())
        .limit(20)
        .all()
    )

    # Estatísticas do cliente
    total_compras = cliente.compras.count()
    total_compras_mes = cliente.get_total_compras_mes()
    agora = datetime.utcnow()

    return render_template(
        "clientes/ver.html",
        cliente=cliente,
        compras=compras,
        total_compras=total_compras,
        total_compras_mes=total_compras_mes,
        now=agora,
    )

@app.route("/clientes/<int:id>/editar", methods=["GET", "POST"])
@login_required
@permission_required("pode_editar_cliente")
def editar_cliente(id):
    """Editar cliente existente"""
    cliente = Cliente.query.get_or_404(id)

    # Verificar permissão
    if not verificar_acesso_cliente(cliente):
        flash("Você não tem permissão para editar este cliente.", "danger")
        return redirect(url_for("lista_clientes"))

    form = ClienteForm(
        cliente_id=cliente.id, empresa_id=current_user.empresa_id, obj=cliente
    )

    # Pré-preencher formas de pagamento e vendedor
    if request.method == "GET":
        form.formas_pagamento.data = cliente.get_formas_pagamento_list()
        form.vendedor_id.data = cliente.vendedor_id

        # Preencher supervisor (relacionamento correto supervisor_obj)
        if cliente.vendedor and getattr(cliente.vendedor, "supervisor_obj", None):
            form.supervisor_nome.data = cliente.vendedor.supervisor_obj.nome

    if form.validate_on_submit():
        try:

            # Atualizar vendedor
            cliente.vendedor_id = form.vendedor_id.data

            # Atualizar dados básicos
            cliente.nome = form.nome.data.strip() if form.nome.data else None
            cliente.razao_social = (
                form.razao_social.data.strip()
                if form.razao_social.data
                else None
            )
            cliente.sigla = (
                form.sigla.data.strip() if form.sigla.data else None
            )
            cliente.cpf = (
                re.sub(r"\D", "", form.cpf.data) if form.cpf.data else None
            )
            cliente.cnpj = (
                re.sub(r"\D", "", form.cnpj.data) if form.cnpj.data else None
            )
            cliente.inscricao_estadual = (
                form.inscricao_estadual.data.strip()
                if form.inscricao_estadual.data
                else None
            )
            cliente.codigo_bp = (
                form.codigo_bp.data.strip() if form.codigo_bp.data else None
            )

            # Atualizar endereço
            cliente.logradouro = (
                form.logradouro.data.strip() if form.logradouro.data else None
            )
            municipio = (
                form.municipio.data.strip()
                if form.municipio.data
                else (form.cidade.data.strip() if form.cidade.data else None)
            )
            cliente.municipio = municipio
            cliente.cidade = municipio  # Mantém compatibilidade
            cliente.cep = (
                re.sub(r"\D", "", form.cep.data) if form.cep.data else None
            )
            cliente.bairro = (
                form.bairro.data.strip() if form.bairro.data else None
            )
            cliente.ponto_referencia = (
                form.ponto_referencia.data.strip()
                if form.ponto_referencia.data
                else None
            )
            cliente.coordenada_x = (
                form.coordenada_x.data.strip()
                if form.coordenada_x.data
                else None
            )
            cliente.coordenada_y = (
                form.coordenada_y.data.strip()
                if form.coordenada_y.data
                else None
            )

            # Atualizar contato
            cliente.telefone = (
                re.sub(r"\D", "", form.telefone.data)
                if form.telefone.data
                else None
            )
            cliente.telefone2 = (
                re.sub(r"\D", "", form.telefone2.data)
                if form.telefone2.data
                else None
            )
            cliente.celular = (
                re.sub(r"\D", "", form.celular.data)
                if form.celular.data
                else None
            )
            cliente.email = (
                form.email.data.strip().lower() if form.email.data else None
            )

            # Atualizar informações comerciais
            cliente.dia_visita = (
                form.dia_visita.data if form.dia_visita.data else None
            )
            cliente.observacoes = (
                form.observacoes.data.strip()
                if form.observacoes.data
                else None
            )
            cliente.ativo = (
                form.ativo.data if form.ativo.data is not None else True
            )

            # Atualizar formas de pagamento
            if form.formas_pagamento.data:
                cliente.set_formas_pagamento_list(form.formas_pagamento.data)
            else:
                cliente.set_formas_pagamento_list([])

            db.session.commit()

            flash("Cliente atualizado com sucesso!", "success")
            return redirect(url_for("ver_cliente", id=cliente.id))

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Erro ao atualizar cliente {id}: {str(e)}")
            flash(f"Erro ao atualizar cliente: {str(e)}", "danger")

    return render_template(
        "clientes/form.html",
        form=form,
        titulo="Editar Cliente",
        cliente=cliente,
    )

@app.route("/clientes/<int:id>/deletar", methods=["POST"])
@login_required
def deletar_cliente(id):
    """Desativar cliente (soft delete)"""
    cliente = Cliente.query.get_or_404(id)

    # Verificar permissão
    if not verificar_acesso_cliente(cliente):
        flash("Você não tem permissão para deletar este cliente.", "danger")
        return redirect(url_for("lista_clientes"))

    try:
        cliente.ativo = False
        db.session.commit()

        flash("Cliente removido com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao remover cliente: {str(e)}", "danger")

    return redirect(url_for("lista_clientes"))

@app.route("/clientes/<int:id>/compra", methods=["GET", "POST"])
@login_required
def registrar_compra(id):
    """Registrar compra do cliente"""
    cliente = Cliente.query.get_or_404(id)

    # Produtos de estoque disponíveis para venda (mesma empresa)
    # Exibe apenas os 10 mais vendidos para simplificar a escolha do vendedor
    from sqlalchemy import func, desc

    vendas_subq = (
        db.session.query(
            EstoqueMovimento.produto_id,
            func.sum(EstoqueMovimento.quantidade).label("total_vendido"),
        )
        .filter(
            EstoqueMovimento.empresa_id == cliente.empresa_id,
            EstoqueMovimento.tipo == "saida",
            EstoqueMovimento.motivo == "venda",
        )
        .group_by(EstoqueMovimento.produto_id)
        .subquery()
    )

    produtos = (
        Produto.query.filter_by(empresa_id=cliente.empresa_id, ativo=True)
        .outerjoin(vendas_subq, Produto.id == vendas_subq.c.produto_id)
        .order_by(desc(vendas_subq.c.total_vendido), Produto.nome)
        .limit(10)
        .all()
    )

    # Verificar permissão
    if not verificar_acesso_cliente(cliente):
        flash(
            "Você não tem permissão para registrar compras para este cliente.",
            "danger",
        )
        return redirect(url_for("lista_clientes"))

    form = CompraClienteForm()
    # Garante que o campo cliente_id tenha uma opção válida para evitar erros
    # de validação do WTForms ("Not a valid choice") durante o submit
    form.cliente_id.choices = [(cliente.id, cliente.nome)]
    form.cliente_id.data = cliente.id

    if form.validate_on_submit():
        try:
            # Verificar se pode comprar no mês
            if not cliente.pode_comprar_no_mes():
                flash(
                    "Cliente já atingiu o limite de compras para este mês.",
                    "warning",
                )
                return redirect(url_for("ver_cliente", id=cliente.id))

            # Mapear itens selecionados a partir dos campos produto_<id>
            itens_venda = []
            valor_total_itens = 0.0

            for key, value in request.form.items():
                if not key.startswith("produto_"):
                    continue

                try:
                    produto_id = int(key.split("_", 1)[1])
                except (ValueError, IndexError):
                    continue

                try:
                    qtd = float((value or "0").replace(",", "."))
                except ValueError:
                    qtd = 0

                if qtd <= 0:
                    continue

                produto = (
                    Produto.query.filter_by(
                        id=produto_id,
                        empresa_id=cliente.empresa_id,
                        ativo=True,
                    )
                    .with_for_update()
                    .first()
                )

                if not produto:
                    continue

                estoque_disponivel = float(produto.estoque_atual or 0)
                if estoque_disponivel <= 0:
                    continue

                quantidade_utilizada = min(qtd, estoque_disponivel)
                if quantidade_utilizada <= 0:
                    continue

                preco_unitario = float(produto.preco_venda or 0)
                subtotal = quantidade_utilizada * preco_unitario
                valor_total_itens += subtotal

                itens_venda.append(
                    {
                        "produto": produto,
                        "quantidade": quantidade_utilizada,
                        "preco_unitario": preco_unitario,
                        "subtotal": subtotal,
                    }
                )

            # Se nenhum item foi selecionado, não é possível registrar a venda
            if not itens_venda:
                flash(
                    "Selecione pelo menos um produto para registrar a venda.",
                    "warning",
                )
                return render_template(
                    "clientes/compra.html",
                    form=form,
                    cliente=cliente,
                    produtos=produtos,
                )

            # Valor da compra passa a ser obrigatoriamente o total dos itens
            valor_compra = valor_total_itens

            # Criar compra
            compra = CompraCliente(
                cliente_id=cliente.id,
                vendedor_id=cliente.vendedor_id,
                empresa_id=cliente.empresa_id,
                valor=valor_compra,
                forma_pagamento=form.forma_pagamento.data,
                observacoes=form.observacoes.data,
                data_compra=datetime.utcnow(),
            )

            db.session.add(compra)
            db.session.flush()  # garante ID para relacionar nos movimentos de estoque

            # Registrar movimentos de estoque para cada item da venda
            for item in itens_venda:
                produto = item["produto"]
                quantidade = item["quantidade"]
                preco_unitario = item["preco_unitario"]
                subtotal = item["subtotal"]

                # Atualizar estoque atual do produto
                estoque_atual = float(produto.estoque_atual or 0)
                novo_estoque = max(0.0, estoque_atual - quantidade)
                produto.estoque_atual = novo_estoque

                movimento = EstoqueMovimento(
                    produto_id=produto.id,
                    tipo="saida",
                    motivo="venda",
                    quantidade=quantidade,
                    valor_unitario=preco_unitario,
                    valor_total=subtotal,
                    documento=f"Venda #{compra.id}",
                    observacoes=form.observacoes.data,
                    cliente_id=cliente.id,
                    usuario_id=current_user.id,
                    empresa_id=cliente.empresa_id,
                )

                db.session.add(movimento)

            # Atualizar meta do vendedor com o valor total da compra
            # utilizando a mesma data da compra para encontrar a meta do mês
            try:
                _update_meta_for_compra(
                    cliente.vendedor_id,
                    valor_compra,
                    compra.data_compra,
                )
            except Exception as e:
                # Em caso de erro na atualização de meta, registramos log
                # mas não impedimos o registro da venda/estoque.
                app.logger.error(
                    f"Erro ao atualizar meta para venda {compra.id}: {e}"
                )

            db.session.commit()

            flash("Compra registrada com sucesso!", "success")
            return redirect(url_for("ver_cliente", id=cliente.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao registrar compra: {str(e)}", "danger")

    return render_template(
        "clientes/compra.html",
        form=form,
        cliente=cliente,
        produtos=produtos,
    )

@app.route("/clientes/relatorio")
@login_required
def relatorio_clientes():
    """Relatório de clientes por vendedor/supervisor"""
    
    # CACHE: Cache key baseado em usuário (10 minutos, dados atualizam menos)
    cache_key = f"relatorio_clientes_{current_user.id}"
    
    if cache:
        cached_result = cache.get(cache_key)
        if cached_result:
            return cached_result
    
    # Determinar escopo
    if current_user.cargo == "vendedor" and current_user.vendedor_id:
        # Relatório do vendedor
        vendedor = Vendedor.query.get(current_user.vendedor_id)
        vendedores = [vendedor]
        titulo = f"Relatório de Clientes - {vendedor.nome}"
    elif current_user.cargo == "supervisor":
        # Relatório de todos vendedores do supervisor
        vendedores = current_user.vendedores
        titulo = "Relatório de Clientes - Supervisão"
    elif current_user.is_super_admin:
        # Todos vendedores
        vendedores = Vendedor.query.filter_by(ativo=True).all()
        titulo = "Relatório Geral de Clientes"
    else:
        # Admin da empresa
        vendedores = Vendedor.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        ).all()
        titulo = "Relatório de Clientes da Empresa"

    # Compilar dados
    relatorio = []
    for vendedor in vendedores:
        clientes = Cliente.query.filter_by(
            vendedor_id=vendedor.id, ativo=True
        ).all()

        total = len(clientes)
        verde = len([c for c in clientes if c.get_status_cor() == "verde"])
        amarelo = len([c for c in clientes if c.get_status_cor() == "amarelo"])
        vermelho = len(
            [c for c in clientes if c.get_status_cor() == "vermelho"]
        )

        relatorio.append(
            {
                "vendedor": vendedor,
                "total": total,
                "verde": verde,
                "amarelo": amarelo,
                "vermelho": vermelho,
                "percentual_positivado": (
                    (verde / total * 100) if total > 0 else 0
                ),
            }
        )

    result = render_template(
        "clientes/relatorio.html", relatorio=relatorio, titulo=titulo
    )
    
    # Armazenar no cache (10 minutos)
    if cache:
        cache.set(cache_key, result, timeout=600)
    
    return result

@app.route("/clientes/relatorio-vendas")
@login_required
def relatorio_vendas():
    """Relatório detalhado de vendas por cliente"""
    from sqlalchemy import extract

    # Filtros
    ano = request.args.get("ano", datetime.now().year, type=int)
    vendedor_id = request.args.get("vendedor_id", type=int)
    supervisor_id = request.args.get("supervisor_id", type=int)
    cidade = request.args.get("cidade", "")
    bairro = request.args.get("bairro", "")
    status = request.args.get("status", "")

    # Query base - Join com Usuario (supervisor) via alias
    SupervisorAlias = db.aliased(Usuario)

    query = (
        db.session.query(
            Cliente.id,
            Cliente.cpf,
            Cliente.cnpj,
            Cliente.nome,
            Cliente.cidade,
            Cliente.bairro,
            Vendedor.nome.label("vendedor_nome"),
            SupervisorAlias.nome.label("supervisor_nome"),
        )
        .join(Vendedor, Cliente.vendedor_id == Vendedor.id)
        .outerjoin(
            SupervisorAlias, Vendedor.supervisor_id == SupervisorAlias.id
        )
    )

    # Aplicar filtros de permissão
    if current_user.cargo == "vendedor" and current_user.vendedor_id:
        query = query.filter(Cliente.vendedor_id == current_user.vendedor_id)
    elif current_user.cargo == "supervisor":
        # Buscar vendedores deste supervisor
        vendedores_ids = [
            v.id
            for v in Vendedor.query.filter_by(
                supervisor_id=current_user.id, ativo=True
            ).all()
        ]
        if vendedores_ids:
            query = query.filter(Cliente.vendedor_id.in_(vendedores_ids))
        else:
            # Retorna vazio se não tiver vendedores
            query = query.filter(Cliente.id == -1)
    elif not current_user.is_super_admin:
        query = query.filter(Cliente.empresa_id == current_user.empresa_id)

    # Aplicar filtros adicionais
    if vendedor_id:
        query = query.filter(Cliente.vendedor_id == vendedor_id)
    if supervisor_id:
        query = query.filter(Vendedor.supervisor_id == supervisor_id)
    if cidade:
        query = query.filter(Cliente.cidade.ilike(f"%{cidade}%"))
    if bairro:
        query = query.filter(Cliente.bairro.ilike(f"%{bairro}%"))

    query = query.filter(Cliente.ativo)

    # Executar query principal (clientes base)
    clientes_base = query.all()

    # Se não houver clientes, retorna relatório vazio rapidamente
    if not clientes_base:
        relatorio_dados = []
        total_compras_geral = 0
        total_valor_geral = 0
    else:
        # Otimização: buscar compras em lote para evitar N+1 queries
        cliente_ids = [row.id for row in clientes_base]

        # Compras do ano para todos os clientes filtrados
        compras_ano_todos = CompraCliente.query.filter(
            CompraCliente.cliente_id.in_(cliente_ids),
            extract("year", CompraCliente.data_compra) == ano,
        ).all()

        from collections import defaultdict

        compras_ano_por_cliente = defaultdict(list)
        for compra in compras_ano_todos:
            compras_ano_por_cliente[compra.cliente_id].append(compra)

        # Última compra (todas as datas) por cliente, via agregação
        ultima_compra_rows = (
            db.session.query(
                CompraCliente.cliente_id,
                db.func.max(CompraCliente.data_compra).label("ultima_data"),
            )
            .filter(CompraCliente.cliente_id.in_(cliente_ids))
            .group_by(CompraCliente.cliente_id)
            .all()
        )
        ultima_compra_por_cliente = {
            row.cliente_id: row.ultima_data for row in ultima_compra_rows
        }

        # Buscar objetos Cliente em lote
        clientes_objs = Cliente.query.filter(Cliente.id.in_(cliente_ids)).all()
        clientes_map = {c.id: c for c in clientes_objs}

        # Processar dados detalhados
        relatorio_dados = []
        total_compras_geral = 0
        total_valor_geral = 0

        for cliente_row in clientes_base:
            cliente = clientes_map.get(cliente_row.id)
            if not cliente:
                continue

            compras_ano = compras_ano_por_cliente.get(cliente.id, [])

            total_compras_ano = len(compras_ano)
            valor_total_ano = (
                sum(c.valor for c in compras_ano) if compras_ano else 0
            )

            # Última compra (qualquer ano)
            ultima_data = ultima_compra_por_cliente.get(cliente.id)

            # Formas de pagamento usadas (apenas no ano filtrado)
            formas_usadas = {
                c.forma_pagamento
                for c in compras_ano
                if c.forma_pagamento
            }

            # Status replicando lógica de Cliente.get_status_cor
            if not ultima_data:
                status_cor = "vermelho"
                dias_sem_compra = None
            else:
                dias_sem_compra = (datetime.now() - ultima_data).days
                if dias_sem_compra <= 30:
                    status_cor = "verde"
                elif dias_sem_compra <= 38:
                    status_cor = "amarelo"
                else:
                    status_cor = "vermelho"

            # Filtrar por status se necessário
            if status and status_cor != status:
                continue

            relatorio_dados.append(
                {
                    "id": cliente.id,
                    "cpf": cliente.cpf or "-",
                    "cnpj": cliente.cnpj or "-",
                    "nome": cliente.nome,
                    "cidade": cliente.cidade or "-",
                    "bairro": cliente.bairro or "-",
                    "vendedor": cliente_row.vendedor_nome,
                    "supervisor": cliente_row.supervisor_nome or "-",
                    "compras_ano": total_compras_ano,
                    "valor_total_ano": valor_total_ano,
                    "ultima_compra": ultima_data,
                    "formas_pagamento": (
                        ", ".join(formas_usadas) if formas_usadas else "-"
                    ),
                    "status": status_cor,
                    "dias_sem_compra": dias_sem_compra,
                }
            )

            total_compras_geral += total_compras_ano
            total_valor_geral += valor_total_ano

    # Ordenar por compras do ano (decrescente)
    relatorio_dados.sort(key=lambda x: x["compras_ano"], reverse=True)

    # Buscar listas para filtros
    if current_user.cargo == "vendedor":
        vendedores = [Vendedor.query.get(current_user.vendedor_id)]
    elif current_user.cargo == "supervisor":
        vendedores = Vendedor.query.filter_by(
            supervisor_id=current_user.id, ativo=True
        ).all()
    elif current_user.is_super_admin:
        vendedores = Vendedor.query.filter_by(ativo=True).all()
    else:
        vendedores = Vendedor.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        ).all()

    # Buscar supervisores (usuários com cargo='supervisor')
    supervisores = Usuario.query.filter_by(
        cargo="supervisor", ativo=True
    ).all()

    # Buscar cidades e bairros únicos
    cidades = (
        db.session.query(Cliente.cidade)
        .filter(Cliente.cidade.isnot(None))
        .distinct()
        .all()
    )
    cidades = [c[0] for c in cidades if c[0]]

    bairros = (
        db.session.query(Cliente.bairro)
        .filter(Cliente.bairro.isnot(None))
        .distinct()
        .all()
    )
    bairros = [b[0] for b in bairros if b[0]]

    return render_template(
        "clientes/relatorio_vendas.html",
        relatorio=relatorio_dados,
        ano=ano,
        vendedores=vendedores,
        supervisores=supervisores,
        cidades=cidades,
        bairros=bairros,
        total_clientes=len(relatorio_dados),
        total_compras=total_compras_geral,
        total_valor=total_valor_geral,
        filtros={
            "vendedor_id": vendedor_id,
            "supervisor_id": supervisor_id,
            "cidade": cidade,
            "bairro": bairro,
            "status": status,
        },
    )


@app.route("/clientes/relatorio-vendas/exportar")
@login_required
def exportar_relatorio_vendas():
    """Exportar Relatório de Vendas por Cliente para Excel.

    Usa exatamente os mesmos filtros e a mesma lógica do relatório em tela,
    gerando uma planilha com as colunas principais de análise, mantendo a
    experiência visual profissional para quem exporta os dados.
    """
    from sqlalchemy import extract

    # Verificar dependências do Excel
    if not EXCEL_AVAILABLE and not ensure_excel_available():
        flash(
            "Erro: Bibliotecas Excel não instaladas. Contate o administrador.",
            "danger",
        )
        return redirect(url_for("relatorio_vendas"))

    # Verificar permissão de exportação (mesma regra de clientes)
    if not pode_exportar(current_user, "clientes"):
        cargos_permitidos = ", ".join(
            get_cargos_permitidos_exportacao("clientes")
        )
        flash(
            f"Apenas {cargos_permitidos} podem exportar o relatório de vendas.",
            "danger",
        )
        return redirect(url_for("relatorio_vendas"))

    try:
        # Filtros (mesmos da tela)
        ano = request.args.get("ano", datetime.now().year, type=int)
        vendedor_id = request.args.get("vendedor_id", type=int)
        supervisor_id = request.args.get("supervisor_id", type=int)
        cidade = request.args.get("cidade", "")
        bairro = request.args.get("bairro", "")
        status = request.args.get("status", "")

        SupervisorAlias = db.aliased(Usuario)

        query = (
            db.session.query(
                Cliente.id,
                Cliente.cpf,
                Cliente.cnpj,
                Cliente.nome,
                Cliente.cidade,
                Cliente.bairro,
                Vendedor.nome.label("vendedor_nome"),
                SupervisorAlias.nome.label("supervisor_nome"),
            )
            .join(Vendedor, Cliente.vendedor_id == Vendedor.id)
            .outerjoin(
                SupervisorAlias, Vendedor.supervisor_id == SupervisorAlias.id
            )
        )

        # Escopo por permissão
        if current_user.cargo == "vendedor" and current_user.vendedor_id:
            query = query.filter(Cliente.vendedor_id == current_user.vendedor_id)
        elif current_user.cargo == "supervisor":
            vendedores_ids = [
                v.id
                for v in Vendedor.query.filter_by(
                    supervisor_id=current_user.id, ativo=True
                ).all()
            ]
            if vendedores_ids:
                query = query.filter(Cliente.vendedor_id.in_(vendedores_ids))
            else:
                query = query.filter(Cliente.id == -1)
        elif not current_user.is_super_admin:
            query = query.filter(Cliente.empresa_id == current_user.empresa_id)

        # Demais filtros
        if vendedor_id:
            query = query.filter(Cliente.vendedor_id == vendedor_id)
        if supervisor_id:
            query = query.filter(Vendedor.supervisor_id == supervisor_id)
        if cidade:
            query = query.filter(Cliente.cidade.ilike(f"%{cidade}%"))
        if bairro:
            query = query.filter(Cliente.bairro.ilike(f"%{bairro}%"))

        query = query.filter(Cliente.ativo)

        clientes_base = query.all()

        # Montar dados do relatório (mesma regra da tela)
        relatorio_dados = []
        total_compras_geral = 0
        total_valor_geral = 0

        for cliente_row in clientes_base:
            cliente = Cliente.query.get(cliente_row.id)

            compras_ano = CompraCliente.query.filter(
                CompraCliente.cliente_id == cliente.id,
                extract("year", CompraCliente.data_compra) == ano,
            ).all()

            total_compras_ano = len(compras_ano)
            valor_total_ano = (
                sum([c.valor for c in compras_ano]) if compras_ano else 0
            )

            ultima_compra = cliente.compras.order_by(
                CompraCliente.data_compra.desc()
            ).first()

            formas_usadas = set(
                [c.forma_pagamento for c in compras_ano if c.forma_pagamento]
            )

            status_cor = cliente.get_status_cor()

            if status and status_cor != status:
                continue

            relatorio_dados.append(
                {
                    "nome": cliente.nome,
                    "cpf": cliente.cpf or "-",
                    "cnpj": cliente.cnpj or "-",
                    "cidade": cliente.cidade or "-",
                    "bairro": cliente.bairro or "-",
                    "vendedor": cliente_row.vendedor_nome,
                    "supervisor": cliente_row.supervisor_nome or "-",
                    "compras_ano": total_compras_ano,
                    "valor_total_ano": valor_total_ano,
                    "ultima_compra": ultima_compra.data_compra if ultima_compra else None,
                    "formas_pagamento": ", ".join(formas_usadas)
                    if formas_usadas
                    else "-",
                    "status": status_cor,
                    "dias_sem_compra": (
                        (datetime.now() - ultima_compra.data_compra).days
                        if ultima_compra
                        else None
                    ),
                }
            )

            total_compras_geral += total_compras_ano
            total_valor_geral += valor_total_ano

        # Ordenar como na tela
        relatorio_dados.sort(key=lambda x: x["compras_ano"], reverse=True)

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Relatorio_Vendas_{ano}"

        header_fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "Nome do Cliente",
            "CPF",
            "CNPJ",
            "Cidade",
            "Bairro",
            "Vendedor",
            "Supervisor",
            f"Compras {ano}",
            "Valor Total (R$)",
            "Última Compra",
            "Dias sem Compra",
            "Formas Pagamento",
            "Status",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        status_map = {
            "verde": "Positivado",
            "amarelo": "Atenção",
            "vermelho": "Sem Compras",
        }

        row_idx = 2
        for item in relatorio_dados:
            ultima = (
                item["ultima_compra"].strftime("%d/%m/%Y")
                if item["ultima_compra"]
                else "Nunca"
            )
            dias_sem = (
                item["dias_sem_compra"]
                if item["dias_sem_compra"] is not None
                else "-"
            )
            status_legivel = status_map.get(item["status"], item["status"])

            ws.cell(row=row_idx, column=1, value=item["nome"])
            ws.cell(row=row_idx, column=2, value=item["cpf"])
            ws.cell(row=row_idx, column=3, value=item["cnpj"])
            ws.cell(row=row_idx, column=4, value=item["cidade"])
            ws.cell(row=row_idx, column=5, value=item["bairro"])
            ws.cell(row=row_idx, column=6, value=item["vendedor"])
            ws.cell(row=row_idx, column=7, value=item["supervisor"])
            ws.cell(row=row_idx, column=8, value=item["compras_ano"])
            ws.cell(row=row_idx, column=9, value=item["valor_total_ano"])
            ws.cell(row=row_idx, column=10, value=ultima)
            ws.cell(row=row_idx, column=11, value=dias_sem)
            ws.cell(row=row_idx, column=12, value=item["formas_pagamento"])
            ws.cell(row=row_idx, column=13, value=status_legivel)

            row_idx += 1

        # Linha de totais
        ws.cell(row=row_idx, column=1, value="TOTAIS:")
        ws.cell(row=row_idx, column=8, value=total_compras_geral)
        ws.cell(row=row_idx, column=9, value=total_valor_geral)

        # Ajustar larguras
        column_widths = [
            35,  # Nome
            16,  # CPF
            20,  # CNPJ
            20,  # Cidade
            20,  # Bairro
            25,  # Vendedor
            25,  # Supervisor
            14,  # Compras ano
            18,  # Valor total
            16,  # Última compra
            16,  # Dias sem compra
            28,  # Formas pagamento
            14,  # Status
        ]

        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relatorio_vendas_{ano}_{timestamp}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        flash(f"Erro ao exportar relatório de vendas: {str(e)}", "danger")
        return redirect(url_for("relatorio_vendas"))

@app.route("/clientes/exportar")
@login_required
def exportar_clientes():
    """Exportar clientes para Excel - Formato Simples ou Estendido"""
    if not EXCEL_AVAILABLE and not ensure_excel_available():
        flash("Erro: Bibliotecas Excel não instaladas. Contate o administrador.", "danger")
        return redirect(url_for("lista_clientes"))
    
    # Verificar permissão de exportação
    if not pode_exportar(current_user, "clientes"):
        cargos_permitidos = ", ".join(get_cargos_permitidos_exportacao("clientes"))
        flash(
            f"Apenas {cargos_permitidos} podem exportar clientes.",
            "danger"
        )
        return redirect(url_for("lista_clientes"))
    
    try:
        # Obter formato solicitado (simples ou estendido)
        formato = request.args.get("formato", "simples")

        # Buscar clientes (com filtros de permissão)
        if current_user.cargo == "vendedor" and current_user.vendedor_id:
            clientes = Cliente.query.filter_by(
                vendedor_id=current_user.vendedor_id,
                empresa_id=current_user.empresa_id,
            ).all()
        elif current_user.cargo == "supervisor" and current_user.supervisor_id:
            supervisor = Usuario.query.get(current_user.supervisor_id)
            vendedores_ids = [v.id for v in supervisor.vendedores]
            clientes = Cliente.query.filter(
                Cliente.vendedor_id.in_(vendedores_ids),
                Cliente.empresa_id == current_user.empresa_id,
            ).all()
        elif current_user.is_super_admin:
            clientes = Cliente.query.all()
        else:
            clientes = Cliente.query.filter_by(
                empresa_id=current_user.empresa_id
            ).all()

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Estilo do cabeçalho
        header_fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Definir cabeçalhos conforme formato
        if formato == "estendido":
            # Formato estendido - todos os campos conforme solicitado (27 colunas)
            headers = [
                "Nome",
                "Razão Social",
                "CPF",
                "CNPJ",
                "Inscr.Estadual",
                "Logradouro",
                "Município",
                "Bairro",
                "Estado",
                "CEP",
                "Celular 1",
                "Celular 2",
                "Fone",
                "Longitude",
                "Latitude",
                "Código-BW",
                "Id",
                "Vendedor",
                "Supervisor",
                "Sigla",
                "Código-BP",
                "Email",
                "Dia de Visita",
                "Formas de Pagamento",
                "Observações",
                "Status",
                "Ativo",
            ]
        else:  # formato simples
            headers = [
                "Nome",
                "CPF",
                "CNPJ",
                "Telefone",
                "Email",
                "Cidade",
                "Bairro",
                "Ponto de Referência",
                "Dia de Visita",
                "Formas de Pagamento",
                "Observações",
            ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Dados dos clientes
        for row, cliente in enumerate(clientes, 2):
            # Buscar última compra
            ultima_compra = cliente.compras.order_by(
                CompraCliente.data_compra.desc()
            ).first()
            data_ultima = (
                ultima_compra.data_compra.strftime("%d/%m/%Y")
                if ultima_compra
                else "Nunca"
            )

            # Formas de pagamento
            formas_pagamento = ", ".join(cliente.get_formas_pagamento_list())

            # Status
            status_map = {
                "verde": "VERDE",
                "amarelo": "AMARELO",
                "vermelho": "VERMELHO",
            }
            status = status_map.get(cliente.get_status_cor(), "N/A")

            # Vendedor
            vendedor_nome = (
                cliente.vendedor.nome if cliente.vendedor else "N/A"
            )
            
            # Supervisor
            supervisor_nome = (
                cliente.supervisor.nome if cliente.supervisor else ""
            )

            # Preencher linha conforme formato
            if formato == "estendido":
                # Formato estendido - todos os campos (27 colunas)
                ws.cell(row=row, column=1, value=cliente.nome)  # Nome
                ws.cell(row=row, column=2, value=cliente.razao_social or "")  # Razão Social
                ws.cell(row=row, column=3, value=cliente.cpf or "")  # CPF
                ws.cell(row=row, column=4, value=cliente.cnpj or "")  # CNPJ
                ws.cell(row=row, column=5, value=cliente.inscricao_estadual or "")  # Inscr.Estadual
                ws.cell(row=row, column=6, value=cliente.logradouro or "")  # Logradouro
                ws.cell(row=row, column=7, value=cliente.municipio or cliente.cidade or "")  # Município
                ws.cell(row=row, column=8, value=cliente.bairro or "")  # Bairro
                ws.cell(row=row, column=9, value=cliente.estado or "")  # Estado
                ws.cell(row=row, column=10, value=cliente.cep or "")  # CEP
                ws.cell(row=row, column=11, value=cliente.celular or "")  # Celular 1
                ws.cell(row=row, column=12, value=cliente.celular2 or "")  # Celular 2
                ws.cell(row=row, column=13, value=cliente.telefone or "")  # Fone
                ws.cell(row=row, column=14, value=cliente.longitude or cliente.coordenada_x or "")  # Longitude
                ws.cell(row=row, column=15, value=cliente.latitude or cliente.coordenada_y or "")  # Latitude
                ws.cell(row=row, column=16, value=cliente.codigo_bw or "")  # Código-BW
                ws.cell(row=row, column=17, value=cliente.id)  # Id
                ws.cell(row=row, column=18, value=vendedor_nome)  # Vendedor
                ws.cell(row=row, column=19, value=supervisor_nome)  # Supervisor
                ws.cell(row=row, column=20, value=cliente.sigla or "")  # Sigla
                ws.cell(row=row, column=21, value=cliente.codigo_bp or "")  # Código-BP
                ws.cell(row=row, column=22, value=cliente.email or "")  # Email
                ws.cell(row=row, column=23, value=cliente.dia_visita or "")  # Dia de Visita
                ws.cell(row=row, column=24, value=formas_pagamento)  # Formas de Pagamento
                ws.cell(row=row, column=25, value=cliente.observacoes or "")  # Observações
                ws.cell(row=row, column=26, value=status)  # Status
                ws.cell(row=row, column=27, value="Sim" if cliente.ativo else "Não")  # Ativo
            else:
                # Formato simples (11 colunas)
                ws.cell(row=row, column=1, value=cliente.nome)
                ws.cell(row=row, column=2, value=cliente.cpf or "")
                ws.cell(row=row, column=3, value=cliente.cnpj or "")
                ws.cell(row=row, column=4, value=cliente.telefone or "")
                ws.cell(row=row, column=5, value=cliente.email or "")
                ws.cell(row=row, column=6, value=cliente.cidade or "")
                ws.cell(row=row, column=7, value=cliente.bairro or "")
                ws.cell(
                    row=row, column=8, value=cliente.ponto_referencia or ""
                )
                ws.cell(row=row, column=9, value=cliente.dia_visita or "")
                ws.cell(row=row, column=10, value=formas_pagamento)
                ws.cell(row=row, column=11, value=cliente.observacoes or "")

        # Ajustar largura das colunas conforme formato
        if formato == "estendido":
            # 27 colunas com larguras otimizadas
            column_widths = [
                30,  # Nome
                35,  # Razão Social
                15,  # CPF
                18,  # CNPJ
                20,  # Inscr.Estadual
                40,  # Logradouro
                25,  # Município
                20,  # Bairro
                10,  # Estado
                12,  # CEP
                15,  # Celular 1
                15,  # Celular 2
                15,  # Fone
                15,  # Longitude
                15,  # Latitude
                15,  # Código-BW
                10,  # Id
                25,  # Vendedor
                25,  # Supervisor
                15,  # Sigla
                15,  # Código-BP
                30,  # Email
                15,  # Dia de Visita
                25,  # Formas de Pagamento
                40,  # Observações
                12,  # Status
                8,   # Ativo
            ]
        else:
            # 11 colunas
            column_widths = [30, 15, 18, 15, 25, 20, 20, 30, 15, 25, 40]

        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = width

        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Gerar nome do arquivo com indicação do formato
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"clientes_{formato}_{timestamp}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        flash(f"Erro ao exportar clientes: {str(e)}", "danger")
        return redirect(url_for("lista_clientes"))

@app.route("/clientes/modelo-importacao")
@login_required
def modelo_importacao_clientes():
    """Baixar modelo em branco para importação de clientes - Formato Simples ou Estendido"""
    # Verificar permissão de importação
    if not pode_importar(current_user, "clientes"):
        cargos_permitidos = ", ".join(get_cargos_permitidos_importacao("clientes"))
        flash(
            f"Apenas {cargos_permitidos} podem baixar o modelo de importação.",
            "danger"
        )
        return redirect(url_for("lista_clientes"))
    
    if not EXCEL_AVAILABLE and not ensure_excel_available():
        error_msg = "Erro: Bibliotecas Excel não instaladas."
        if EXCEL_ERROR_MESSAGE:
            print(f"📊 Erro Excel na rota modelo_importacao_clientes: {EXCEL_ERROR_MESSAGE}")
            error_msg += f" Detalhes: {EXCEL_ERROR_MESSAGE}"
        flash(error_msg + " Contate o administrador.", "danger")
        return redirect(url_for("lista_clientes"))
    
    try:
        # Obter formato solicitado (simples ou estendido)
        formato = request.args.get("formato", "simples")

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo Clientes"

        # Estilo do cabeçalho
        header_fill = PatternFill(
            start_color="0066CC", end_color="0066CC", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Definir cabeçalhos e exemplos conforme formato
        if formato == "estendido":
            # Formato estendido - todos os campos (27 colunas)
            headers = [
                "Nome",
                "Razão Social",
                "CPF",
                "CNPJ",
                "Inscr.Estadual",
                "Logradouro",
                "Município",
                "Bairro",
                "Estado",
                "CEP",
                "Celular 1",
                "Celular 2",
                "Fone",
                "Longitude",
                "Latitude",
                "Código-BW",
                "Id",
                "Vendedor",
                "Supervisor",
                "Sigla",
                "Código-BP",
                "Email",
                "Dia de Visita",
                "Formas de Pagamento",
                "Observações",
                "Status",
                "Ativo",
            ]
            exemplos = [
                "João Silva",  # Nome
                "Silva & Cia Ltda",  # Razão Social
                "123.456.789-00",  # CPF
                "12.345.678/0001-90",  # CNPJ
                "123.456.789.012",  # Inscr.Estadual
                "Rua das Flores, 123",  # Logradouro
                "São Paulo",  # Município
                "Centro",  # Bairro
                "SP",  # Estado
                "01310-100",  # CEP
                "(11) 98765-4321",  # Celular 1
                "(11) 98765-4322",  # Celular 2
                "(11) 3456-7890",  # Fone
                "-46.633308",  # Longitude
                "-23.550520",  # Latitude
                "BW-123456",  # Código-BW
                "",  # Id (deixar em branco para novo)
                "Nome do Vendedor",  # Vendedor
                "Nome do Supervisor",  # Supervisor
                "SilvaGrill",  # Sigla
                "BP-123456",  # Código-BP
                "joao@email.com",  # Email
                "segunda",  # Dia de Visita
                "dinheiro, pix, cartao",  # Formas de Pagamento
                "Cliente preferencial",  # Observações
                "",  # Status (calculado automaticamente)
                "Sim",  # Ativo
            ]
            column_widths = [
                30,  # Nome
                35,  # Razão Social
                15,  # CPF
                18,  # CNPJ
                20,  # Inscr.Estadual
                40,  # Logradouro
                25,  # Município
                20,  # Bairro
                10,  # Estado
                12,  # CEP
                15,  # Celular 1
                15,  # Celular 2
                15,  # Fone
                15,  # Longitude
                15,  # Latitude
                15,  # Código-BW
                10,  # Id
                25,  # Vendedor
                25,  # Supervisor
                15,  # Sigla
                15,  # Código-BP
                30,  # Email
                15,  # Dia de Visita
                25,  # Formas de Pagamento
                40,  # Observações
                12,  # Status
                8,   # Ativo
            ]
        else:
            # Formato simples (colunas originais)
            headers = [
                "Nome",
                "CPF",
                "CNPJ",
                "Telefone",
                "Email",
                "Cidade",
                "Bairro",
                "Ponto de Referência",
                "Dia de Visita",
                "Formas de Pagamento",
                "Observações",
            ]
            exemplos = [
                "João Silva",
                "123.456.789-00",
                "",
                "(11) 98765-4321",
                "joao@email.com",
                "São Paulo",
                "Centro",
                "Próximo ao mercado",
                "segunda",
                "dinheiro, pix, cartao_credito",
                "Cliente preferencial",
            ]
            column_widths = [30, 15, 18, 15, 25, 20, 20, 30, 15, 25, 40]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Adicionar linha de exemplo
        for col, exemplo in enumerate(exemplos, 1):
            ws.cell(row=2, column=col, value=exemplo)

        # Ajustar largura das colunas
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col).column_letter
            ].width = width

        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nome do arquivo com indicação do formato
        filename = f"modelo_importacao_clientes_{formato}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        flash(f"Erro ao gerar modelo: {str(e)}", "danger")
        return redirect(url_for("lista_clientes"))

@app.route("/clientes/importar", methods=["GET", "POST"])
@login_required
def importar_clientes():
    """Importar clientes via planilha Excel - Admin, Supervisor e RH"""
    # Verificar permissões: admin, supervisor e RH podem importar clientes
    if not pode_importar(current_user, "clientes"):
        flash(
            "Acesso negado! Apenas Administradores, Supervisores e RH podem importar clientes em lote.",
            "danger",
        )
        flash(
            "💡 Gerentes e Vendedores devem cadastrar clientes manualmente.",
            "info",
        )
        return redirect(url_for("lista_clientes"))

    # Buscar vendedores disponíveis para seleção
    vendedores = (
        Vendedor.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        )
        .order_by(Vendedor.nome)
        .all()
    )

    if request.method == "POST":
        # Obter modo de vendedor selecionado
        modo_vendedor = request.form.get("modo_vendedor", "planilha")
        vendedor_id = None
        usar_vendedor_planilha = False

        if modo_vendedor == "fixo":
            # Validar vendedor selecionado
            vendedor_id = request.form.get("vendedor_id")
            if not vendedor_id:
                flash(
                    'Selecione um vendedor quando escolher "Atribuir a um vendedor específico"!',
                    "danger",
                )
                return redirect(request.url)

            # Validar se vendedor existe e pertence à empresa
            vendedor = Vendedor.query.filter_by(
                id=vendedor_id, empresa_id=current_user.empresa_id
            ).first()
            if not vendedor:
                flash("Vendedor inválido!", "danger")
                return redirect(request.url)
        elif modo_vendedor == "planilha":
            usar_vendedor_planilha = True
        elif modo_vendedor == "depois":
            vendedor_id = None  # Será None, clientes sem vendedor

        # Validar arquivo Excel
        arquivo, erro = validar_arquivo_excel(request)
        if erro:
            flash(erro, "danger")
            return redirect(request.url)

        # Verificar disponibilidade do Excel com mensagens detalhadas
        if not EXCEL_AVAILABLE and not ensure_excel_available():
            error_msg = "❌ Erro: Funcionalidade de importação Excel indisponível."
            
            if EXCEL_ERROR_MESSAGE:
                print(f"📊 Erro Excel: {EXCEL_ERROR_MESSAGE}")
                
                # Detectar tipo de erro
                if "libstdc++" in EXCEL_ERROR_MESSAGE or ".so" in EXCEL_ERROR_MESSAGE:
                    flash(error_msg, "danger")
                    flash("🔧 Erro de biblioteca do sistema detectado. O administrador foi notificado.", "warning")
                    flash("💡 Enquanto isso, cadastre clientes manualmente.", "info")
                else:
                    flash(error_msg + " Contate o administrador.", "danger")
            else:
                flash(error_msg + " Contate o administrador.", "danger")
            
            return redirect(request.url)

        try:
            # Ler Excel
            print(f"📂 Lendo arquivo Excel: {arquivo.filename}")
            df = pd.read_excel(arquivo)

            # Normalizar nomes das colunas: remover espaços e converter para
            # minúsculas
            df.columns = df.columns.str.strip().str.lower()

            # Mapear colunas (formato simples + formato estendido)
            # Mapeamento de variantes de nomes de colunas para nomes
            # padronizados
            colunas_map = {
                "nome": [
                    "nome",
                    "nome completo",
                    "cliente",
                    "razao social",
                    "razão social",
                ],
                "cpf": ["cpf", "documento cpf"],
                "cnpj": ["cnpj", "documento cnpj"],
                "cpf_cnpj": [
                    "cpf/cnpj",
                    "cpf / cnpj",
                    "cnpj/cpf",
                    "documento único",
                    "documento unico",
                    "cpf ou cnpj",
                    "documento",
                    "doc",
                ],
                "razao_social": [
                    "razao social",
                    "razão social",
                    "razão",
                    "razao",
                ],
                "sigla": ["sigla", "apelido", "sigla/apelido"],
                "inscricao_estadual": [
                    "inscr.estadual",
                    "inscrição estadual",
                    "inscricao estadual",
                    "ie",
                    "i.e.",
                    "insc. estadual",
                ],
                "codigo_bp": [
                    "codigo-bp",
                    "código bp",
                    "codigo bp",
                    "bp",
                    "código erp",
                    "codigo erp",
                    "cod bp",
                    "cod. bp",
                ],
                "logradouro": [
                    "logradouro",
                    "endereço",
                    "endereco",
                    "rua",
                    "avenida",
                    "av",
                    "end",
                    "end.",
                ],
                "telefone": [
                    "fone(1)",
                    "telefone",
                    "fone",
                    "fone 1",
                    "telefone 1",
                    "contato",
                    "tel",
                    "tel 1",
                ],
                "telefone2": [
                    "fone(2)",
                    "telefone 2",
                    "fone 2",
                    "fone2",
                    "telefone2",
                    "tel 2",
                    "tel2",
                ],
                "celular": [
                    "cel(1)",
                    "celular",
                    "cel",
                    "cel 1",
                    "celular 1",
                    "móvel",
                    "movel",
                    "whatsapp",
                ],
                "celular2": [
                    "cel(2)",
                    "celular 2",
                    "cel 2",
                    "cel2",
                    "celular2",
                    "móvel 2",
                    "movel 2",
                ],
                "email": ["email", "e-mail", "e mail", "e_mail"],
                "municipio": ["município", "municipio", "cidade", "mun"],
                "cidade": ["cidade"],  # Mantido para compatibilidade
                "bairro": ["bairro", "região", "regiao"],
                "estado": ["estado", "uf", "u.f.", "sigla estado"],
                "cep": ["cep", "código postal", "codigo postal", "cod postal"],
                "coordenada_x": [
                    "coordenada-x",
                    "coordenada x",
                    "coordenadax",
                    "longitude",
                    "long",
                    "coord x",
                ],
                "coordenada_y": [
                    "coordenada-y",
                    "coordenada y",
                    "coordenaday",
                    "latitude",
                    "lat",
                    "coord y",
                ],
                "longitude": ["longitude", "long", "lng"],
                "latitude": ["latitude", "lat"],
                "codigo_bw": [
                    "código-bw",
                    "codigo-bw",
                    "código bw",
                    "codigo bw",
                    "bw",
                    "cod bw",
                    "cod-bw",
                ],
                "ponto_referencia": [
                    "ponto de referência",
                    "ponto de referencia",
                    "referência",
                    "referencia",
                    "ponto ref",
                ],
                "dia_visita": [
                    "dia de visita",
                    "dia visita",
                    "dia",
                    "dia semana",
                ],
                "formas_pagamento": [
                    "formas de pagamento",
                    "pagamento",
                    "formas pagamento",
                    "forma pgto",
                    "forma de pagamento",
                ],
                "observacoes": [
                    "observações",
                    "observacoes",
                    "obs",
                    "observação",
                ],
                "vendedor": [
                    "vendedor",
                    "vendedor responsável",
                    "vendedor responsavel",
                    "id vendedor",
                    "vendedor id",
                    "nome vendedor",
                ],
                "supervisor": [
                    "supervisor",
                    "supervisor responsável",
                    "supervisor responsavel",
                    "id supervisor",
                    "supervisor id",
                    "nome supervisor",
                ],
            }

            # Renomear colunas com base no mapeamento
            rename_dict = {}
            for col in df.columns:
                # Procurar qual coluna padronizada corresponde a esta coluna
                for col_padrao, variantes in colunas_map.items():
                    if col in variantes and col != col_padrao:
                        rename_dict[col] = col_padrao
                        break

            # Aplicar renomeações
            if rename_dict:
                df.rename(columns=rename_dict, inplace=True)

            # Validar colunas obrigatórias
            colunas_obrigatorias = ["nome"]
            colunas_faltando = [
                col for col in colunas_obrigatorias if col not in df.columns
            ]

            if colunas_faltando:
                # Mostrar as colunas encontradas para ajudar no debug
                colunas_encontradas = ", ".join(df.columns.tolist()[:10])
                if len(df.columns) > 10:
                    colunas_encontradas += "..."

                msg = f'Coluna obrigatória "Nome" não encontrada na planilha! '
                msg += f'Certifique-se de que sua planilha tem uma coluna chamada "Nome". '
                msg += f"Colunas detectadas: {colunas_encontradas}"
                flash(msg, "danger")
                return redirect(request.url)

            # Processar registros
            importados = 0
            atualizados = 0
            erros = []
            pulados = 0  # Contador de linhas vazias

            for index, row in df.iterrows():
                try:
                    # Validar nome - PERMITIR PULAR LINHAS VAZIAS
                    nome = str(row.get("nome", "")).strip()
                    if not nome or nome == "nan" or nome == "None":
                        # Pular silenciosamente linhas sem nome (provavelmente
                        # vazias)
                        pulados += 1
                        continue

                    # Helper para limpar números vindos do Excel (remove .0 de floats)
                    def limpar_numero_excel(valor):
                        if pd.isna(valor): return ""
                        s = str(valor).strip()
                        if s.endswith('.0'):
                            s = s[:-2]
                        return re.sub(r'\D', '', s)

                    # Processar CPF/CNPJ (formato combinado ou separado)
                    cpf = None
                    cnpj = None

                    # Verificar se tem coluna combinada CPF/CNPJ
                    if "cpf_cnpj" in df.columns:
                        cpf_cnpj_raw = row.get("cpf_cnpj")
                        doc = limpar_numero_excel(cpf_cnpj_raw)

                        if doc:
                            if len(doc) == 11:
                                cpf = doc
                            elif len(doc) == 14:
                                cnpj = doc
                            elif len(doc) > 0 and len(doc) < 11:
                                # Menos de 11 dígitos, provavelmente CPF incompleto
                                cpf = doc.zfill(11)
                            elif len(doc) > 11 and len(doc) < 14:
                                # Entre 11 e 14, provavelmente CNPJ incompleto
                                cnpj = doc.zfill(14)
                            else:
                                # Tamanho inválido, mas vamos tentar salvar no campo mais provável
                                # Se < 11, assume CPF. Se > 11, assume CNPJ.
                                if len(doc) < 11:
                                    cpf = doc.zfill(11)
                                else:
                                    cnpj = doc.zfill(14)

                    # Se não tiver coluna combinada, buscar CPF e CNPJ separados
                    if cpf is None and "cpf" in df.columns:
                        cpf_raw = row.get("cpf")
                        cpf_limpo = limpar_numero_excel(cpf_raw)
                        if cpf_limpo:
                            if len(cpf_limpo) <= 11:
                                cpf = cpf_limpo.zfill(11)
                            else:
                                # Se tem mais de 11 dígitos no campo CPF, pode ser um erro ou CNPJ no lugar errado
                                # Vamos tentar aproveitar se for 14 dígitos
                                if len(cpf_limpo) == 14:
                                    cnpj = cpf_limpo
                                else:
                                    cpf = None # Inválido

                    if cnpj is None and "cnpj" in df.columns:
                        cnpj_raw = row.get("cnpj")
                        cnpj_limpo = limpar_numero_excel(cnpj_raw)
                        if cnpj_limpo:
                            if len(cnpj_limpo) <= 14:
                                cnpj = cnpj_limpo.zfill(14)
                            else:
                                cnpj = None # Inválido

                    # Verificar duplicidade e decidir se atualiza ou cria
                    cliente_existente = None

                    if cpf:
                        cliente_existente = Cliente.query.filter_by(
                            cpf=cpf, empresa_id=current_user.empresa_id
                        ).first()

                    if not cliente_existente and cnpj:
                        cliente_existente = Cliente.query.filter_by(
                            cnpj=cnpj, empresa_id=current_user.empresa_id
                        ).first()

                    # Helper function para limpar e validar strings
                    def limpar_valor(valor):
                        """Limpa valor, retorna None se vazio ou 'nan'"""
                        if pd.isna(valor) or valor is None:
                            return None
                        valor_str = str(valor).strip()
                        if (
                            not valor_str or
                            valor_str.lower() == "nan" or
                            valor_str == ""
                        ):
                            return None
                        return valor_str

                    # Processar formas de pagamento
                    formas_str = limpar_valor(row.get("formas_pagamento"))
                    formas_pagamento = []
                    if formas_str:
                        # Aceitar separados por vírgula ou ponto-e-vírgula
                        formas_list = re.split(r"[,;]", formas_str)
                        formas_validas = [
                            "dinheiro",
                            "pix",
                            "cartao_debito",
                            "cartao_credito",
                            "boleto",
                        ]
                        for forma in formas_list:
                            forma_limpa = (
                                forma.strip().lower().replace(" ", "_")
                            )
                            if forma_limpa in formas_validas:
                                formas_pagamento.append(forma_limpa)

                    # Extrair e limpar valores
                    razao_social = limpar_valor(row.get("razao_social"))
                    sigla = limpar_valor(row.get("sigla"))
                    inscricao_estadual = limpar_valor(
                        row.get("inscricao_estadual")
                    )
                    codigo_bp = limpar_valor(row.get("codigo_bp"))
                    logradouro = limpar_valor(row.get("logradouro"))
                    telefone = limpar_valor(row.get("telefone"))
                    telefone2 = limpar_valor(row.get("telefone2"))
                    celular = limpar_valor(row.get("celular"))
                    celular2 = limpar_valor(row.get("celular2"))
                    email = limpar_valor(row.get("email"))
                    municipio = limpar_valor(row.get("municipio"))
                    cidade = (
                        municipio
                        if municipio
                        else limpar_valor(row.get("cidade"))
                    )  # Compatibilidade
                    bairro = limpar_valor(row.get("bairro"))
                    estado = limpar_valor(row.get("estado"))
                    cep = limpar_valor(row.get("cep"))
                    coordenada_x = limpar_valor(row.get("coordenada_x"))
                    coordenada_y = limpar_valor(row.get("coordenada_y"))
                    longitude = limpar_valor(row.get("longitude")) or coordenada_x
                    latitude = limpar_valor(row.get("latitude")) or coordenada_y
                    codigo_bw = limpar_valor(row.get("codigo_bw"))
                    ponto_referencia = limpar_valor(
                        row.get("ponto_referencia")
                    )
                    dia_visita = limpar_valor(row.get("dia_visita"))
                    observacoes = limpar_valor(row.get("observacoes"))

                    # Determinar vendedor_id para este cliente
                    cliente_vendedor_id = (
                        vendedor_id  # Usar o vendedor fixo se foi definido
                    )

                    if usar_vendedor_planilha and "vendedor" in df.columns:
                        # Tentar pegar vendedor da planilha
                        vendedor_valor = limpar_valor(row.get("vendedor"))
                        if vendedor_valor:
                            # Pode ser ID numérico ou nome
                            try:
                                # Tentar como ID
                                vendedor_id_planilha = int(vendedor_valor)
                                vendedor_encontrado = Vendedor.query.filter_by(
                                    id=vendedor_id_planilha,
                                    empresa_id=current_user.empresa_id,
                                    ativo=True,
                                ).first()
                                if vendedor_encontrado:
                                    cliente_vendedor_id = (
                                        vendedor_encontrado.id
                                    )
                            except ValueError:
                                # Não é número, tentar como nome
                                vendedor_encontrado = Vendedor.query.filter(
                                    Vendedor.nome.ilike(f"%{vendedor_valor}%"),
                                    Vendedor.empresa_id ==
                                    current_user.empresa_id,
                                    Vendedor.ativo,
                                ).first()
                                if vendedor_encontrado:
                                    cliente_vendedor_id = (
                                        vendedor_encontrado.id
                                    )

                    # Processar supervisor (se houver na planilha)
                    cliente_supervisor_id = None
                    if "supervisor" in df.columns:
                        supervisor_valor = limpar_valor(row.get("supervisor"))
                        if supervisor_valor:
                            try:
                                # Tentar como ID
                                supervisor_id_planilha = int(supervisor_valor)
                                supervisor_encontrado = Usuario.query.filter_by(
                                    id=supervisor_id_planilha,
                                    empresa_id=current_user.empresa_id,
                                    cargo='supervisor',
                                    ativo=True,
                                ).first()
                                if supervisor_encontrado:
                                    cliente_supervisor_id = supervisor_encontrado.id
                            except ValueError:
                                # Não é número, tentar como nome
                                supervisor_encontrado = Usuario.query.filter(
                                    Usuario.nome.ilike(f"%{supervisor_valor}%"),
                                    Usuario.empresa_id == current_user.empresa_id,
                                    Usuario.cargo == 'supervisor',
                                    Usuario.ativo == True,
                                ).first()
                                if supervisor_encontrado:
                                    cliente_supervisor_id = supervisor_encontrado.id

                    if cliente_existente:
                        # ATUALIZAR CLIENTE EXISTENTE
                        cliente = cliente_existente
                        cliente.nome = nome
                        if razao_social: cliente.razao_social = razao_social
                        if sigla: cliente.sigla = sigla
                        if inscricao_estadual: cliente.inscricao_estadual = inscricao_estadual
                        if codigo_bp: cliente.codigo_bp = codigo_bp

                        # Atualizar endereço
                        if logradouro: cliente.logradouro = logradouro
                        if municipio: cliente.municipio = municipio
                        if cidade: cliente.cidade = cidade
                        if bairro: cliente.bairro = bairro
                        if estado: cliente.estado = estado
                        if cep: cliente.cep = cep
                        if ponto_referencia: cliente.ponto_referencia = ponto_referencia
                        if coordenada_x: cliente.coordenada_x = coordenada_x
                        if coordenada_y: cliente.coordenada_y = coordenada_y
                        if longitude: cliente.longitude = longitude
                        if latitude: cliente.latitude = latitude

                        # Atualizar contato
                        if telefone: cliente.telefone = telefone
                        if telefone2: cliente.telefone2 = telefone2
                        if celular: cliente.celular = celular
                        if celular2: cliente.celular2 = celular2
                        if email: cliente.email = email

                        # Atualizar códigos
                        if codigo_bw: cliente.codigo_bw = codigo_bw

                        # Atualizar outros
                        if dia_visita: cliente.dia_visita = dia_visita
                        if observacoes: cliente.observacoes = observacoes

                        # Atualizar vendedor se fornecido
                        if cliente_vendedor_id:
                            cliente.vendedor_id = cliente_vendedor_id
                        
                        # Atualizar supervisor se fornecido
                        if cliente_supervisor_id:
                            cliente.supervisor_id = cliente_supervisor_id

                        # Atualizar formas de pagamento
                        if formas_pagamento:
                            cliente.set_formas_pagamento_list(formas_pagamento)

                        cliente.data_atualizacao = datetime.utcnow()
                        atualizados += 1

                    else:
                        # CRIAR NOVO CLIENTE
                        # Gerar código único do cliente baseado no município
                        municipio_codigo = (
                            municipio
                            if municipio
                            else cidade if cidade else "SEM_CIDADE"
                        )
                        codigo_cliente = Cliente.gerar_codigo_cliente(
                            municipio_codigo, current_user.empresa_id
                        )

                        # Criar cliente
                        cliente = Cliente(
                            nome=nome,
                            cpf=cpf,
                            cnpj=cnpj,
                            razao_social=razao_social,
                            sigla=sigla,
                            inscricao_estadual=inscricao_estadual,
                            codigo_bp=codigo_bp,
                            codigo_bw=codigo_bw,
                            codigo_cliente=codigo_cliente,
                            logradouro=logradouro,
                            municipio=municipio,
                            cidade=cidade,  # Mantém para compatibilidade
                            bairro=bairro,
                            estado=estado,
                            cep=cep,
                            coordenada_x=coordenada_x,
                            coordenada_y=coordenada_y,
                            longitude=longitude,
                            latitude=latitude,
                            ponto_referencia=ponto_referencia,
                            telefone=telefone,
                            telefone2=telefone2,
                            celular=celular,
                            celular2=celular2,
                            email=email,
                            dia_visita=dia_visita,
                            observacoes=observacoes,
                            vendedor_id=cliente_vendedor_id,
                            supervisor_id=cliente_supervisor_id,
                            empresa_id=current_user.empresa_id,
                            ativo=True,
                        )

                        # Configurar formas de pagamento
                        if formas_pagamento:
                            cliente.set_formas_pagamento_list(formas_pagamento)

                        db.session.add(cliente)
                        importados += 1

                except Exception as e:
                    erros.append(f"Linha {index + 2}: {str(e)}")

            # Commit
            db.session.commit()

            # Feedback
            total_processados = importados + atualizados
            if total_processados > 0:
                msg_sucesso = (
                    f"Processamento concluído! {importados} novos clientes importados e {atualizados} atualizados."
                )
                if pulados > 0:
                    msg_sucesso += (
                        f" ({pulados} linha(s) vazia(s) ignorada(s))"
                    )
                flash(msg_sucesso, "success")
            elif pulados > 0 and not erros:
                flash(
                    f"⚠️ Nenhum cliente importado. {pulados} linha(s) vazia(s) foram ignoradas. Verifique se a planilha contém dados válidos.",
                    "warning",
                )

            if erros:
                # Mostrar apenas os primeiros 3 erros para não sobrecarregar a
                # interface
                erros_msg = "<br>".join(erros[:3])
                if len(erros) > 3:
                    erros_msg += f"<br>... e mais {len(erros) - 3} erro(s)"
                flash(f"⚠️ Erros encontrados:<br>{erros_msg}", "warning")

            return redirect(url_for("lista_clientes"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao importar arquivo: {str(e)}", "danger")
            return redirect(request.url)

    return render_template("clientes/importar.html", vendedores=vendedores)

def verificar_acesso_cliente(cliente):
    """Verifica se o usuário atual pode acessar o cliente"""
    if current_user.is_super_admin:
        return True

    if current_user.cargo == "vendedor" and current_user.vendedor_id:
        return cliente.vendedor_id == current_user.vendedor_id

    if current_user.cargo == "supervisor":
        vendedores_ids = [v.id for v in current_user.vendedores]
        return cliente.vendedor_id in vendedores_ids

    # Admin da empresa
    return cliente.empresa_id == current_user.empresa_id

# ===== ROTAS DE MENSAGENS =====

@app.route("/mensagens")
@login_required
@permission_required("pode_enviar_mensagens")
def caixa_entrada():
    """Caixa de entrada - mensagens recebidas"""
    # Buscar mensagens recebidas não arquivadas
    mensagens = (
        Mensagem.query.filter_by(
            destinatario_id=current_user.id, arquivada_destinatario=False
        )
        .order_by(Mensagem.data_envio.desc())
        .all()
    )

    # Contar não lidas
    nao_lidas = Mensagem.query.filter_by(
        destinatario_id=current_user.id,
        lida=False,
        arquivada_destinatario=False,
    ).count()

    return render_template(
        "mensagens/caixa_entrada.html",
        mensagens=mensagens,
        nao_lidas=nao_lidas,
    )

@app.route("/mensagens/enviadas")
@login_required
@permission_required("pode_enviar_mensagens")
def mensagens_enviadas():
    """Mensagens enviadas pelo usuário"""
    mensagens = (
        Mensagem.query.filter_by(
            remetente_id=current_user.id, arquivada_remetente=False
        )
        .order_by(Mensagem.data_envio.desc())
        .all()
    )

    return render_template("mensagens/enviadas.html", mensagens=mensagens)

@app.route("/mensagens/nova", methods=["GET", "POST"])
@login_required
@permission_required("pode_enviar_mensagens")
def nova_mensagem():
    """Enviar nova mensagem"""
    if request.method == "POST":
        destinatario_id = request.form.get("destinatario_id")
        assunto = request.form.get("assunto")
        mensagem_texto = request.form.get("mensagem")
        prioridade = request.form.get("prioridade", "normal")

        if not destinatario_id or not assunto or not mensagem_texto:
            flash("Preencha todos os campos obrigatórios!", "danger")
            return redirect(url_for("nova_mensagem"))

        try:
            # Verificar se destinatário existe e está na mesma empresa
            destinatario = Usuario.query.get(int(destinatario_id))
            if not destinatario:
                flash("Destinatário não encontrado!", "danger")
                return redirect(url_for("nova_mensagem"))

            # Verificar se é da mesma empresa (exceto super admin)
            if not current_user.is_super_admin:
                if destinatario.empresa_id != current_user.empresa_id:
                    flash(
                        "Você só pode enviar mensagens para usuários da sua empresa!",
                        "danger",
                    )
                    return redirect(url_for("nova_mensagem"))

            # Criar mensagem
            mensagem = Mensagem(
                remetente_id=current_user.id,
                destinatario_id=int(destinatario_id),
                assunto=assunto,
                mensagem=mensagem_texto,
                prioridade=prioridade,
                tipo="individual",  # Mensagem individual
            )

            db.session.add(mensagem)
            db.session.commit()

            flash("Mensagem enviada com sucesso!", "success")
            return redirect(url_for("mensagens_enviadas"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao enviar mensagem: {str(e)}", "danger")

    # Buscar usuários da mesma empresa para o select
    if current_user.is_super_admin:
        usuarios = (
            Usuario.query.filter(Usuario.id != current_user.id, Usuario.ativo)
            .order_by(Usuario.nome)
            .all()
        )
    elif current_user.cargo == "supervisor":
        # Supervisor vê apenas seus vendedores
        # Buscar IDs dos vendedores supervisionados
        vendedores_ids = (
            db.session.query(Vendedor.id)
            .filter_by(supervisor_id=current_user.id, ativo=True)
            .all()
        )
        vendedores_ids = [v[0] for v in vendedores_ids]

        # Buscar usuários vinculados aos vendedores
        usuarios = (
            Usuario.query.filter(
                Usuario.vendedor_id.in_(vendedores_ids),
                Usuario.id != current_user.id,
                Usuario.ativo,
            )
            .order_by(Usuario.nome)
            .all()
        )
    elif current_user.cargo == "vendedor" and current_user.vendedor_id:
        # Vendedor vê apenas vendedores da sua equipe
        vendedor_atual = Vendedor.query.get(current_user.vendedor_id)
        if vendedor_atual and vendedor_atual.equipe_id:
            # Buscar vendedores da mesma equipe
            vendedores_equipe = (
                db.session.query(Vendedor.id)
                .filter_by(equipe_id=vendedor_atual.equipe_id, ativo=True)
                .all()
            )
            vendedores_ids = [v[0] for v in vendedores_equipe]

            # Buscar usuários vinculados aos vendedores da equipe
            usuarios = (
                Usuario.query.filter(
                    Usuario.vendedor_id.in_(vendedores_ids),
                    Usuario.id != current_user.id,
                    Usuario.ativo,
                )
                .order_by(Usuario.nome)
                .all()
            )
        else:
            # Se não tem equipe, não pode enviar mensagens
            usuarios = []
    else:
        usuarios = (
            Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.id != current_user.id,
                Usuario.ativo,
            )
            .order_by(Usuario.nome)
            .all()
        )

    return render_template("mensagens/nova.html", usuarios=usuarios)

@app.route("/mensagens/<int:id>")
@login_required
@permission_required("pode_enviar_mensagens")
def ver_mensagem(id):
    """Visualizar mensagem"""
    mensagem = Mensagem.query.get_or_404(id)

    # Verificar permissão:
    # - Mensagens individuais: apenas remetente e destinatário
    # - Mensagens de grupo: qualquer membro da equipe pode ver
    pode_ver = False

    if (
        mensagem.remetente_id == current_user.id or
        mensagem.destinatario_id == current_user.id
    ):
        # É remetente ou destinatário
        pode_ver = True
    elif mensagem.tipo == "grupo":
        # Mensagem de grupo: verificar se usuário faz parte da mesma equipe
        # Buscar vendedor do remetente para identificar a equipe
        remetente = Usuario.query.get(mensagem.remetente_id)
        if remetente and remetente.vendedor_id:
            vendedor_remetente = Vendedor.query.get(remetente.vendedor_id)
            if vendedor_remetente and vendedor_remetente.equipe_id:
                # Verificar se current_user está na mesma equipe
                if current_user.vendedor_id:
                    vendedor_atual = Vendedor.query.get(
                        current_user.vendedor_id
                    )
                    if (
                        vendedor_atual and
                        vendedor_atual.equipe_id ==
                        vendedor_remetente.equipe_id
                    ):
                        pode_ver = True
                # Supervisor da equipe também pode ver
                elif current_user.cargo == "supervisor":
                    if vendedor_remetente.supervisor_id == current_user.id:
                        pode_ver = True

    if not pode_ver:
        flash("Você não tem permissão para ver esta mensagem!", "danger")
        return redirect(url_for("caixa_entrada"))

    # Se for destinatário e não lida, marcar como lida
    if mensagem.destinatario_id == current_user.id and not mensagem.lida:
        mensagem.marcar_como_lida()

    return render_template("mensagens/ver.html", mensagem=mensagem)

@app.route("/mensagens/<int:id>/arquivar", methods=["POST"])
@login_required
@permission_required("pode_enviar_mensagens")
def arquivar_mensagem(id):
    """Arquivar mensagem"""
    mensagem = Mensagem.query.get_or_404(id)

    # Verificar se o usuário é remetente ou destinatário
    if (
        mensagem.remetente_id != current_user.id and
        mensagem.destinatario_id != current_user.id
    ):
        flash("Você não tem permissão para arquivar esta mensagem!", "danger")
        return redirect(url_for("caixa_entrada"))

    # Arquivar conforme remetente ou destinatário
    if mensagem.remetente_id == current_user.id:
        mensagem.arquivada_remetente = True
    if mensagem.destinatario_id == current_user.id:
        mensagem.arquivada_destinatario = True

    db.session.commit()
    flash("Mensagem arquivada com sucesso!", "success")

    # Redirecionar de volta
    return redirect(request.referrer or url_for("caixa_entrada"))

@app.route("/mensagens/<int:id>/marcar-lida", methods=["POST"])
@login_required
@permission_required("pode_enviar_mensagens")
def marcar_como_lida(id):
    """Marcar mensagem como lida"""
    mensagem = Mensagem.query.get_or_404(id)

    # Apenas destinatário pode marcar como lida
    if mensagem.destinatario_id != current_user.id:
        flash("Você não tem permissão para modificar esta mensagem!", "danger")
        return redirect(url_for("caixa_entrada"))

    mensagem.marcar_como_lida()

    return redirect(request.referrer or url_for("caixa_entrada"))

@app.route("/mensagens/<int:id>/deletar", methods=["POST"])
@login_required
@permission_required("pode_enviar_mensagens")
def deletar_mensagem(id):
    """Deletar mensagem"""
    mensagem = Mensagem.query.get_or_404(id)

    # Verificar se o usuário é remetente ou destinatário
    if (
        mensagem.remetente_id != current_user.id and
        mensagem.destinatario_id != current_user.id
    ):
        flash("Você não tem permissão para deletar esta mensagem!", "danger")
        return redirect(url_for("caixa_entrada"))

    # Deletar mensagem
    db.session.delete(mensagem)
    db.session.commit()

    flash("Mensagem deletada com sucesso!", "success")
    return redirect(request.referrer or url_for("caixa_entrada"))

@app.route("/mensagens/enviar-equipe", methods=["GET", "POST"])
@login_required
@permission_required("pode_enviar_mensagens")
def enviar_mensagem_equipe():
    """Enviar mensagem para todos da equipe"""
    # Verificar se usuário é vendedor ou supervisor
    if current_user.cargo not in [
        "vendedor",
        "supervisor",
        "gerente",
        "admin",
    ]:
        flash(
            "Apenas vendedores, supervisores, gerentes e admins podem enviar mensagens para equipe!",
            "danger",
        )
        return redirect(url_for("caixa_entrada"))

    if request.method == "POST":
        equipe_id = request.form.get("equipe_id")
        assunto = request.form.get("assunto")
        mensagem_texto = request.form.get("mensagem")
        prioridade = request.form.get("prioridade", "normal")

        if not equipe_id or not assunto or not mensagem_texto:
            flash("Preencha todos os campos obrigatórios!", "danger")
            return redirect(url_for("enviar_mensagem_equipe"))

        try:
            # Buscar equipe
            equipe = Equipe.query.get(int(equipe_id))
            if not equipe:
                flash("Equipe não encontrada!", "danger")
                return redirect(url_for("enviar_mensagem_equipe"))

            # Buscar vendedores da equipe
            vendedores = Vendedor.query.filter_by(
                equipe_id=equipe.id, ativo=True
            ).all()

            enviados = 0
            for vendedor in vendedores:
                # Buscar usuário do vendedor
                usuario_vendedor = Usuario.query.filter_by(
                    vendedor_id=vendedor.id, ativo=True
                ).first()

                if usuario_vendedor and usuario_vendedor.id != current_user.id:
                    # Criar mensagem
                    mensagem = Mensagem(
                        remetente_id=current_user.id,
                        destinatario_id=usuario_vendedor.id,
                        assunto=f"[Equipe {equipe.nome}] {assunto}",
                        mensagem=mensagem_texto,
                        prioridade=prioridade,
                        tipo="grupo",  # Mensagem de grupo
                    )
                    db.session.add(mensagem)
                    enviados += 1

            db.session.commit()
            flash(f"Mensagem enviada para {enviados} membros da equipe {equipe.nome}!", "success")
            return redirect(url_for("mensagens_enviadas"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao enviar mensagem: {str(e)}", "danger")

    # Buscar equipes disponíveis
    if current_user.is_super_admin:
        equipes = Equipe.query.filter_by(ativa=True).all()
    else:
        equipes = Equipe.query.filter_by(
            empresa_id=current_user.empresa_id, ativa=True
        ).all()

    return render_template("mensagens/enviar_equipe.html", equipes=equipes)

# ===== FUNÇÕES AUXILIARES PARA METAS =====

def _prepare_meta_form(form):
    """Prepara o formulário de meta, populando os vendedores."""
    vendedores = filtrar_vendedores_por_escopo(current_user)
    form.vendedor_id.choices = [(v.id, v.nome) for v in vendedores]

def _save_meta_from_form(form, meta=None):
    """Salva uma meta (nova ou existente) a partir dos dados do formulário."""
    vendedor = Vendedor.query.get(form.vendedor_id.data)
    if not vendedor:
        raise ValueError("Vendedor não encontrado.")

    if not current_user.is_super_admin and vendedor.empresa_id != current_user.empresa_id:
        raise PermissionError("Você só pode gerenciar metas da sua empresa.")

    if meta:  # Edição
        meta.valor_meta = form.valor_meta.data
        meta.mes = form.mes.data
        meta.ano = form.ano.data
        meta.vendedor_id = form.vendedor_id.data
    else:  # Criação
        meta = Meta(
            valor_meta=form.valor_meta.data,
            mes=form.mes.data,
            ano=form.ano.data,
            vendedor_id=form.vendedor_id.data,
            empresa_id=vendedor.empresa_id,
        )
        db.session.add(meta)
    
    db.session.commit()
    return meta

# ===== FUNÇÕES AUXILIARES PARA COMPRAS DE CLIENTES =====

def _prepare_compra_cliente_form(form, cliente_id=None):
    """Prepara o formulário de compra, populando clientes e produtos."""
    clientes = filtrar_clientes_por_escopo(current_user)
    form.cliente_id.choices = [(c.id, c.nome) for c in clientes]

    produtos_query = Produto.query.filter_by(ativo=True)
    if not current_user.is_super_admin:
        produtos_query = produtos_query.filter_by(empresa_id=current_user.empresa_id)
    
    produtos = produtos_query.order_by(Produto.nome).all()
    form.produto_id.choices = [(p.id, f"{p.nome} - {formatar_moeda(p.preco)}") for p in produtos]

    if cliente_id:
        form.cliente_id.data = cliente_id

def _update_meta_for_compra(vendedor_id, valor_compra, data_compra, anular=False):
    """Atualiza a receita alcançada e a comissão na meta do vendedor.

    Esta função é utilizada sempre que uma compra é registrada ou anulada,
    garantindo que o Ranking de Metas e os relatórios enxerguem a receita
    real do vendedor no período e a comissão correspondente.
    """
    meta = Meta.query.filter_by(
        vendedor_id=vendedor_id,
        mes=data_compra.month,
        ano=data_compra.year,
    ).first()
    if meta:
        if anular:
            meta.receita_alcancada = (meta.receita_alcancada or 0) - valor_compra
        else:
            meta.receita_alcancada = (meta.receita_alcancada or 0) + valor_compra

        # Recalcular percentual de alcance e comissão da meta
        try:
            meta.calcular_comissao()
        except Exception as e:
            app.logger.error(
                f"Erro ao recalcular comissão da meta {meta.id}: {e}"
            )

def _save_compra_cliente_from_form(form, compra=None):
    """Salva uma compra (nova ou existente) a partir dos dados do formulário."""
    cliente = Cliente.query.get(form.cliente_id.data)
    if not cliente:
        raise ValueError("Cliente não encontrado.")

    if not current_user.is_super_admin and cliente.empresa_id != current_user.empresa_id:
        raise PermissionError("Você só pode registrar compras para clientes da sua empresa.")

    produto = Produto.query.get(form.produto_id.data)
    if not produto:
        raise ValueError("Produto não encontrado.")

    valor_total = produto.preco * form.quantidade.data
    data_compra = form.data_compra.data

    if compra:  # Edição
        # Reverter valor antigo da meta
        _update_meta_for_compra(compra.vendedor_id, compra.valor_total, compra.data_compra, anular=True)

        compra.cliente_id = cliente.id
        compra.produto_id = produto.id
        compra.quantidade = form.quantidade.data
        compra.data_compra = data_compra
        compra.valor_total = valor_total
        compra.vendedor_id = cliente.vendedor_id
        compra.empresa_id = cliente.empresa_id
    else:  # Criação
        compra = CompraCliente(
            cliente_id=cliente.id,
            produto_id=produto.id,
            quantidade=form.quantidade.data,
            data_compra=data_compra,
            valor_total=valor_total,
            vendedor_id=cliente.vendedor_id,
            empresa_id=cliente.empresa_id,
        )
        db.session.add(compra)
    
    # Atualizar meta do vendedor com o novo valor e depois
    # persistir tanto a compra quanto a meta em uma única transação
    _update_meta_for_compra(cliente.vendedor_id, valor_total, data_compra)
    db.session.commit()

    return compra

# ===== ROTAS DE METAS =====

@app.route("/metas/importar", methods=["GET", "POST"])
@login_required
def importar_metas():
    """Importar metas via planilha Excel - Admin, Supervisor, RH e Gerente"""
    # Verificar permissões: admin, supervisor, RH e gerente podem importar metas
    if not pode_importar(current_user, "metas"):
        flash(
            "Você não tem permissão para importar metas. Apenas Administradores, Supervisores, RH e Gerentes podem realizar esta ação.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        # Validar arquivo Excel
        arquivo, erro = validar_arquivo_excel(request)
        if erro:
            flash(erro, "danger")
            return redirect(request.url)

        # Verificar disponibilidade do Excel
        if not EXCEL_AVAILABLE and not ensure_excel_available():
            flash("Erro: Bibliotecas Excel não instaladas. Contate o administrador.", "danger")
            return redirect(request.url)

        try:
            # Ler arquivo Excel
            df = pd.read_excel(arquivo)

            # Validar colunas obrigatórias
            colunas_obrigatorias = [
                "Vendedor Email",
                "Mês",
                "Ano",
                "Meta Vendas",
            ]
            colunas_faltando = [
                col for col in colunas_obrigatorias if col not in df.columns
            ]

            if colunas_faltando:
                flash(
                    f'Colunas obrigatórias faltando: {", ".join(colunas_faltando)}',
                    "danger",
                )
                return redirect(request.url)

            # Processar cada linha
            erros = []
            sucesso = 0

            for index, row in df.iterrows():
                try:
                    # Buscar vendedor por email
                    vendedor = Vendedor.query.filter_by(
                        email=row["Vendedor Email"]
                    ).first()

                    if not vendedor:
                        erros.append(f"Linha {index + 2}: Vendedor com email {row['Vendedor Email']} não encontrado")
                        continue

                    # Verificar se usuário pode criar meta para este vendedor
                    if (
                        not current_user.is_super_admin and
                        vendedor.empresa_id != current_user.empresa_id
                    ):
                        erros.append(f"Linha {index + 2}: Sem permissão para criar meta para este vendedor")
                        continue

                    # Verificar se meta já existe
                    meta_existente = Meta.query.filter_by(
                        vendedor_id=vendedor.id,
                        mes=int(row["Mês"]),
                        ano=int(row["Ano"]),
                    ).first()

                    if meta_existente:
                        # ATUALIZAR META EXISTENTE (Upsert)
                        meta_existente.valor_meta = float(row["Meta Vendas"])
                        sucesso += 1
                        continue

                    # Criar meta
                    meta = Meta(
                        vendedor_id=vendedor.id,
                        mes=int(row["Mês"]),
                        ano=int(row["Ano"]),
                        valor_meta=float(row["Meta Vendas"]),
                    )

                    db.session.add(meta)
                    sucesso += 1

                except Exception as e:
                    erros.append(
                        f"Linha {index + 2}: Erro ao processar - {str(e)}"
                    )
                    continue

            if erros:
                db.session.rollback()
                flash(
                    f"Erros encontrados durante importação. Nenhuma meta foi importada.",
                    "danger",
                )
                for erro in erros[:10]:  # Mostrar até 10 erros
                    flash(erro, "warning")
                if len(erros) > 10:
                    flash(f"... e mais {len(erros) - 10} erro(s)", "warning")
                return redirect(request.url)

            db.session.commit()
            flash(f"{sucesso} meta(s) processada(s) (criadas ou atualizadas) com sucesso!", "success")
            return redirect(url_for("lista_metas"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao processar arquivo: {str(e)}", "danger")
            return redirect(request.url)

    return render_template("metas/importar.html")

@app.route("/metas")
@login_required
def lista_metas():
    """Lista todas as metas"""
    mes = request.args.get("mes", datetime.now().month, type=int)
    ano = request.args.get("ano", datetime.now().year, type=int)
    ordenar = request.args.get("ordenar", "vendas", type=str)
    page = request.args.get("page", 1, type=int)

    # Quantidade de registros por página para manter a tela leve
    per_page = 20

    # Filtrar metas por empresa (exceto super admin)
    if current_user.is_super_admin:
        query = Meta.query.filter_by(mes=mes, ano=ano).join(Vendedor)
    else:
        query = (
            Meta.query.filter_by(mes=mes, ano=ano)
            .join(Vendedor)
            .filter(Vendedor.empresa_id == current_user.empresa_id)
        )

    # Aplicar ordenação
    if ordenar == "vendas":
        # Ordenar por receita alcançada (maior primeiro)
        metas_ordenadas = query.order_by(Meta.receita_alcancada.desc()).all()
    elif ordenar == "manutencao":
        # Ordenar por percentual de alcance (maior primeiro)
        metas_todas = query.all()
        metas_ordenadas = sorted(
            metas_todas,
            key=lambda m: (
                (m.receita_alcancada / m.valor_meta * 100)
                if m.valor_meta > 0
                else 0
            ),
            reverse=True,
        )
    else:
        metas_ordenadas = query.all()

    # Calcular totais globais (todas as metas do período)
    total_meta = sum(m.valor_meta for m in metas_ordenadas)
    total_receita = sum(m.receita_alcancada for m in metas_ordenadas)
    total_comissao = sum(m.comissao_total for m in metas_ordenadas)

    # Paginação em memória (lista já está ordenada)
    total_registros = len(metas_ordenadas)
    total_paginas = (total_registros + per_page - 1) // per_page if total_registros else 1

    # Garantir que a página esteja dentro dos limites
    if page < 1:
        page = 1
    if page > total_paginas:
        page = total_paginas

    start = (page - 1) * per_page
    end = start + per_page
    metas = metas_ordenadas[start:end]

    pagination = {
        "page": page,
        "per_page": per_page,
        "total": total_registros,
        "pages": total_paginas,
        "has_prev": page > 1,
        "has_next": page < total_paginas,
        "prev_num": page - 1,
        "next_num": page + 1 if page < total_paginas else None,
    }
    # Faixas dinâmicas de comissão para exibição na legenda
    if current_user.is_super_admin:
        faixas_vendedor = (
            FaixaComissaoVendedor.query.filter(
                FaixaComissaoVendedor.empresa_id.is_(None)
            ).order_by(FaixaComissaoVendedor.ordem).all()
        )
    else:
        faixas_vendedor = (
            FaixaComissaoVendedor.query.filter_by(
                empresa_id=current_user.empresa_id
            ).order_by(FaixaComissaoVendedor.ordem).all()
        )
        # fallback globais
        if not faixas_vendedor:
            faixas_vendedor = (
                FaixaComissaoVendedor.query.filter(
                    FaixaComissaoVendedor.empresa_id.is_(None)
                ).order_by(FaixaComissaoVendedor.ordem).all()
            )

    return render_template(
        "metas/lista.html",
        metas=metas,
        mes=mes,
        ano=ano,
        ordenar=ordenar,
        total_meta=total_meta,
        total_receita=total_receita,
        total_comissao=total_comissao,
        pagination=pagination,
        faixas_vendedor=faixas_vendedor,
    )

@app.route("/metas/nova", methods=["GET", "POST"])
@login_required
def nova_meta():
    """Cadastrar nova meta"""
    # Apenas admin, gerente e supervisor podem criar metas
    if current_user.cargo not in ["admin", "gerente", "supervisor"]:
        flash(
            "Você não tem permissão para criar metas. Apenas Administradores, Gerentes e Supervisores podem realizar esta ação.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    form = MetaForm()

    # Preencher choices de vendedores da mesma empresa
    if current_user.is_super_admin:
        vendedores = Vendedor.query.filter_by(ativo=True).all()
    else:
        vendedores = Vendedor.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        ).all()
    form.vendedor_id.choices = [(v.id, v.nome) for v in vendedores]

    if form.validate_on_submit():
        # Verificar se já existe meta para este vendedor neste período
        meta_existente = Meta.query.filter_by(
            vendedor_id=form.vendedor_id.data,
            mes=form.mes.data,
            ano=form.ano.data,
        ).first()

        if meta_existente:
            msg = "Já existe uma meta para este vendedor " "neste período!"
            flash(msg, "warning")
            return render_template(
                "metas/form.html", form=form, titulo="Nova Meta"
            )

        meta = Meta(
            vendedor_id=form.vendedor_id.data,
            mes=form.mes.data,
            ano=form.ano.data,
            valor_meta=form.valor_meta.data,
            receita_alcancada=form.receita_alcancada.data,
            status_comissao=form.status_comissao.data,
            observacoes=form.observacoes.data,
        )

        # Calcular comissão
        meta.calcular_comissao()

        db.session.add(meta)
        db.session.commit()

        flash("Meta cadastrada com sucesso!", "success")
        return redirect(url_for("lista_metas"))

    return render_template("metas/form.html", form=form, titulo="Nova Meta")

@app.route("/metas/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_meta(id):
    """Editar meta existente"""
    # Apenas admin, gerente e supervisor podem editar metas
    if current_user.cargo not in ["admin", "gerente", "supervisor"]:
        flash(
            "Você não tem permissão para editar metas. Apenas Administradores, Gerentes e Supervisores podem realizar esta ação.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    meta = Meta.query.get_or_404(id)

    # Verificar se a meta pertence à empresa do usuário
    # (exceto super admin)
    if not current_user.is_super_admin:
        if meta.vendedor.empresa_id != current_user.empresa_id:
            msg = "Você não tem permissão " "para editar esta meta."
            flash(msg, "danger")
            return redirect(url_for("lista_metas"))

    form = MetaForm(obj=meta)

    # Preencher choices de vendedores da mesma empresa
    if current_user.is_super_admin:
        vendedores = Vendedor.query.filter_by(ativo=True).all()
    else:
        vendedores = Vendedor.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        ).all()
    form.vendedor_id.choices = [(v.id, v.nome) for v in vendedores]

    if form.validate_on_submit():
        # Verificar se há outra meta para este vendedor
        # neste período (exceto a atual)
        meta_existente = (
            Meta.query.filter_by(
                vendedor_id=form.vendedor_id.data,
                mes=form.mes.data,
                ano=form.ano.data,
            )
            .filter(Meta.id != id)
            .first()
        )

        if meta_existente:
            msg = "Já existe outra meta para este vendedor " "neste período!"
            flash(msg, "warning")
            return render_template(
                "metas/form.html", form=form, meta=meta, titulo="Editar Meta"
            )

        meta.vendedor_id = form.vendedor_id.data
        meta.mes = form.mes.data
        meta.ano = form.ano.data
        meta.valor_meta = form.valor_meta.data
        meta.receita_alcancada = form.receita_alcancada.data
        meta.status_comissao = form.status_comissao.data
        meta.observacoes = form.observacoes.data

        # Recalcular comissão
        meta.calcular_comissao()

        db.session.commit()
        flash("Meta atualizada com sucesso!", "success")
        return redirect(url_for("lista_metas"))

    return render_template(
        "metas/form.html", form=form, titulo="Editar Meta", meta=meta
    )

@app.route("/metas/<int:id>/deletar", methods=["POST"])
@login_required
def deletar_meta(id):
    """Deletar meta"""
    # Apenas admin, gerente e supervisor podem deletar metas
    if current_user.cargo not in ["admin", "gerente", "supervisor"]:
        flash(
            "Você não tem permissão para deletar metas. Apenas Administradores, Gerentes e Supervisores podem realizar esta ação.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    meta = Meta.query.get_or_404(id)

    # Verificar se a meta pertence à empresa do usuário
    # (exceto super admin)
    if not current_user.is_super_admin:
        if meta.vendedor.empresa_id != current_user.empresa_id:
            msg = "Você não tem permissão " "para deletar esta meta."
            flash(msg, "danger")
            return redirect(url_for("lista_metas"))

    db.session.delete(meta)
    db.session.commit()
    flash("Meta deletada com sucesso!", "success")
    return redirect(url_for("lista_metas"))

@app.route("/metas/exportar-pdf")
@login_required
def exportar_pdf_metas():
    """Exportar relatório de metas em PDF"""
    mes = request.args.get("mes", datetime.now().month, type=int)
    ano = request.args.get("ano", datetime.now().year, type=int)

    query = Meta.query.filter_by(mes=mes, ano=ano).join(Vendedor)

    if not current_user.is_super_admin:
        query = query.filter(Vendedor.empresa_id == current_user.empresa_id)

    metas = query.all()

    try:
        pdf_buffer = gerar_pdf_metas(metas, mes, ano)
        if not pdf_buffer:
            raise RuntimeError("Geração de PDF retornou buffer vazio")

        meses = [
            "Janeiro",
            "Fevereiro",
            "Março",
            "Abril",
            "Maio",
            "Junho",
            "Julho",
            "Agosto",
            "Setembro",
            "Outubro",
            "Novembro",
            "Dezembro",
        ]
        filename = f"Relatorio_Metas_{meses[mes - 1]}_{ano}.pdf"

        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        app.logger.error(f"Erro ao gerar PDF de metas: {e}", exc_info=True)
        flash(
            "Não foi possível gerar o PDF. Verifique os logs e tente novamente.",
            "danger",
        )
        return redirect(url_for("lista_metas"))

@app.route("/dashboard/exportar-pdf")
@login_required
def exportar_pdf_dashboard():
    """Exportar relatório do dashboard em PDF"""
    # Calcular resumo global
    mes_atual = datetime.now().month
    ano_atual = datetime.now().year

    query = Meta.query.filter_by(mes=mes_atual, ano=ano_atual).join(Vendedor)

    if not current_user.is_super_admin:
        query = query.filter(Vendedor.empresa_id == current_user.empresa_id)

    metas_mes = query.all()

    # Filtrar vendedores ativos da empresa
    if current_user.is_super_admin:
        total_vendedores = Vendedor.query.filter_by(ativo=True).count()
    else:
        total_vendedores = Vendedor.query.filter_by(
            ativo=True, empresa_id=current_user.empresa_id
        ).count()

    resumo = {
        "total_vendedores": total_vendedores,
        "receita_total": sum(m.receita_alcancada for m in metas_mes),
        "meta_total": sum(m.valor_meta for m in metas_mes),
        "comissao_total": sum(m.comissao_total for m in metas_mes),
    }

    # Buscar TODOS os vendedores com dados completos
    hoje = datetime.now()
    vendedores = []
    equipes_dict = {}
    supervisores_dict = {}

    metas_sorted = sorted(
        metas_mes, key=lambda m: m.receita_alcancada, reverse=True
    )

    for meta in metas_sorted:
        # Tratar supervisor com segurança
        try:
            supervisor_nome = (
                meta.vendedor.supervisor.nome
                if meta.vendedor.supervisor
                else "Sem supervisor"
            )
        except AttributeError:
            supervisor_nome = "Sem supervisor"

        # Tratar equipe com segurança
        try:
            equipe_nome = (
                meta.vendedor.equipe_obj.nome
                if meta.vendedor.equipe_obj
                else "Sem Equipe"
            )
        except AttributeError:
            equipe_nome = "Sem Equipe"

        # Calcular projeção do vendedor
        projecao_vendedor = calcular_projecao_mes(
            receita_atual=meta.receita_alcancada,
            meta_mes=meta.valor_meta,
            ano=ano_atual,
            mes=mes_atual,
            dia_atual=hoje.day,
        )

        vendedor_data = {
            "nome": meta.vendedor.nome,
            "supervisor": supervisor_nome,
            "equipe": equipe_nome,
            "receita": meta.receita_alcancada,
            "meta": meta.valor_meta,
            "percentual": meta.percentual_alcance,
            "comissao": meta.comissao_total,
            "projecao": projecao_vendedor,
        }
        vendedores.append(vendedor_data)

        # Agrupar por equipes
        if equipe_nome not in equipes_dict:
            equipes_dict[equipe_nome] = {
                "nome": equipe_nome,
                "vendedores_count": 0,
                "receita_total": 0.0,
                "meta_total": 0.0,
            }
        equipes_dict[equipe_nome]["vendedores_count"] += 1
        equipes_dict[equipe_nome]["receita_total"] += meta.receita_alcancada
        equipes_dict[equipe_nome]["meta_total"] += meta.valor_meta

        # Agrupar por supervisores
        if supervisor_nome not in supervisores_dict:
            supervisores_dict[supervisor_nome] = {
                "nome": supervisor_nome,
                "vendedores_count": 0,
                "receita_total": 0.0,
                "meta_total": 0.0,
            }
        supervisores_dict[supervisor_nome]["vendedores_count"] += 1
        supervisores_dict[supervisor_nome][
            "receita_total"
        ] += meta.receita_alcancada
        supervisores_dict[supervisor_nome]["meta_total"] += meta.valor_meta

    # Calcular projeções e percentuais para equipes
    equipes = []
    for eq in equipes_dict.values():
        eq["percentual_alcance"] = (
            (eq["receita_total"] / eq["meta_total"] * 100)
            if eq["meta_total"] > 0
            else 0
        )
        eq["projecao"] = calcular_projecao_mes(
            receita_atual=eq["receita_total"],
            meta_mes=eq["meta_total"],
            ano=ano_atual,
            mes=mes_atual,
            dia_atual=hoje.day,
        )
        equipes.append(eq)
    equipes.sort(key=lambda x: x["percentual_alcance"], reverse=True)

    # Calcular projeções e percentuais para supervisores
    supervisores = []
    for sup in supervisores_dict.values():
        sup["percentual_alcance"] = (
            (sup["receita_total"] / sup["meta_total"] * 100)
            if sup["meta_total"] > 0
            else 0
        )
        sup["projecao"] = calcular_projecao_mes(
            receita_atual=sup["receita_total"],
            meta_mes=sup["meta_total"],
            ano=ano_atual,
            mes=mes_atual,
            dia_atual=hoje.day,
        )
        # Comissão do supervisor baseada nas faixas de supervisor
        taxa_sup = _obter_taxa_por_alcance(
            "supervisor",
            current_user.empresa_id if current_user.cargo != "super_admin" else None,
            sup["percentual_alcance"],
        )
        sup["taxa_supervisor"] = taxa_sup
        sup["comissao_supervisor"] = sup["receita_total"] * taxa_sup
        supervisores.append(sup)
    supervisores.sort(key=lambda x: x["percentual_alcance"], reverse=True)

    # Calcular projeção global
    total_receita = resumo["receita_total"]
    total_meta = resumo["meta_total"]
    projecao_global = calcular_projecao_mes(
        receita_atual=total_receita,
        meta_mes=total_meta,
        ano=ano_atual,
        mes=mes_atual,
        dia_atual=hoje.day,
    )
    resumo["alcance_geral"] = (
        (total_receita / total_meta * 100) if total_meta > 0 else 0
    )
    resumo["projecao_global"] = projecao_global

    try:
        # Log dos dados que serão enviados ao PDF
        app.logger.info(f"Gerando PDF Dashboard - Vendedores: {len(vendedores)}, Equipes: {len(equipes)}, Supervisores: {len(supervisores)}")

        pdf_buffer = gerar_pdf_dashboard(
            resumo, vendedores, mes_atual, ano_atual, equipes, supervisores
        )

        if not pdf_buffer:
            app.logger.error("PDF buffer retornou None")
            raise RuntimeError("Geração de PDF retornou buffer vazio")

        meses = [
            "Janeiro",
            "Fevereiro",
            "Março",
            "Abril",
            "Maio",
            "Junho",
            "Julho",
            "Agosto",
            "Setembro",
            "Outubro",
            "Novembro",
            "Dezembro",
        ]
        filename = f"Dashboard_Completo_{meses[mes_atual - 1]}_{ano_atual}.pdf"

        app.logger.info(f"PDF gerado com sucesso: {filename}")

        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        app.logger.error(f"Erro ao gerar PDF do dashboard: {str(e)}", exc_info=True)
        flash(f"Erro ao gerar PDF: {str(e)}", "danger")
        return redirect(url_for("dashboard"))

# ===== API ROUTES =====

@app.route("/api/ranking")
@login_required
def api_ranking():
    """API: Retorna ranking de vendedores"""
    mes = request.args.get("mes", datetime.now().month, type=int)
    ano = request.args.get("ano", datetime.now().year, type=int)

    query = Meta.query.filter_by(mes=mes, ano=ano).join(Vendedor)

    if not current_user.is_super_admin:
        query = query.filter(Vendedor.empresa_id == current_user.empresa_id)

    metas = query.all()

    dados = []
    for meta in metas:
        dados.append(
            {
                "id": meta.id,
                "nome": meta.vendedor.nome,
                "supervisor": meta.vendedor.supervisor_nome,
                "meta": meta.valor_meta,
                "receita_alcancada": meta.receita_alcancada,
                "percentual_alcance": meta.percentual_alcance,
                "comissao_total": meta.comissao_total,
                "status_comissao": meta.status_comissao,
            }
        )

    dados.sort(key=lambda x: x["percentual_alcance"], reverse=True)
    return jsonify(dados)

@app.route("/api/vendedor/<int:vendedor_id>/supervisor")
@login_required
def api_vendedor_supervisor(vendedor_id):
    """API: Retorna o nome do supervisor do vendedor"""
    try:
        vendedor = Vendedor.query.get_or_404(vendedor_id)

        # Verificar permissão
        if (
            not current_user.is_super_admin and
            vendedor.empresa_id != current_user.empresa_id
        ):
            return jsonify({"error": "Sem permissão"}), 403

        supervisor_nome = ""
        if vendedor.supervisor:
            supervisor_nome = vendedor.supervisor.nome

        return jsonify(
            {
                "vendedor_id": vendedor.id,
                "vendedor_nome": vendedor.nome,
                "supervisor": supervisor_nome,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ===== FUNÇÕES AUXILIARES PARA EQUIPES =====

def _prepare_equipe_form(form):
    """Prepara o formulário de equipe, populando os supervisores e vendedores."""
    supervisores = Usuario.query.filter(
        Usuario.cargo == "supervisor",
        Usuario.ativo == True,
        Usuario.empresa_id == current_user.empresa_id,
    ).all()
    vendedores = Vendedor.query.filter(
        Vendedor.ativo == True, Vendedor.empresa_id == current_user.empresa_id
    ).all()

    form.supervisor_id.choices = [(s.id, s.nome) for s in supervisores]
    form.vendedores.choices = [(v.id, v.nome) for v in vendedores]

def _save_equipe_from_form(form, equipe=None):
    """Salva uma equipe (nova ou existente) a partir dos dados do formulário."""
    if not current_user.is_super_admin and form.empresa_id.data != current_user.empresa_id:
        raise PermissionError("Você só pode gerenciar equipes da sua empresa.")

    if equipe:  # Edição
        equipe.nome = form.nome.data
        equipe.supervisor_id = form.supervisor_id.data
        equipe.empresa_id = form.empresa_id.data
    else:  # Criação
        equipe = Equipe(
            nome=form.nome.data,
            supervisor_id=form.supervisor_id.data,
            empresa_id=form.empresa_id.data,
        )
        db.session.add(equipe)
        db.session.flush()  # Garante que o ID da equipe esteja disponível

    # Atualizar vendedores da equipe
    vendedores_selecionados = Vendedor.query.filter(
        Vendedor.id.in_(form.vendedores.data)
    ).all()
    equipe.vendedores = vendedores_selecionados
    
    db.session.commit()
    return equipe

# ===== ROTAS DE EQUIPES =====

@app.route("/equipes")
@login_required
def lista_equipes():
    """Lista todas as equipes"""
    page = request.args.get("page", 1, type=int)

    # Filtrar equipes por empresa (exceto super admin)
    if current_user.is_super_admin:
        base_query = Equipe.query.filter_by(ativa=True)
    else:
        base_query = Equipe.query.filter_by(
            empresa_id=current_user.empresa_id, ativa=True
        )

    # Estatísticas globais
    total_equipes = base_query.count()
    total_vendedores_alocados = 0

    # Para contar vendedores por equipe globalmente, usamos uma lista de IDs
    equipe_ids = [e.id for e in base_query.with_entities(Equipe.id).all()]
    if equipe_ids:
        total_vendedores_alocados = Vendedor.query.filter(
            Vendedor.ativo.is_(True), Vendedor.equipe_id.in_(equipe_ids)
        ).count()

    stats = {
        "total_equipes": total_equipes,
        "total_vendedores": total_vendedores_alocados,
    }

    # Paginação ordenada por nome de equipe
    pagination = base_query.order_by(Equipe.nome).paginate(
        page=page, per_page=15, error_out=False
    )
    equipes_pagina = pagination.items

    # Adiciona estatísticas de cada equipe (apenas da página)
    equipes_data = []
    for equipe in equipes_pagina:
        vendedores_count = Vendedor.query.filter_by(
            equipe_id=equipe.id, ativo=True
        ).count()
        equipes_data.append(
            {
                "equipe": equipe,
                "vendedores_count": vendedores_count,
                "supervisor": Usuario.query.get(equipe.supervisor_id),
            }
        )

    return render_template(
        "equipes/lista.html",
        equipes=equipes_data,
        pagination=pagination,
        stats=stats,
    )

@app.route("/equipes/nova", methods=["GET", "POST"])
@login_required
def nova_equipe():
    """Cadastrar nova equipe"""
    form = EquipeForm()

    # Preencher choices de supervisores da mesma empresa
    if current_user.is_super_admin:
        supervisores = Usuario.query.filter(
            Usuario.cargo.in_(["supervisor", "supervisor_manutencao"]),
            Usuario.ativo.is_(True),
        ).all()
    else:
        supervisores = Usuario.query.filter(
            Usuario.empresa_id == current_user.empresa_id,
            Usuario.cargo.in_(["supervisor", "supervisor_manutencao"]),
            Usuario.ativo.is_(True),
        ).all()
    form.supervisor_id.choices = [(s.id, s.nome) for s in supervisores]

    if form.validate_on_submit():
        try:
            equipe = Equipe(
                nome=form.nome.data,
                descricao=form.descricao.data,
                supervisor_id=form.supervisor_id.data,
                empresa_id=current_user.empresa_id,
            )

            db.session.add(equipe)
            db.session.commit()

            flash(f"Equipe {equipe.nome} cadastrada com sucesso!", "success")
            return redirect(url_for("lista_equipes"))
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar equipe: {str(e)}", "danger")

    return render_template(
        "equipes/form.html", form=form, titulo="Nova Equipe"
    )

@app.route("/equipes/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_equipe(id):
    """Editar equipe existente"""
    try:
        equipe = Equipe.query.get_or_404(id)

        # Verificar se a equipe pertence à empresa do usuário
        # (exceto super admin)
        if not current_user.is_super_admin:
            if equipe.empresa_id != current_user.empresa_id:
                msg = "Você não tem permissão " "para editar esta equipe."
                flash(msg, "danger")
                return redirect(url_for("lista_equipes"))

        form = EquipeForm(equipe_id=id, obj=equipe)

        # Preencher choices de supervisores da mesma empresa
        if current_user.is_super_admin:
            supervisores = Usuario.query.filter(
                Usuario.cargo.in_(["supervisor", "supervisor_manutencao"]),
                Usuario.ativo.is_(True),
            ).all()
        else:
            supervisores = Usuario.query.filter(
                Usuario.empresa_id == current_user.empresa_id,
                Usuario.cargo.in_(["supervisor", "supervisor_manutencao"]),
                Usuario.ativo.is_(True),
            ).all()
        form.supervisor_id.choices = [(s.id, s.nome) for s in supervisores]

        if form.validate_on_submit():
            try:
                equipe.nome = form.nome.data
                equipe.descricao = form.descricao.data
                equipe.supervisor_id = form.supervisor_id.data
                # Nota: empresa_id não deve ser alterado na edição

                db.session.commit()
                flash(f"Equipe {equipe.nome} atualizada com sucesso!", "success")
                return redirect(url_for("lista_equipes"))
            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao atualizar equipe: {str(e)}", "danger")

        return render_template(
            "equipes/form.html",
            form=form,
            equipe=equipe,
            titulo="Editar Equipe",
        )
    except Exception as e:
        flash(f"Equipe não encontrada. {str(e)}", "danger")
        return redirect(url_for("lista_equipes"))

@app.route("/equipes/<int:id>/deletar", methods=["POST"])
@login_required
def deletar_equipe(id):
    """Deletar equipe (soft delete)"""
    try:
        equipe = Equipe.query.get_or_404(id)

        # Verificar se a equipe pertence à empresa do usuário
        # (exceto super admin)
        if not current_user.is_super_admin:
            if equipe.empresa_id != current_user.empresa_id:
                msg = "Você não tem permissão " "para deletar esta equipe."
                flash(msg, "danger")
                return redirect(url_for("lista_equipes"))

        nome_equipe = equipe.nome
        equipe.ativa = False
        db.session.commit()

        flash(f"Equipe {nome_equipe} desativada com sucesso!", "success")
        return redirect(url_for("lista_equipes"))
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao desativar equipe: {str(e)}", "danger")
        return redirect(url_for("lista_equipes"))

@app.route("/equipes/<int:id>/detalhes")
@login_required
def detalhes_equipe(id):
    """Ver detalhes da equipe com vendedores e metas"""
    equipe = Equipe.query.get_or_404(id)

    # Verificar se a equipe pertence à empresa do usuário
    # (exceto super admin)
    if not current_user.is_super_admin:
        if equipe.empresa_id != current_user.empresa_id:
            msg = "Você não tem permissão " "para visualizar esta equipe."
            flash(msg, "danger")
            return redirect(url_for("lista_equipes"))

    vendedores = Vendedor.query.filter_by(
        equipe_id=equipe.id, ativo=True
    ).all()
    supervisor = Usuario.query.get(equipe.supervisor_id)

    # Buscar metas do mês atual para os vendedores da equipe
    mes_atual = datetime.now().month
    ano_atual = datetime.now().year

    vendedores_data = []
    for vendedor in vendedores:
        meta = Meta.query.filter_by(
            vendedor_id=vendedor.id, mes=mes_atual, ano=ano_atual
        ).first()

        vendedores_data.append({"vendedor": vendedor, "meta": meta})

    return render_template(
        "equipes/detalhes.html",
        equipe=equipe,
        supervisor=supervisor,
        vendedores=vendedores_data,
        mes=mes_atual,
        ano=ano_atual,
    )

# ===== CONFIGURAÇÕES DE COMISSÃO =====

# Utilitário: recalcular comissões das metas do mês atual da empresa
def _recalcular_comissoes_mes_atual_empresa(empresa_id: int):
    try:
        if not empresa_id:
            return
        hoje = datetime.utcnow()
        metas = (
            Meta.query.join(Vendedor)
            .filter(
                Vendedor.empresa_id == empresa_id,
                Meta.mes == hoje.month,
                Meta.ano == hoje.year,
            )
            .all()
        )
        for m in metas:
            m.calcular_comissao()
        db.session.commit()
    except Exception as e:
        app.logger.error(
            f"Erro ao recalcular comissões (empresa_id={empresa_id}): {e}"
        )

# Utilitário: obter taxa de comissão por percentual de alcance a partir das faixas
def _obter_taxa_por_alcance(tipo: str, empresa_id: int | None, percentual_alcance: float) -> float:
    try:
        modelo = FaixaComissaoSupervisor if tipo == "supervisor" else FaixaComissaoVendedor
        query = modelo.query.filter_by(ativa=True)
        faixas = []
        if empresa_id:
            faixas = query.filter_by(empresa_id=empresa_id).order_by(modelo.alcance_min).all()
        if not faixas:
            faixas = (
                modelo.query.filter(
                    modelo.empresa_id.is_(None),
                    modelo.ativa.is_(True),
                ).order_by(modelo.alcance_min).all()
            )
        if not faixas:
            return 0.0
        perc = percentual_alcance or 0.0
        faixas_ordenadas = sorted(faixas, key=lambda f: (f.alcance_min or 0))
        taxa = faixas_ordenadas[-1].taxa_comissao
        for f in faixas_ordenadas:
            if perc <= (f.alcance_max or perc):
                taxa = f.taxa_comissao
                break
        return float(taxa or 0.0)
    except Exception as e:
        app.logger.error(f"Erro ao obter taxa por alcance ({tipo}): {e}")
        return 0.0

@app.route("/configuracoes/comissoes")
@login_required
def configuracoes_comissoes():
    """Lista as faixas de comissão configuradas (vendedores e supervisores)"""
    if current_user.cargo not in ["admin", "super_admin"]:
        flash("Acesso negado. Apenas administradores podem acessar.", "danger")
        return redirect(url_for("dashboard"))

    # Busca faixas da empresa ou globais para VENDEDORES
    if current_user.cargo == "super_admin":
        faixas_vendedor = (
            FaixaComissaoVendedor.query.filter(
                FaixaComissaoVendedor.empresa_id.is_(None)
            )
            .order_by(FaixaComissaoVendedor.ordem)
            .all()
        )
        faixas_supervisor = (
            FaixaComissaoSupervisor.query.filter(
                FaixaComissaoSupervisor.empresa_id.is_(None)
            )
            .order_by(FaixaComissaoSupervisor.ordem)
            .all()
        )
    else:
        faixas_vendedor = (
            FaixaComissaoVendedor.query.filter_by(
                empresa_id=current_user.empresa_id
            )
            .order_by(FaixaComissaoVendedor.ordem)
            .all()
        )

        faixas_supervisor = (
            FaixaComissaoSupervisor.query.filter_by(
                empresa_id=current_user.empresa_id
            )
            .order_by(FaixaComissaoSupervisor.ordem)
            .all()
        )

        # Se não houver faixas customizadas, usa as globais
        if not faixas_vendedor:
            faixas_vendedor = (
                FaixaComissaoVendedor.query.filter(
                    FaixaComissaoVendedor.empresa_id.is_(None)
                )
                .order_by(FaixaComissaoVendedor.ordem)
                .all()
            )

        if not faixas_supervisor:
            faixas_supervisor = (
                FaixaComissaoSupervisor.query.filter(
                    FaixaComissaoSupervisor.empresa_id.is_(None)
                )
                .order_by(FaixaComissaoSupervisor.ordem)
                .all()
            )

    # Flag de sincronização automática (por empresa; super_admin usa chave global)
    empresa_contexto = None if current_user.cargo == "super_admin" else current_user.empresa_id
    sincronizacao_automatica = _get_sync_flag(empresa_contexto)

    return render_template(
        "configuracoes/comissoes.html",
        faixas_vendedor=faixas_vendedor,
        faixas_supervisor=faixas_supervisor,
        sincronizacao_automatica=sincronizacao_automatica,
    )

@app.route("/configuracoes/comissoes/criar", methods=["GET", "POST"])
@login_required
def criar_faixa_comissao():
    """Cria nova faixa de comissão"""
    if current_user.cargo not in ["admin", "super_admin"]:
        flash("Acesso negado.", "danger")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        try:
            # Validação de dados
            # 'vendedor' ou 'supervisor'
            tipo = request.form.get("tipo", "vendedor")
            alcance_min = request.form.get("alcance_min", "0")
            alcance_max = request.form.get("alcance_max", "100")
            taxa_comissao = request.form.get("taxa_comissao", "1")
            # Sincronização automática: opcional por configuração
            empresa_id_atual = (
                current_user.empresa_id
                if current_user.cargo != "super_admin"
                else None
            )
            copiar_para_outro = _get_sync_flag(empresa_id_atual)

            # Conversão segura
            alcance_min = float(alcance_min) if alcance_min else 0.0
            alcance_max = float(alcance_max) if alcance_max else 100.0
            taxa = float(taxa_comissao) / 100 if taxa_comissao else 0.01
            cor = request.form.get("cor", "primary")
            ordem = (
                int(request.form.get("ordem", "0"))
                if request.form.get("ordem")
                else 0
            )

            # Validações
            if alcance_min >= alcance_max:
                flash(
                    "O alcance mínimo deve ser menor que o máximo.", "danger"
                )
                return redirect(url_for("criar_faixa_comissao"))

            if taxa <= 0:
                flash("A taxa de comissão deve ser maior que zero.", "danger")
                return redirect(url_for("criar_faixa_comissao"))

            # empresa_id_atual já definido

            # Cria faixa do tipo selecionado
            if tipo == "vendedor":
                nova_faixa = FaixaComissaoVendedor(
                    empresa_id=empresa_id_atual,
                    alcance_min=alcance_min,
                    alcance_max=alcance_max,
                    taxa_comissao=taxa,
                    cor=cor,
                    ordem=ordem,
                )
            else:
                nova_faixa = FaixaComissaoSupervisor(
                    empresa_id=empresa_id_atual,
                    alcance_min=alcance_min,
                    alcance_max=alcance_max,
                    taxa_comissao=taxa,
                    cor=cor,
                    ordem=ordem,
                )

            db.session.add(nova_faixa)

            # Se solicitado, copia para o outro tipo
            if copiar_para_outro:
                if tipo == "vendedor":
                    faixa_copiada = nova_faixa.copiar_para_supervisor(
                        empresa_id_atual
                    )
                else:
                    faixa_copiada = nova_faixa.copiar_para_vendedor(
                        empresa_id_atual
                    )
                db.session.add(faixa_copiada)

            db.session.commit()

            tipo_nome = "vendedor" if tipo == "vendedor" else "supervisor"
            msg = f"Faixa de comissão para {tipo_nome} criada com sucesso!"
            if copiar_para_outro:
                outro_tipo = "supervisor" if tipo == "vendedor" else "vendedor"
                msg += f" (copiada também para {outro_tipo})"
            flash(msg, "success")
            # Recalcular comissões das metas do mês atual (empresa do usuário)
            if current_user.cargo != "super_admin":
                _recalcular_comissoes_mes_atual_empresa(current_user.empresa_id)
            return redirect(url_for("configuracoes_comissoes"))

        except ValueError as e:
            db.session.rollback()
            flash(f"Erro nos valores informados: {str(e)}", "danger")
            return redirect(url_for("criar_faixa_comissao"))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Erro ao criar faixa de comissão: {str(e)}")
            flash(
                f"Erro ao criar faixa. Por favor, tente novamente.", "danger"
            )
            return redirect(url_for("criar_faixa_comissao"))

    return render_template(
        "configuracoes/comissao_form.html",
        faixa=None,
        tipo=None,
        sincronizacao_automatica=_get_sync_flag(None if current_user.cargo == "super_admin" else current_user.empresa_id),
    )

@app.route(
    "/configuracoes/comissoes/<string:tipo>/<int:id>/editar",
    methods=["GET", "POST"],
)
@login_required
def editar_faixa_comissao(tipo, id):
    """Edita faixa de comissão existente"""
    if current_user.cargo not in ["admin", "super_admin"]:
        flash("Acesso negado.", "danger")
        return redirect(url_for("dashboard"))

    # Busca a faixa do tipo correto
    if tipo == "vendedor":
        faixa = FaixaComissaoVendedor.query.get_or_404(id)
    elif tipo == "supervisor":
        faixa = FaixaComissaoSupervisor.query.get_or_404(id)
    else:
        flash("Tipo inválido.", "danger")
        return redirect(url_for("configuracoes_comissoes"))

    # Verifica permissão
    if (
        current_user.cargo != "super_admin" and
        faixa.empresa_id != current_user.empresa_id
    ):
        flash("Você não tem permissão para editar esta faixa.", "danger")
        return redirect(url_for("configuracoes_comissoes"))

    if request.method == "POST":
        try:
            # Validação de dados
            alcance_min = request.form.get("alcance_min", "0")
            alcance_max = request.form.get("alcance_max", "100")
            taxa_comissao = request.form.get("taxa_comissao", "1")
            # Sincronização automática: opcional por configuração
            empresa_id_atual = (
                current_user.empresa_id
                if current_user.cargo != "super_admin"
                else None
            )
            copiar_para_outro = _get_sync_flag(empresa_id_atual)

            # Conversão segura
            alcance_min = float(alcance_min) if alcance_min else 0.0
            alcance_max = float(alcance_max) if alcance_max else 100.0
            taxa = float(taxa_comissao) / 100 if taxa_comissao else 0.01
            cor = request.form.get("cor", "primary")
            ordem = (
                int(request.form.get("ordem", "0"))
                if request.form.get("ordem")
                else 0
            )

            # Validações
            if alcance_min >= alcance_max:
                flash(
                    "O alcance mínimo deve ser menor que o máximo.", "danger"
                )
                return redirect(
                    url_for("editar_faixa_comissao", tipo=tipo, id=id)
                )

            if taxa <= 0:
                flash("A taxa de comissão deve ser maior que zero.", "danger")
                return redirect(
                    url_for("editar_faixa_comissao", tipo=tipo, id=id)
                )

            # Atualiza faixa
            faixa.alcance_min = alcance_min
            faixa.alcance_max = alcance_max
            faixa.taxa_comissao = taxa
            faixa.cor = cor
            faixa.ordem = ordem

            # Se solicitado, copia/atualiza para o outro tipo
            if copiar_para_outro:
                if tipo == "vendedor":
                    # Procura se já existe uma faixa supervisor correspondente
                    faixa_outro = FaixaComissaoSupervisor.query.filter_by(
                        empresa_id=empresa_id_atual, ordem=ordem
                    ).first()
                    if faixa_outro:
                        faixa_outro.alcance_min = alcance_min
                        faixa_outro.alcance_max = alcance_max
                        faixa_outro.taxa_comissao = taxa
                        faixa_outro.cor = cor
                    else:
                        nova = faixa.copiar_para_supervisor(empresa_id_atual)
                        db.session.add(nova)
                else:
                    faixa_outro = FaixaComissaoVendedor.query.filter_by(
                        empresa_id=empresa_id_atual, ordem=ordem
                    ).first()
                    if faixa_outro:
                        faixa_outro.alcance_min = alcance_min
                        faixa_outro.alcance_max = alcance_max
                        faixa_outro.taxa_comissao = taxa
                        faixa_outro.cor = cor
                    else:
                        nova = faixa.copiar_para_vendedor(empresa_id_atual)
                        db.session.add(nova)

            db.session.commit()

            msg = "Faixa de comissão atualizada com sucesso!"
            if copiar_para_outro:
                outro_tipo = "supervisor" if tipo == "vendedor" else "vendedor"
                msg += f" (sincronizada com {outro_tipo})"
            flash(msg, "success")
            # Recalcular comissões das metas do mês atual (empresa do usuário)
            if current_user.cargo != "super_admin":
                _recalcular_comissoes_mes_atual_empresa(current_user.empresa_id)
            return redirect(url_for("configuracoes_comissoes"))

        except ValueError as e:
            db.session.rollback()
            flash(f"Erro nos valores informados: {str(e)}", "danger")
            return redirect(url_for("editar_faixa_comissao", tipo=tipo, id=id))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Erro ao atualizar faixa de comissão: {str(e)}")
            flash(
                f"Erro ao atualizar faixa. Por favor, tente novamente.",
                "danger",
            )
            return redirect(url_for("editar_faixa_comissao", tipo=tipo, id=id))

    return render_template(
        "configuracoes/comissao_form.html", faixa=faixa, tipo=tipo, sincronizacao_automatica=_get_sync_flag(None if current_user.cargo == "super_admin" else current_user.empresa_id)
    )

@app.route(
    "/configuracoes/comissoes/<string:tipo>/<int:id>/deletar", methods=["POST"]
)
@login_required
def deletar_faixa_comissao(tipo, id):
    """Deleta faixa de comissão"""
    if current_user.cargo not in ["admin", "super_admin"]:
        flash("Acesso negado.", "danger")
        return redirect(url_for("dashboard"))

    # Busca a faixa do tipo correto
    if tipo == "vendedor":
        faixa = FaixaComissaoVendedor.query.get_or_404(id)
    elif tipo == "supervisor":
        faixa = FaixaComissaoSupervisor.query.get_or_404(id)
    else:
        flash("Tipo inválido.", "danger")
        return redirect(url_for("configuracoes_comissoes"))

    # Verifica permissão
    if (
        current_user.cargo != "super_admin" and
        faixa.empresa_id != current_user.empresa_id
    ):
        flash("Você não tem permissão para deletar esta faixa.", "danger")
        return redirect(url_for("configuracoes_comissoes"))

    try:
        # Excluir faixa atual
        empresa_id_alvo = faixa.empresa_id
        ordem_alvo = faixa.ordem
        db.session.delete(faixa)

        # Sincronização automática de exclusão: remover faixa espelhada do outro tipo se habilitado
        remover_espelho = _get_sync_flag(
            None if current_user.cargo == "super_admin" else current_user.empresa_id
        )
        faixa_espelho = None
        if remover_espelho:
            if tipo == "vendedor":
                faixa_espelho = FaixaComissaoSupervisor.query.filter_by(
                    empresa_id=empresa_id_alvo, ordem=ordem_alvo
                ).first()
            else:
                faixa_espelho = FaixaComissaoVendedor.query.filter_by(
                    empresa_id=empresa_id_alvo, ordem=ordem_alvo
                ).first()

            if faixa_espelho:
                db.session.delete(faixa_espelho)

        db.session.commit()

        tipo_nome = "vendedor" if tipo == "vendedor" else "supervisor"
        msg_sinc = " (faixa espelhada também removida)" if faixa_espelho else ""
        flash(
            f"Faixa de comissão de {tipo_nome} excluída com sucesso!{msg_sinc}",
            "success",
        )

        # Recalcular comissões das metas do mês atual (empresa do usuário)
        if current_user.cargo != "super_admin":
            _recalcular_comissoes_mes_atual_empresa(current_user.empresa_id)
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir faixa: {str(e)}", "danger")

    return redirect(url_for("configuracoes_comissoes"))

@app.route("/api/comissoes/faixas")
@login_required
def api_faixas_comissoes():
    """API JSON para obter faixas de comissão (DEPRECADO - usar rotas específicas por tipo)"""
    tipo = request.args.get("tipo", "vendedor")

    if tipo == "supervisor":
        modelo = FaixaComissaoSupervisor
    else:
        modelo = FaixaComissaoVendedor

    if current_user.cargo == "super_admin":
        faixas = (
            modelo.query.filter(modelo.empresa_id.is_(None))
            .order_by(modelo.ordem)
            .all()
        )
    else:
        faixas = (
            modelo.query.filter_by(empresa_id=current_user.empresa_id)
            .order_by(modelo.ordem)
            .all()
        )

        if not faixas:
            faixas = (
                modelo.query.filter(modelo.empresa_id.is_(None))
                .order_by(modelo.ordem)
                .all()
            )

    return jsonify([faixa.to_dict() for faixa in faixas])

# ===== Sincronização automática: helpers e rota =====

def _sync_chave(empresa_id: int | None) -> str:
    return (
        "sincronizacao_faixas_comissao_global"
        if empresa_id is None
        else f"sincronizacao_faixas_comissao_empresa_{empresa_id}"
    )

def _get_sync_flag(empresa_id: int | None) -> bool:
    try:
        chave = _sync_chave(empresa_id)
        cfg = Configuracao.query.filter_by(chave=chave).first()
        if not cfg or not cfg.valor:
            return True  # padrão: ativo
        return str(cfg.valor).strip().lower() in ("1", "true", "on", "yes")
    except Exception:
        return True

def _set_sync_flag(empresa_id: int | None, enabled: bool) -> None:
    chave = _sync_chave(empresa_id)
    cfg = Configuracao.query.filter_by(chave=chave).first()
    if not cfg:
        cfg = Configuracao(chave=chave)
    cfg.valor = "true" if enabled else "false"
    db.session.add(cfg)
    db.session.commit()

@app.route("/configuracoes/comissoes/sincronizacao", methods=["POST"])
@login_required
def configurar_sincronizacao_comissoes():
    if current_user.cargo not in ["admin", "super_admin"]:
        flash("Acesso negado.", "danger")
        return redirect(url_for("dashboard"))

    empresa_contexto = None if current_user.cargo == "super_admin" else current_user.empresa_id
    enabled = request.form.get("sincronizacao") == "on"
    try:
        _set_sync_flag(empresa_contexto, enabled)
        if enabled:
            flash("Sincronização automática ativada.", "success")
        else:
            flash("Sincronização automática desativada.", "warning")
    except Exception as e:
        flash(f"Erro ao salvar sincronização: {e}", "danger")
    return redirect(url_for("configuracoes_comissoes"))

# ===== COMANDOS CLI =====

@app.cli.command()
def init_db():
    """Inicializa o banco de dados"""
    db.create_all()
    print("[OK] Banco de dados criado com sucesso!")

@app.cli.command()
def create_admin():
    """Cria um usuário administrador"""
    admin = Usuario(
        nome="Administrador", email="admin@metas.com", cargo="admin"
    )
    admin.set_senha("admin123")

    db.session.add(admin)
    db.session.commit()
    print("[OK] Usuario administrador criado!")
    print("   Email: admin@metas.com")
    print("   Senha: admin123")

# Inicialização automática do banco de dados  # Apenas se não estiver sendo importado por outro script
if __name__ != "__main__" and os.environ.get("SKIP_INIT") != "1":
    with app.app_context():
        try:
            db.create_all()
            print("[OK] Tabelas do banco de dados criadas/verificadas!")

            # Iniciar sistema de backup automático (com try/except para não
            # bloquear o app)
            try:
                # Só iniciar scheduler se não estiver em modo de init
                if not os.environ.get("INIT_DB_ONLY"):
                    iniciar_backup_automatico()
            except Exception as e:
                print(f"[AVISO] Aviso ao iniciar backup automatico: {e}")
        except Exception as e:
            print(f"[AVISO] Aviso na inicializacao do BD: {e}")

# ===== ROTAS DE CONFIGURAÇÃO DE BACKUP =====

@app.route("/super-admin/backups/config")
@super_admin_required
def backup_config_page():
    """Página de configuração de backups automáticos"""
    # Obter próxima execução do backup
    proxima_execucao = None
    if scheduler.get_job("backup_automatico"):
        job = scheduler.get_job("backup_automatico")
        proxima_execucao = job.next_run_time

    return render_template(
        "super_admin/backup_config.html",
        config=backup_config,
        proxima_execucao=proxima_execucao,
        scheduler_ativo=scheduler.running,
    )

@app.route("/super-admin/backups/config/salvar", methods=["POST"])
@super_admin_required
def salvar_config_backup():
    """Salva configurações de backup automático"""
    try:
        backup_config["enabled"] = request.form.get("enabled") == "on"
        backup_config["frequency"] = request.form.get("frequency", "daily")
        backup_config["time"] = request.form.get("time", "02:00")
        backup_config["keep_last"] = int(request.form.get("keep_last", 7))
        backup_config["auto_cleanup"] = (
            request.form.get("auto_cleanup") == "on"
        )

        # Reiniciar scheduler com novas configurações
        if scheduler.get_job("backup_automatico"):
            scheduler.remove_job("backup_automatico")

        if backup_config["enabled"]:
            iniciar_backup_automatico()
            flash(
                "Configurações de backup salvas e agendamento atualizado!",
                "success",
            )
        else:
            flash("⚠️ Backup automático desativado.", "warning")

    except Exception as e:
        flash(f"Erro ao salvar configurações: {str(e)}", "danger")

    return redirect(url_for("backup_config_page"))

@app.route("/super-admin/backups/executar-agora", methods=["POST"])
@super_admin_required
def executar_backup_agora():
    """Executa backup imediatamente"""
    try:
        criar_backup_automatico()
        flash("Backup manual executado com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao executar backup: {str(e)}", "danger")

    return redirect(url_for("super_admin_backups"))


@app.route("/super-admin/limpar-clientes", methods=["POST"])
@super_admin_required
def limpar_todos_clientes():
    """Limpa todos os clientes cadastrados no sistema"""
    try:
        # Contar antes de deletar
        total_compras = CompraCliente.query.count()
        total_clientes = Cliente.query.count()
        
        # Deletar compras primeiro (relacionamento)
        if total_compras > 0:
            CompraCliente.query.delete()
        
        # Deletar clientes
        if total_clientes > 0:
            Cliente.query.delete()
        
        db.session.commit()
        
        flash(f"{total_clientes} clientes e {total_compras} compras removidos com sucesso!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao limpar clientes: {str(e)}", "danger")
    
    return redirect(url_for("super_admin_backups"))

# ====================================================================
# ROTAS DE METAS AVANÇADAS  # ====================================================================

@app.route("/metas/configurar", methods=["GET", "POST"])
@login_required
def configurar_metas():
    """Interface para configurar metas com balanceamento"""
    from calculo_balanceamento import calcular_meta_balanceada

    # Verificar permissão (apenas supervisor, admin, super_admin)
    if current_user.cargo not in ["supervisor", "admin", "super_admin"]:
        flash("Você não tem permissão para configurar metas.", "danger")
        return redirect(url_for("dashboard"))

    # Buscar vendedores disponíveis
    if current_user.cargo == "supervisor":
        vendedores = Vendedor.query.filter_by(
            supervisor_id=current_user.id, ativo=True
        ).all()
    elif current_user.is_super_admin:
        vendedores = Vendedor.query.filter_by(ativo=True).all()
    else:
        vendedores = Vendedor.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        ).all()

    dados_calculo = None

    if request.method == "POST":
        vendedor_id = request.form.get("vendedor_id", type=int)
        tipo_meta = request.form.get("tipo_meta", "valor")
        periodo_historico = request.form.get("periodo_historico", 6, type=int)
        tipo_balanceamento = request.form.get("tipo_balanceamento", "simples")
        mes = request.form.get("mes", type=int)
        ano = request.form.get("ano", type=int)
        ajuste_manual = request.form.get("ajuste_manual", type=float)

        # Calcular meta balanceada
        dados_calculo = calcular_meta_balanceada(
            vendedor_id, periodo_historico, tipo_balanceamento
        )

        # Se foi solicitado salvar
        if "salvar" in request.form:
            try:
                # Verificar se já existe meta para este vendedor/mês/ano
                meta_existente = Meta.query.filter_by(
                    vendedor_id=vendedor_id, mes=mes, ano=ano
                ).first()

                if meta_existente:
                    flash(
                        "Já existe uma meta para este vendedor neste mês/ano. Exclua a meta existente primeiro.",
                        "warning",
                    )
                else:
                    # Criar nova meta
                    nova_meta = Meta(
                        vendedor_id=vendedor_id,
                        mes=mes,
                        ano=ano,
                        tipo_meta=tipo_meta,
                        periodo_historico=periodo_historico,
                        meta_balanceada=True,
                        data_base_calculo=datetime.now(),
                    )

                    if tipo_meta == "valor":
                        valor_final = (
                            ajuste_manual
                            if ajuste_manual
                            else dados_calculo["meta_valor"]
                        )
                        nova_meta.valor_meta = valor_final
                        nova_meta.media_mensal_historico = dados_calculo[
                            "media_mensal_valor"
                        ]
                    else:  # volume
                        volume_final = (
                            int(ajuste_manual)
                            if ajuste_manual
                            else dados_calculo["meta_volume"]
                        )
                        nova_meta.volume_meta = volume_final
                        nova_meta.media_mensal_historico = dados_calculo[
                            "media_mensal_volume"
                        ]

                    nova_meta.tendencia_calculada = dados_calculo["tendencia"]

                    db.session.add(nova_meta)
                    db.session.commit()

                    flash(
                        f"Meta de {tipo_meta} criada com sucesso para {mes}/{ano}!",
                        "success",
                    )
                    return redirect(url_for("configurar_metas"))

            except Exception as e:
                db.session.rollback()
                flash(f"Erro ao criar meta: {str(e)}", "danger")

    # Anos disponíveis (atual e próximos 2)
    ano_atual = datetime.now().year
    anos = [ano_atual, ano_atual + 1, ano_atual + 2]

    return render_template(
        "metas/configurar.html",
        vendedores=vendedores,
        dados_calculo=dados_calculo,
        anos=anos,
    )

@app.route("/relatorios/metas-avancado")
@login_required
def relatorio_metas_avancado():
    """Relatório avançado de metas com gráficos"""
    from calculo_balanceamento import obter_ranking_meses

    # Filtros
    visao = request.args.get("visao", "vendedor")  # 'vendedor' | 'supervisor'
    vendedor_id = request.args.get("vendedor_id", type=int)
    supervisor_id = request.args.get("supervisor_id", type=int)
    tipo_meta = request.args.get("tipo_meta", "")
    ano = request.args.get("ano", datetime.now().year, type=int)
    mes = request.args.get("mes", type=int)

    # Query base
    query = Meta.query

    # Aplicar filtros de permissão
    if current_user.cargo == "vendedor" and current_user.vendedor_id:
        query = query.join(Vendedor).filter(
            Meta.vendedor_id == current_user.vendedor_id
        )
    elif current_user.cargo == "supervisor":
        vendedores_ids = [
            v.id
            for v in Vendedor.query.filter_by(
                supervisor_id=current_user.id, ativo=True
            ).all()
        ]
        query = query.filter(Meta.vendedor_id.in_(vendedores_ids))
    elif not current_user.is_super_admin:
        query = query.join(Vendedor).filter(
            Vendedor.empresa_id == current_user.empresa_id
        )

    # Aplicar filtros adicionais
    if vendedor_id:
        query = query.filter(Meta.vendedor_id == vendedor_id)
    # Filtro por supervisor (quando fornecido ou quando visão é supervisor)
    if supervisor_id:
        query = query.join(Vendedor).filter(Vendedor.supervisor_id == supervisor_id)
    if tipo_meta:
        query = query.filter(Meta.tipo_meta == tipo_meta)
    if ano:
        query = query.filter(Meta.ano == ano)
    if mes:
        query = query.filter(Meta.mes == mes)

    # Buscar metas
    metas = query.order_by(Meta.ano.desc(), Meta.mes.desc()).all()

    # Recalcular comissões se necessário
    for meta in metas:
        if meta.comissao_total is None or meta.comissao_total == 0:
            meta.calcular_comissao()

    # Estatísticas gerais (independente da visão)
    total_metas = len(metas)
    metas_atingidas = sum(1 for m in metas if m.percentual_alcance >= 100)
    taxa_sucesso = (
        (metas_atingidas / total_metas * 100) if total_metas > 0 else 0
    )

    # Comissão total
    comissao_total = sum(m.comissao_total or 0 for m in metas)

    # Buscar vendedores para filtro
    if current_user.cargo == "vendedor":
        vendedores = [Vendedor.query.get(current_user.vendedor_id)]
    elif current_user.cargo == "supervisor":
        vendedores = Vendedor.query.filter_by(
            supervisor_id=current_user.id, ativo=True
        ).all()
    elif current_user.is_super_admin:
        vendedores = Vendedor.query.filter_by(ativo=True).all()
    else:
        vendedores = Vendedor.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        ).all()

    # Buscar supervisores para filtro
    from models import Usuario
    if current_user.cargo == "supervisor":
        supervisores = [Usuario.query.get(current_user.id)]
    elif current_user.is_super_admin:
        supervisores = Usuario.query.filter_by(cargo="supervisor", ativo=True).all()
    else:
        supervisores = Usuario.query.filter_by(
            empresa_id=current_user.empresa_id, cargo="supervisor", ativo=True
        ).all()

    # Ranking de meses (somente na visão por vendedor e com vendedor selecionado)
    ranking = None
    if visao == "vendedor" and vendedor_id:
        ranking = obter_ranking_meses(vendedor_id, ano)

    # Agregação por Supervisor quando solicitado
    supervisores_resumo = []
    if visao == "supervisor":
        # Agrupar metas por supervisor
        grupos = {}
        for m in metas:
            sup_id = m.vendedor.supervisor_id if m.vendedor else None
            sup_nome = m.vendedor.supervisor_nome if m.vendedor else "N/A"
            if sup_id is None:
                # Ignora metas sem supervisor vinculado
                continue
            if sup_id not in grupos:
                grupos[sup_id] = {
                    "supervisor_id": sup_id,
                    "supervisor_nome": sup_nome,
                    "tipo_meta": tipo_meta or m.tipo_meta,
                    "mes": mes,
                    "ano": ano,
                    "meta_total": 0.0,
                    "realizado_total": 0.0,
                    "volume_meta_total": 0,
                    "volume_realizado_total": 0,
                    "comissao_total_vendedores": 0.0,
                }
            if m.tipo_meta == "valor":
                grupos[sup_id]["meta_total"] += float(m.valor_meta or 0.0)
                grupos[sup_id]["realizado_total"] += float(m.receita_alcancada or 0.0)
            else:
                grupos[sup_id]["volume_meta_total"] += int(m.volume_meta or 0)
                grupos[sup_id]["volume_realizado_total"] += int(m.volume_alcancado or 0)
            grupos[sup_id]["comissao_total_vendedores"] += float(m.comissao_total or 0.0)

        # Calcular percentual e comissão do supervisor pelas faixas
        empresa_id_ctx = None if current_user.is_super_admin else current_user.empresa_id
        for g in grupos.values():
            if (g["tipo_meta"] or "valor") == "valor":
                alcance = (
                    (g["realizado_total"] / g["meta_total"] * 100)
                    if g["meta_total"] > 0
                    else 0
                )
                taxa_sup = _obter_taxa_por_alcance("supervisor", empresa_id_ctx, alcance)
                comissao_sup = float(g["realizado_total"]) * float(taxa_sup)
            else:
                alcance = (
                    (g["volume_realizado_total"] / g["volume_meta_total"] * 100)
                    if g["volume_meta_total"] > 0
                    else 0
                )
                # Para metas de volume, comissionar pela soma das comissões dos vendedores
                taxa_sup = None
                comissao_sup = g["comissao_total_vendedores"]

            supervisores_resumo.append({
                "id": g["supervisor_id"],
                "nome": g["supervisor_nome"],
                "tipo_meta": g["tipo_meta"],
                "periodo": f"{mes:02d}/{ano}" if (mes and ano) else "—",
                "meta_total": g["meta_total"] if (g["tipo_meta"] == "valor") else g["volume_meta_total"],
                "realizado_total": g["realizado_total"] if (g["tipo_meta"] == "valor") else g["volume_realizado_total"],
                "percentual_alcance": alcance,
                "taxa_supervisor": taxa_sup,
                "comissao_supervisor": comissao_sup,
            })

        # Ordenar por percentual
        supervisores_resumo.sort(key=lambda x: x["percentual_alcance"], reverse=True)

    # Anos disponíveis
    anos_disponiveis = (
        db.session.query(Meta.ano).distinct().order_by(Meta.ano.desc()).all()
    )
    anos_disponiveis = (
        [a[0] for a in anos_disponiveis]
        if anos_disponiveis
        else [datetime.now().year]
    )

    return render_template(
        "relatorios/metas_avancado.html",
        metas=metas,
        vendedores=vendedores,
        supervisores=supervisores,
        estatisticas={
            "total_metas": total_metas,
            "metas_atingidas": metas_atingidas,
            "taxa_sucesso": taxa_sucesso,
            "total_comissoes": comissao_total,
        },
        ranking_meses=ranking,
        anos=anos_disponiveis,
        filtros={
            "visao": visao,
            "vendedor_id": vendedor_id,
            "supervisor_id": supervisor_id,
            "tipo_meta": tipo_meta,
            "ano": ano,
            "mes": mes,
        },
        supervisores_resumo=supervisores_resumo,
    )

@app.route("/relatorios/metas-avancado/export")
@login_required
def exportar_pdf_metas_avancado():
    """Exporta o Relatório de Metas Avançado em PDF nas visões Vendedor ou Supervisor."""
    visao = request.args.get("visao", "vendedor")
    vendedor_id = request.args.get("vendedor_id", type=int)
    supervisor_id = request.args.get("supervisor_id", type=int)
    tipo_meta = request.args.get("tipo_meta", "")
    ano = request.args.get("ano", datetime.now().year, type=int)
    mes = request.args.get("mes", datetime.now().month, type=int)

    # Construir query semelhante à página
    query = Meta.query

    # Escopo por cargo
    if current_user.cargo == "vendedor" and current_user.vendedor_id:
        query = query.join(Vendedor).filter(Meta.vendedor_id == current_user.vendedor_id)
    elif current_user.cargo == "supervisor":
        vendedores_ids = [v.id for v in Vendedor.query.filter_by(supervisor_id=current_user.id, ativo=True).all()]
        query = query.filter(Meta.vendedor_id.in_(vendedores_ids))
    elif not current_user.is_super_admin:
        query = query.join(Vendedor).filter(Vendedor.empresa_id == current_user.empresa_id)

    # Filtros adicionais
    if vendedor_id:
        query = query.filter(Meta.vendedor_id == vendedor_id)
    if supervisor_id:
        query = query.join(Vendedor).filter(Vendedor.supervisor_id == supervisor_id)
    if tipo_meta:
        query = query.filter(Meta.tipo_meta == tipo_meta)
    if ano:
        query = query.filter(Meta.ano == ano)
    if mes:
        query = query.filter(Meta.mes == mes)

    metas = query.order_by(Meta.ano.desc(), Meta.mes.desc()).all()

    # Recalcular comissões quando necessário
    for meta in metas:
        if meta.comissao_total is None or meta.comissao_total == 0:
            try:
                meta.calcular_comissao()
            except Exception:
                pass

    try:
        if visao == "supervisor":
            # Agregação por supervisor (mesmo cálculo da página)
            grupos = {}
            for m in metas:
                sup_id = m.vendedor.supervisor_id if m.vendedor else None
                sup_nome = m.vendedor.supervisor_nome if m.vendedor else "N/A"
                if sup_id is None:
                    continue
                if sup_id not in grupos:
                    grupos[sup_id] = {
                        "supervisor_id": sup_id,
                        "supervisor_nome": sup_nome,
                        "tipo_meta": tipo_meta or m.tipo_meta,
                        "mes": mes,
                        "ano": ano,
                        "meta_total": 0.0,
                        "realizado_total": 0.0,
                        "volume_meta_total": 0,
                        "volume_realizado_total": 0,
                        "comissao_total_vendedores": 0.0,
                    }
                if m.tipo_meta == "valor":
                    grupos[sup_id]["meta_total"] += float(m.valor_meta or 0.0)
                    grupos[sup_id]["realizado_total"] += float(m.receita_alcancada or 0.0)
                else:
                    grupos[sup_id]["volume_meta_total"] += int(m.volume_meta or 0)
                    grupos[sup_id]["volume_realizado_total"] += int(m.volume_alcancado or 0)
                grupos[sup_id]["comissao_total_vendedores"] += float(m.comissao_total or 0.0)

            empresa_id_ctx = None if current_user.is_super_admin else current_user.empresa_id

            supervisores_resumo = []
            for g in grupos.values():
                if (g["tipo_meta"] or "valor") == "valor":
                    alcance = (g["realizado_total"] / g["meta_total"] * 100) if g["meta_total"] > 0 else 0
                    taxa_sup = _obter_taxa_por_alcance("supervisor", empresa_id_ctx, alcance)
                    comissao_sup = float(g["realizado_total"]) * float(taxa_sup)
                else:
                    alcance = (g["volume_realizado_total"] / g["volume_meta_total"] * 100) if g["volume_meta_total"] > 0 else 0
                    taxa_sup = None
                    comissao_sup = g["comissao_total_vendedores"]

                supervisores_resumo.append({
                    "id": g["supervisor_id"],
                    "nome": g["supervisor_nome"],
                    "tipo_meta": g["tipo_meta"],
                    "periodo": f"{mes:02d}/{ano}" if (mes and ano) else "—",
                    "meta_total": g["meta_total"] if (g["tipo_meta"] == "valor") else g["volume_meta_total"],
                    "realizado_total": g["realizado_total"] if (g["tipo_meta"] == "valor") else g["volume_realizado_total"],
                    "percentual_alcance": alcance,
                    "taxa_supervisor": taxa_sup,
                    "comissao_supervisor": comissao_sup,
                })

            supervisores_resumo.sort(key=lambda x: x["percentual_alcance"], reverse=True)

            pdf_buffer = gerar_pdf_metas_supervisor(supervisores_resumo, mes, ano)
            meses = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
            nome_periodo = f"_{meses[mes-1]}_{ano}" if (mes and ano and 1 <= mes <= 12) else ""
            filename = f"Relatorio_Metas_Supervisores{nome_periodo}.pdf"
        else:
            # Visão Vendedor: usar gerador existente
            pdf_buffer = gerar_pdf_metas(metas, mes, ano)
            meses = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
            nome_periodo = f"_{meses[mes-1]}_{ano}" if (mes and ano and 1 <= mes <= 12) else ""
            filename = f"Relatorio_Metas_Vendedores{nome_periodo}.pdf"

        return send_file(pdf_buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)
    except Exception as e:
        app.logger.error(f"Erro ao exportar PDF de metas avançado: {e}", exc_info=True)
        flash("Não foi possível gerar o PDF. Verifique os logs e tente novamente.", "danger")
        return redirect(url_for("relatorio_metas_avancado", **request.args))

@app.route("/api/metas/dados-grafico/<int:vendedor_id>")
@login_required
def api_dados_grafico_metas(vendedor_id):
    """API para retornar dados de gráficos de metas"""
    from calculo_balanceamento import (
        obter_dados_grafico_evolucao,
        obter_ranking_meses,
    )

    periodo_meses = request.args.get("periodo", 12, type=int)
    ano = request.args.get("ano", datetime.now().year, type=int)

    # Verificar permissão
    if (
        current_user.cargo == "vendedor" and
        current_user.vendedor_id != vendedor_id
    ):
        return jsonify({"error": "Sem permissão"}), 403
    elif current_user.cargo == "supervisor":
        vendedor = Vendedor.query.get_or_404(vendedor_id)
        if vendedor.supervisor_id != current_user.id:
            return jsonify({"error": "Sem permissão"}), 403

    # Dados de evolução
    evolucao = obter_dados_grafico_evolucao(vendedor_id, periodo_meses)

    # Ranking de meses
    ranking = obter_ranking_meses(vendedor_id, ano)

    return jsonify({"evolucao": evolucao, "ranking": ranking})

# ====================================================================
# FIM DAS ROTAS DE METAS AVANÇADAS  # ====================================================================

# ====================================================================
# ROTAS DE ORDENS DE SERVIÇO (MANUTENÇÃO)  # ====================================================================

@app.route("/os")
@login_required
def lista_ordens_servico():
    """Lista todas as ordens de serviço com filtros"""
    # Verificar permissões: admin, gerente_manutencao, supervisor_manutencao,
    # administrativo, tecnico e auxiliar (apenas visualização)
    cargos_permitidos = [
        "admin",
        "gerente_manutencao",
        "supervisor_manutencao",
        "administrativo",
        "tecnico",
        "auxiliar",
    ]
    if current_user.cargo not in cargos_permitidos:
        flash(
            "Acesso negado! Você não tem permissão para acessar ordens de serviço.",
            "danger",
        )
        return redirect(url_for("dashboard"))

    # Obter filtros
    status = request.args.get("status", "")
    prioridade = request.args.get("prioridade", "")
    tecnico_id = request.args.get("tecnico_id", "")
    busca = request.args.get("busca", "")
    page = request.args.get("page", 1, type=int)

    # Quantidade de registros por página (equilíbrio entre performance e usabilidade)
    per_page = 15

    # Query base
    query = OrdemServico.query.filter_by(empresa_id=current_user.empresa_id)

    # Aplicar filtros de permissão por cargo
    if current_user.cargo == "tecnico":
        # Técnico só vê suas próprias OS
        tecnico = Tecnico.query.filter_by(
            usuario_id=current_user.id, empresa_id=current_user.empresa_id
        ).first()
        if tecnico:
            query = query.filter_by(tecnico_id=tecnico.id)
        else:
            flash(
                "⚠️ Você não está vinculado a um técnico no sistema.", "warning"
            )
            query = query.filter_by(id=0)  # Sem resultados

    # Aplicar filtros da interface
    if status:
        query = query.filter_by(status=status)
    if prioridade:
        query = query.filter_by(prioridade=prioridade)
    if tecnico_id:
        query = query.filter_by(tecnico_id=tecnico_id)
    if busca:
        query = query.filter(
            db.or_(
                OrdemServico.numero_os.ilike(f"%{busca}%"),
                OrdemServico.titulo.ilike(f"%{busca}%"),
                OrdemServico.descricao_problema.ilike(f"%{busca}%"),
            )
        )

    # Estatísticas por status (baseadas no conjunto filtrado, antes da paginação)
    stats_query = query
    stats = {
        "aguardando_aprovacao": stats_query.filter_by(status="aguardando_aprovacao").count(),
        "em_andamento": stats_query.filter_by(status="em_andamento").count(),
        "concluida": stats_query.filter_by(status="concluida").count(),
        "total": stats_query.count(),
    }

    # Paginação: ordenar por data de abertura (mais recentes primeiro)
    pagination = query.order_by(OrdemServico.data_abertura.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    ordens = pagination.items

    # Buscar técnicos para o filtro
    tecnicos = (
        Tecnico.query.filter_by(empresa_id=current_user.empresa_id, ativo=True)
        .order_by(Tecnico.nome)
        .all()
    )

    return render_template(
        "os/lista.html",
        ordens=ordens,
        pagination=pagination,
        stats=stats,
        tecnicos=tecnicos,
        filtros={
            "status": status,
            "prioridade": prioridade,
            "tecnico_id": tecnico_id,
            "busca": busca,
        },
    )

@app.route("/os/nova", methods=["GET", "POST"])
@login_required
def nova_ordem_servico():
    """Criar nova ordem de serviço - Administrativo"""
    # Apenas administrativo e admin podem criar OS
    if current_user.cargo not in ["admin", "administrativo"]:
        flash(
            "Acesso negado! Apenas Administrativo pode criar ordens de serviço.",
            "danger",
        )
        return redirect(url_for("lista_ordens_servico"))

    form = OrdemServicoForm()

    # Preencher choices de clientes
    clientes = (
        Cliente.query.filter_by(empresa_id=current_user.empresa_id)
        .order_by(Cliente.nome)
        .all()
    )
    form.cliente_id.choices = [(0, "Selecione o cliente")] + [
        (c.id, f"{c.nome} - {c.cpf or c.cnpj}") for c in clientes
    ]

    # Preencher choices de técnicos (necessário para validação do SelectField)
    tecnicos = Tecnico.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
    form.tecnico_id.choices = [(0, "Selecione o técnico (Opcional)")] + [
        (t.id, t.nome) for t in tecnicos
    ]

    if form.validate_on_submit():
        # Usar módulo de serviço para reduzir acoplamento
        try:
            from modules.os_service import criar_os
            ok, os_obj, erro = criar_os(form, current_user)
            if ok and os_obj:
                return redirect(url_for("visualizar_ordem_servico", id=os_obj.id))
            else:
                flash(f"Erro ao criar ordem de serviço: {erro}", "danger")
        except Exception as e:
            flash(f"Erro ao criar ordem de serviço: {str(e)}", "danger")

    return render_template("os/nova.html", form=form)

@app.route("/os/<int:id>")
@login_required
def visualizar_ordem_servico(id):
    """Visualizar detalhes da ordem de serviço"""
    os = OrdemServico.query.get_or_404(id)

    # Verificar empresa
    if os.empresa_id != current_user.empresa_id:
        flash("Ordem de serviço não encontrada.", "danger")
        return redirect(url_for("lista_ordens_servico"))

    # Verificar permissão (técnico só vê suas próprias OS; auxiliar vê
    # apenas OS da própria empresa, validação extra pode ser aplicada depois)
    if current_user.cargo == "tecnico":
        tecnico = Tecnico.query.filter_by(
            usuario_id=current_user.id, empresa_id=current_user.empresa_id
        ).first()
        if not tecnico or os.tecnico_id != tecnico.id:
            flash(
                "Você não tem permissão para visualizar esta ordem de serviço.",
                "danger",
            )
            return redirect(url_for("lista_ordens_servico"))

    # Buscar movimentos de estoque relacionados
    movimentos = EstoqueMovimento.query.filter_by(ordem_servico_id=os.id).all()

    return render_template("os/visualizar.html", os=os, movimentos=movimentos)

@app.route("/os/<int:id>/aprovar", methods=["GET", "POST"])
@login_required
def aprovar_ordem_servico(id):
    """Aprovar ou reprovar ordem de serviço - Supervisor de Manutenção"""
    # Apenas gerente/supervisor de manutenção e admin podem aprovar
    if current_user.cargo not in ["admin", "gerente_manutencao", "supervisor_manutencao"]:
        flash(
            "Acesso negado! Apenas Supervisor de Manutenção pode aprovar ordens.",
            "danger",
        )
        return redirect(url_for("lista_ordens_servico"))

    os = OrdemServico.query.get_or_404(id)

    # Verificar empresa
    if os.empresa_id != current_user.empresa_id:
        flash("Ordem de serviço não encontrada.", "danger")
        return redirect(url_for("lista_ordens_servico"))

    # Verificar se pode aprovar
    if not os.pode_aprovar(current_user):
        flash(
            "Esta ordem de serviço não pode ser aprovada no momento.",
            "danger",
        )
        return redirect(url_for("visualizar_ordem_servico", id=os.id))

    form = OrdemServicoAvaliarForm()

    # Preencher choices de técnicos
    tecnicos = (
        Tecnico.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True, disponivel=True
        )
        .order_by(Tecnico.nome)
        .all()
    )
    form.tecnico_id.choices = [(0, "Selecione o técnico")] + [
        (t.id, f'{t.nome} - {t.especialidades or "Geral"}') for t in tecnicos
    ]

    if form.validate_on_submit():
        try:
            from modules.os_service import aprovar_ou_reprovar_os
            ok, erro = aprovar_ou_reprovar_os(form, os, current_user)
            if ok:
                return redirect(url_for("visualizar_ordem_servico", id=os.id))
            else:
                flash(erro or "Erro ao processar ordem de serviço.", "danger")
        except Exception as e:
            flash(f"Erro ao processar ordem de serviço: {str(e)}", "danger")

    return render_template("os/aprovar.html", form=form, os=os)

@app.route("/os/<int:id>/atualizar", methods=["GET", "POST"])
@login_required
def atualizar_ordem_servico(id):
    """Atualizar andamento da ordem de serviço - Técnico"""
    os = OrdemServico.query.get_or_404(id)

    # Verificar empresa
    if os.empresa_id != current_user.empresa_id:
        flash("Ordem de serviço não encontrada.", "danger")
        return redirect(url_for("lista_ordens_servico"))

    # Verificar se pode editar
    if not os.pode_editar(current_user):
        flash(
            "Você não tem permissão para atualizar esta ordem de serviço.",
            "danger",
        )
        return redirect(url_for("visualizar_ordem_servico", id=os.id))

    form = OrdemServicoAndamentoForm()

    if form.validate_on_submit():
        try:
            from modules.os_service import atualizar_os
            ok, erro = atualizar_os(form, os, current_user)
            if ok:
                return redirect(url_for("visualizar_ordem_servico", id=os.id))
            else:
                flash(erro or "Erro ao atualizar ordem de serviço.", "danger")
        except Exception as e:
            flash(f"Erro ao atualizar ordem de serviço: {str(e)}", "danger")
    else:
        # Preencher form com dados atuais
        form.status.data = os.status
        form.descricao_solucao.data = os.descricao_solucao
        form.valor_mao_obra.data = os.valor_mao_obra
        form.valor_pecas.data = os.valor_pecas
        form.data_previsao.data = os.data_previsao
        form.feedback_tecnico.data = os.feedback_tecnico

    return render_template("os/atualizar.html", form=form, os=os)

@app.route("/os/<int:id>/avaliar", methods=["GET", "POST"])
@login_required
def avaliar_ordem_servico(id):
    """Cliente avalia a ordem de serviço concluída"""
    os = OrdemServico.query.get_or_404(id)

    # Verificar empresa
    if os.empresa_id != current_user.empresa_id:
        flash("Ordem de serviço não encontrada.", "danger")
        return redirect(url_for("lista_ordens_servico"))

    # Verificar se está concluída
    if os.status != "concluida":
        flash(
            "Apenas ordens de serviço concluídas podem ser avaliadas.",
            "danger",
        )
        return redirect(url_for("visualizar_ordem_servico", id=os.id))

    # Verificar se já foi avaliada
    if os.avaliacao_cliente:
        flash("⚠️ Esta ordem de serviço já foi avaliada.", "info")
        return redirect(url_for("visualizar_ordem_servico", id=os.id))

    form = OrdemServicoAvaliacaoForm()

    if form.validate_on_submit():
        try:
            from modules.os_service import avaliar_os
            ok, erro = avaliar_os(form, os, current_user)
            if ok:
                return redirect(url_for("visualizar_ordem_servico", id=os.id))
            else:
                flash(erro or "Erro ao registrar avaliação.", "danger")
        except Exception as e:
            flash(f"Erro ao registrar avaliação: {str(e)}", "danger")

    return render_template("os/avaliar.html", form=form, os=os)

# ====================================================================
# FIM DAS ROTAS DE ORDENS DE SERVIÇO  # ====================================================================

"""
=================================================================
ROTAS DE ESTOQUE E PRODUTOS
=================================================================
"""

@app.route("/estoque")
@login_required
def lista_estoque():
    """Dashboard de estoque com resumo e estatísticas"""
    if not current_user.pode_acessar_estoque and current_user.cargo != "admin":
        flash("Você não tem permissão para acessar o estoque.", "danger")
        return redirect(url_for("dashboard"))

    total_produtos = Produto.query.filter_by(
        empresa_id=current_user.empresa_id, ativo=True
    ).count()

    produtos_zerados = (
        Produto.query.filter_by(empresa_id=current_user.empresa_id, ativo=True)
        .filter(Produto.estoque_atual <= 0)
        .count()
    )

    produtos_baixos = (
        Produto.query.filter_by(empresa_id=current_user.empresa_id, ativo=True)
        .filter(
            Produto.estoque_atual > 0,
            Produto.estoque_atual <= Produto.estoque_minimo,
        )
        .count()
    )

    valor_total = 0
    if current_user.pode_ver_custos or current_user.cargo == "admin":
        produtos = Produto.query.filter_by(
            empresa_id=current_user.empresa_id, ativo=True
        ).all()
        valor_total = sum(p.estoque_atual * p.custo_medio for p in produtos)

    ultimas_movimentacoes = (
        EstoqueMovimento.query.filter_by(empresa_id=current_user.empresa_id)
        .order_by(EstoqueMovimento.data_movimento.desc())
        .limit(10)
        .all()
    )

    produtos_alerta = (
        Produto.query.filter_by(empresa_id=current_user.empresa_id, ativo=True)
        .filter(
            Produto.estoque_atual > 0,
            Produto.estoque_atual <= Produto.estoque_minimo,
        )
        .order_by(Produto.estoque_atual)
        .limit(10)
        .all()
    )

    return render_template(
        "estoque/dashboard.html",
        total_produtos=total_produtos,
        produtos_zerados=produtos_zerados,
        produtos_baixos=produtos_baixos,
        valor_total=valor_total,
        ultimas_movimentacoes=ultimas_movimentacoes,
        produtos_alerta=produtos_alerta,
    )

@app.route("/estoque/produtos")
@login_required
def lista_produtos():
    """Lista todos os produtos do estoque"""
    if not current_user.pode_acessar_estoque and current_user.cargo != "admin":
        flash("Você não tem permissão para acessar o estoque.", "danger")
        return redirect(url_for("dashboard"))

    busca = request.args.get("busca", "")
    categoria = request.args.get("categoria", "")
    status_estoque = request.args.get("status", "")
    page = request.args.get("page", 1, type=int)
    per_page = 25

    query = Produto.query.filter_by(empresa_id=current_user.empresa_id)

    if busca:
        query = query.filter(
            db.or_(
                Produto.codigo.ilike(f"%{busca}%"),
                Produto.nome.ilike(f"%{busca}%"),
                Produto.codigo_barra.ilike(f"%{busca}%"),
                Produto.referencia.ilike(f"%{busca}%"),
            )
        )

    if categoria:
        query = query.filter_by(categoria=categoria)

    if status_estoque == "zerado":
        query = query.filter(Produto.estoque_atual <= 0)
    elif status_estoque == "baixo":
        query = query.filter(
            Produto.estoque_atual > 0,
            Produto.estoque_atual <= Produto.estoque_minimo,
        )
    elif status_estoque == "normal":
        query = query.filter(Produto.estoque_atual > Produto.estoque_minimo)

    pagination = query.order_by(Produto.nome).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )

    produtos = pagination.items

    return render_template(
        "estoque/produtos.html",
        produtos=produtos,
        filtros={
            "busca": busca,
            "categoria": categoria,
            "status": status_estoque,
        },
        pagination=pagination,
    )

@app.route("/estoque/produto/novo", methods=["GET", "POST"])
@login_required
def novo_produto():
    """Cadastrar novo produto"""
    if (
        not current_user.pode_gerenciar_produtos and
        current_user.cargo != "admin"
    ):
        flash("Você não tem permissão para cadastrar produtos.", "danger")
        return redirect(url_for("lista_produtos"))

    form = ProdutoForm()

    if form.validate_on_submit():
        try:
            existe = Produto.query.filter_by(
                codigo=form.codigo.data, empresa_id=current_user.empresa_id
            ).first()

            if existe:
                flash("Já existe um produto com este código!", "danger")
                return render_template(
                    "estoque/produto_form.html",
                    form=form,
                    titulo="Novo Produto",
                )

            produto = Produto(
                codigo=form.codigo.data,
                codigo_barra=form.codigo_barra.data,
                referencia=form.referencia.data,
                ncm=form.ncm.data,
                nome=form.nome.data,
                descricao=form.descricao.data,
                categoria=form.categoria.data,
                fornecedor=form.fornecedor.data,
                unidade=form.unidade.data,
                estoque_minimo=form.estoque_minimo.data or 0,
                estoque_atual=0,
                custo_medio=form.custo_medio.data or 0,
                preco_venda=form.preco_venda.data or 0,
                preco_servico=form.preco_servico.data or 0,
                localizacao=form.localizacao.data,
                ativo=form.ativo.data,
                empresa_id=current_user.empresa_id,
            )

            db.session.add(produto)
            db.session.commit()

            flash(
                f'Produto "{produto.nome}" cadastrado com sucesso!',
                "success",
            )
            return redirect(url_for("visualizar_produto", id=produto.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar produto: {str(e)}", "danger")

    return render_template(
        "estoque/produto_form.html", form=form, titulo="Novo Produto"
    )

@app.route("/estoque/produto/<int:id>")
@login_required
def visualizar_produto(id):
    """Visualizar detalhes do produto"""
    if not current_user.pode_acessar_estoque and current_user.cargo != "admin":
        flash("Você não tem permissão para acessar o estoque.", "danger")
        return redirect(url_for("dashboard"))

    produto = Produto.query.get_or_404(id)

    if produto.empresa_id != current_user.empresa_id:
        flash("Produto não encontrado.", "danger")
        return redirect(url_for("lista_produtos"))

    movimentacoes = (
        EstoqueMovimento.query.filter_by(produto_id=produto.id)
        .order_by(EstoqueMovimento.data_movimento.desc())
        .limit(20)
        .all()
    )

    total_entradas = (
        db.session.query(db.func.sum(EstoqueMovimento.quantidade))
        .filter_by(produto_id=produto.id, tipo="entrada")
        .scalar() or
        0
    )

    total_saidas = (
        db.session.query(db.func.sum(EstoqueMovimento.quantidade))
        .filter_by(produto_id=produto.id, tipo="saida")
        .scalar() or
        0
    )

    return render_template(
        "estoque/produto_visualizar.html",
        produto=produto,
        movimentacoes=movimentacoes,
        total_entradas=total_entradas,
        total_saidas=total_saidas,
    )

@app.route("/estoque/produto/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_produto(id):
    """Editar produto existente"""
    if (
        not current_user.pode_gerenciar_produtos and
        current_user.cargo != "admin"
    ):
        flash("Você não tem permissão para editar produtos.", "danger")
        return redirect(url_for("visualizar_produto", id=id))

    produto = Produto.query.get_or_404(id)

    if produto.empresa_id != current_user.empresa_id:
        flash("Produto não encontrado.", "danger")
        return redirect(url_for("lista_produtos"))

    form = ProdutoForm(obj=produto)

    if form.validate_on_submit():
        try:
            if form.codigo.data != produto.codigo:
                existe = Produto.query.filter_by(
                    codigo=form.codigo.data, empresa_id=current_user.empresa_id
                ).first()

                if existe:
                    flash("Já existe um produto com este código!", "danger")
                    return render_template(
                        "estoque/produto_form.html",
                        form=form,
                        titulo="Editar Produto",
                        produto=produto,
                    )

            produto.codigo = form.codigo.data
            produto.codigo_barra = form.codigo_barra.data
            produto.referencia = form.referencia.data
            produto.ncm = form.ncm.data
            produto.nome = form.nome.data
            produto.descricao = form.descricao.data
            produto.categoria = form.categoria.data
            produto.fornecedor = form.fornecedor.data
            produto.unidade = form.unidade.data
            produto.estoque_minimo = form.estoque_minimo.data or 0
            produto.custo_medio = form.custo_medio.data or 0
            produto.preco_venda = form.preco_venda.data or 0
            produto.preco_servico = form.preco_servico.data or 0
            produto.localizacao = form.localizacao.data
            produto.ativo = form.ativo.data

            db.session.commit()

            flash(
                f'Produto "{produto.nome}" atualizado com sucesso!',
                "success",
            )
            return redirect(url_for("visualizar_produto", id=produto.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar produto: {str(e)}", "danger")

    return render_template(
        "estoque/produto_form.html",
        form=form,
        titulo="Editar Produto",
        produto=produto,
    )

@app.route("/estoque/produtos/download-template")
@login_required
def download_template_produtos():
    """Gera e baixa o template Excel para importação de produtos"""
    # Verificar permissão usando helper
    if not pode_importar(current_user, "produtos"):
        cargos_permitidos = ", ".join(get_cargos_permitidos_importacao("produtos"))
        flash(
            f"Apenas {cargos_permitidos} podem baixar o template de produtos.",
            "danger"
        )
        return redirect(url_for("lista_produtos"))

    if not EXCEL_AVAILABLE and not ensure_excel_available():
        flash("Erro: Bibliotecas Excel não instaladas. Contate o administrador.", "danger")
        return redirect(url_for("lista_produtos"))

    try:
        import io
        import openpyxl

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"

        # Imports adicionais do openpyxl
        from openpyxl.styles import Border, Side
        from openpyxl.utils import get_column_letter

        # Definir estilos
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Cabeçalhos conforme imagem
        headers = [
            "Código",
            "Nome",
            "Referência",
            "Código de barra",
            "NCM",
            "Data Hora",
            "Qte entrada",
            "Qte saída",
            "O.S",
            "Administrativo",
            "Técnico",
            "Vendedor",
            "Supervisor",
            "Gerente",
            "Status",
        ]

        # Aplicar cabeçalhos
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Ajustar largura das colunas
        column_widths = [
            15,
            30,
            20,
            20,
            15,
            20,
            12,
            12,
            15,
            15,
            12,
            12,
            12,
            12,
            12,
        ]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Adicionar linha de exemplo
        exemplo = [
            "PROD001",
            "Produto Exemplo",
            "REF-001",
            "7891234567890",
            "12345678",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Ativo",
        ]
        for col, value in enumerate(exemplo, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Adicionar instruções em nova aba
        ws_info = wb.create_sheet("Instruções")
        ws_info.column_dimensions["A"].width = 80

        instrucoes = [
            "INSTRUÇÕES PARA IMPORTAÇÃO DE PRODUTOS",
            "",
            "1. Preencha os dados na aba 'Produtos'",
            "2. Campos obrigatórios: Código, Nome",
            "3. Campos numéricos: Qte entrada, Qte saída (deixar vazio ou 0 se não aplicável)",
            "4. O.S: Preencher manualmente quando necessário",
            "5. Status: Ativo ou Inativo",
            "6. Produtos com código já cadastrado serão ignorados",
            "7. Novos produtos receberão ID único automaticamente",
            "8. Data Hora será preenchida automaticamente no momento da importação",
            "",
            "COLUNAS DE FUNCIONÁRIOS:",
            "- Use 'Sim' ou deixe vazio nas colunas de funcionários",
            "- Estas colunas são opcionais e servem para controle",
        ]

        for row, texto in enumerate(instrucoes, start=1):
            cell = ws_info.cell(row=row, column=1)
            cell.value = texto
            if row == 1:
                cell.font = Font(bold=True, size=14)
            elif texto.startswith("COLUNAS"):
                cell.font = Font(bold=True, size=12)

        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f'template_produtos_{datetime.now().strftime("%Y%m%d")}.xlsx',
        )

    except ImportError:
        flash(
            "Biblioteca openpyxl não instalada. Execute: pip install openpyxl",
            "danger",
        )
        return redirect(url_for("lista_produtos"))
    except Exception as e:
        flash(f"Erro ao gerar template: {str(e)}", "danger")
        return redirect(url_for("lista_produtos"))

@app.route("/estoque/produtos/importar", methods=["POST"])
@login_required
def importar_produtos():
    """Importa produtos do arquivo Excel"""
    # Verificar permissão usando helper
    if not pode_importar(current_user, "produtos"):
        cargos_permitidos = ", ".join(get_cargos_permitidos_importacao("produtos"))
        flash(
            f"Apenas {cargos_permitidos} podem importar produtos.",
            "danger"
        )
        return redirect(url_for("lista_produtos"))

    # Verificar disponibilidade do Excel (pandas/openpyxl)
    if not EXCEL_AVAILABLE and not ensure_excel_available():
        error_msg = "❌ Erro: Funcionalidade de importação Excel indisponível."
        if EXCEL_ERROR_MESSAGE:
            # Log detalhado no backend; para o usuário mantemos mensagem amigável
            print(f"📊 Erro Excel (importar_produtos): {EXCEL_ERROR_MESSAGE}")
            flash(error_msg + " Contate o administrador.", "danger")
        else:
            flash(error_msg + " Contate o administrador.", "danger")

        return redirect(url_for("lista_produtos"))

    try:
        # Import local para garantir que 'openpyxl' esteja definido neste escopo
        import openpyxl

        arquivo = request.files.get("arquivo")
        if not arquivo:
            flash("Nenhum arquivo foi enviado.", "danger")
            return redirect(url_for("lista_produtos"))

        if not arquivo.filename.endswith((".xlsx", ".xls")):
            flash(
                "Formato de arquivo inválido. Use .xlsx ou .xls", "danger"
            )
            return redirect(url_for("lista_produtos"))

        # Ler arquivo Excel
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active

        produtos_criados = []
        produtos_duplicados = []
        erros = []

        # Processar linhas (pular cabeçalho)
        for row_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            try:
                # Extrair dados das colunas
                codigo = str(row[0]).strip() if row[0] else None
                nome = str(row[1]).strip() if row[1] else None
                referencia = str(row[2]).strip() if row[2] else None
                codigo_barra = str(row[3]).strip() if row[3] else None
                ncm = str(row[4]).strip() if row[4] else None

                # Validar campos obrigatórios
                if not codigo or not nome:
                    erros.append(
                        f"Linha {row_num}: Código e Nome são obrigatórios"
                    )
                    continue

                # Verificar se já existe
                existe = Produto.query.filter_by(
                    codigo=codigo, empresa_id=current_user.empresa_id
                ).first()

                if existe:
                    # ATUALIZAR PRODUTO EXISTENTE (Upsert)
                    existe.nome = nome
                    if referencia: existe.referencia = referencia
                    if codigo_barra: existe.codigo_barra = codigo_barra
                    if ncm: existe.ncm = ncm

                    # Atualizar timestamp
                    existe.data_atualizacao = datetime.utcnow()

                    produtos_criados.append(f"{codigo} - {nome} (Atualizado)")
                    continue

                # Criar novo produto
                produto = Produto(
                    codigo=codigo,
                    nome=nome,
                    referencia=referencia,
                    codigo_barra=codigo_barra,
                    ncm=ncm,
                    categoria="Outro",  # Categoria padrão
                    unidade="UN",  # Unidade padrão
                    empresa_id=current_user.empresa_id,
                    estoque_atual=0,
                    estoque_minimo=0,
                    custo_medio=0,
                    preco_venda=0,
                    preco_servico=0,  # Inicializar preço do serviço
                    ativo=True,
                )

                db.session.add(produto)
                produtos_criados.append(f"{codigo} - {nome}")

            except Exception as e:
                erros.append(f"Linha {row_num}: {str(e)}")

        # Commit das alterações
        if produtos_criados:
            db.session.commit()

        # Montar mensagem de resultado
        mensagens = []
        if produtos_criados:
            mensagens.append(f"{len(produtos_criados)} produto(s) importado(s) com sucesso!")
        if produtos_duplicados:
            mensagens.append(f"⚠️ {len(produtos_duplicados)} produto(s) já cadastrado(s) (ignorados)")
        if erros:
            mensagens.append(f"{len(erros)} erro(s) encontrado(s)")

        # Flash principal
        if produtos_criados:
            flash(mensagens[0], "success")

        # Flash warnings e erros - mostrar apenas quantidade
        if produtos_duplicados:
            # Mostrar apenas os primeiros 3 códigos de forma compacta
            codigos_mostrar = [
                p.split(" - ")[0] for p in produtos_duplicados[:3]
            ]
            msg_duplicados = f"⚠️ {len(produtos_duplicados)} produto(s) já cadastrado(s): {', '.join(codigos_mostrar)}"
            if len(produtos_duplicados) > 3:
                msg_duplicados += f" e mais {len(produtos_duplicados) - 3} produto(s)"
            flash(msg_duplicados, "warning")

        if erros:
            flash(f"{len(erros)} erro(s) encontrado(s). Verifique a planilha e tente novamente.", "danger")

        return redirect(url_for("lista_produtos"))

    except ImportError:
        flash(
            "Biblioteca openpyxl não instalada. Execute: pip install openpyxl",
            "danger",
        )
        return redirect(url_for("lista_produtos"))
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao importar produtos: {str(e)}", "danger")
        return redirect(url_for("lista_produtos"))

@app.route("/estoque/movimentacao/nova", methods=["GET", "POST"])
@login_required
def nova_movimentacao():
    """Registrar nova movimentação de estoque"""
    from permissoes_estoque import (
        get_motivos_permitidos,
        usuario_pode_usar_motivo,
    )

    if (
        not current_user.pode_movimentar_estoque and
        current_user.cargo != "admin"
    ):
        flash("Você não tem permissão para movimentar estoque.", "danger")
        return redirect(url_for("lista_estoque"))

    form = EstoqueMovimentoForm()

    produtos = (
        Produto.query.filter_by(empresa_id=current_user.empresa_id, ativo=True)
        .order_by(Produto.nome)
        .all()
    )
    form.produto_id.choices = [(0, "Selecione o produto")] + [
        (p.id, f"{p.codigo} - {p.nome}") for p in produtos
    ]

    clientes = (
        Cliente.query.filter_by(empresa_id=current_user.empresa_id)
        .order_by(Cliente.nome)
        .all()
    )
    form.cliente_id.choices = [(0, "Nenhum")] + [
        (c.id, c.nome) for c in clientes
    ]

    ordens = (
        OrdemServico.query.filter_by(empresa_id=current_user.empresa_id)
        .filter(
            OrdemServico.status.in_(
                ["aprovada", "em_andamento", "aguardando_peca"]
            )
        )
        .all()
    )
    form.ordem_servico_id.choices = [(0, "Nenhuma")] + [
        (os.id, f"{os.numero_os} - {os.titulo}") for os in ordens
    ]

    # Configurar motivos baseado no cargo do usuário e tipo de movimento
    tipo_movimento = request.form.get("tipo", "entrada")
    motivos_permitidos = get_motivos_permitidos(
        current_user.cargo, tipo_movimento
    )
    form.motivo.choices = [("", "Selecione o motivo")] + motivos_permitidos

    if form.validate_on_submit():
        try:
            # Validar permissão do motivo selecionado
            if not usuario_pode_usar_motivo(
                current_user.cargo, form.tipo.data, form.motivo.data
            ):
                flash(
                    "Você não tem permissão para usar este motivo de movimentação!",
                    "danger",
                )
                return render_template(
                    "estoque/movimentacao_form.html", form=form
                )

            produto = Produto.query.get(form.produto_id.data)
            if not produto:
                flash("Produto não encontrado!", "danger")
                return render_template(
                    "estoque/movimentacao_form.html", form=form
                )

            if form.tipo.data == "saida":
                if form.quantidade.data > produto.estoque_atual:
                    flash(
                        f"Quantidade insuficiente! Estoque atual: {produto.estoque_atual}",
                        "danger",
                    )
                    return render_template(
                        "estoque/movimentacao_form.html", form=form
                    )

            movimento = EstoqueMovimento(
                produto_id=form.produto_id.data,
                tipo=form.tipo.data,
                motivo=form.motivo.data,
                quantidade=form.quantidade.data,
                valor_unitario=form.valor_unitario.data or 0,
                valor_total=(
                    form.quantidade.data * (form.valor_unitario.data or 0)
                ),
                documento=form.documento.data,
                ordem_servico_id=(
                    form.ordem_servico_id.data
                    if form.ordem_servico_id.data
                    else None
                ),
                fornecedor=form.fornecedor.data,
                cliente_id=(
                    form.cliente_id.data if form.cliente_id.data else None
                ),
                observacoes=form.observacoes.data,
                usuario_id=current_user.id,
                empresa_id=current_user.empresa_id,
                data_movimento=datetime.now(),
            )

            if form.tipo.data == "entrada":
                produto.estoque_atual += form.quantidade.data
                if form.valor_unitario.data and form.valor_unitario.data > 0:
                    total_anterior = produto.custo_medio * (
                        produto.estoque_atual - form.quantidade.data
                    )
                    total_novo = (
                        form.valor_unitario.data * form.quantidade.data
                    )
                    produto.custo_medio = (
                        total_anterior + total_novo
                    ) / produto.estoque_atual
            else:
                produto.estoque_atual -= form.quantidade.data

            db.session.add(movimento)
            db.session.commit()

            flash(
                f"Movimentação registrada com sucesso! Estoque atual: {produto.estoque_atual}",
                "success",
            )
            return redirect(url_for("visualizar_produto", id=produto.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao registrar movimentação: {str(e)}", "danger")

    return render_template("estoque/movimentacao_form.html", form=form)

@app.route("/api/estoque/motivos-permitidos/<tipo_movimento>")
@login_required
def get_motivos_permitidos_api(tipo_movimento):
    """API para obter motivos permitidos baseado no cargo do usuário"""
    from permissoes_estoque import get_motivos_permitidos
    from flask import jsonify

    if tipo_movimento not in ["entrada", "saida"]:
        return jsonify({"error": "Tipo de movimento inválido"}), 400

    motivos = get_motivos_permitidos(current_user.cargo, tipo_movimento)

    return jsonify(
        {
            "success": True,
            "motivos": [{"value": m[0], "label": m[1]} for m in motivos],
            "cargo": current_user.cargo,
        }
    )

@app.route("/estoque/permissoes")
@login_required
def permissoes_estoque():
    """Página com documentação da hierarquia de permissões"""
    if current_user.cargo not in ["admin", "gerente"]:
        flash("Acesso restrito a administradores e gerentes.", "danger")
        return redirect(url_for("lista_estoque"))

    return render_template("estoque/permissoes_estoque.html")

@app.route("/estoque/movimentacoes")
@login_required
def lista_movimentacoes():
    """Lista todas as movimentações de estoque"""
    if not current_user.pode_acessar_estoque and current_user.cargo != "admin":
        flash("Você não tem permissão para acessar o estoque.", "danger")
        return redirect(url_for("dashboard"))

    tipo = request.args.get("tipo", "")
    motivo = request.args.get("motivo", "")
    produto_id = request.args.get("produto_id", "")
    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")

    query = EstoqueMovimento.query.filter_by(
        empresa_id=current_user.empresa_id
    )

    if tipo:
        query = query.filter_by(tipo=tipo)
    if motivo:
        query = query.filter_by(motivo=motivo)
    if produto_id:
        query = query.filter_by(produto_id=produto_id)
    if data_inicio:
        query = query.filter(EstoqueMovimento.data_movimento >= data_inicio)
    if data_fim:
        query = query.filter(EstoqueMovimento.data_movimento <= data_fim)

    movimentacoes = query.order_by(
        EstoqueMovimento.data_movimento.desc()
    ).all()

    produtos = (
        Produto.query.filter_by(empresa_id=current_user.empresa_id)
        .order_by(Produto.nome)
        .all()
    )

    return render_template(
        "estoque/movimentacoes.html",
        movimentacoes=movimentacoes,
        produtos=produtos,
        filtros={
            "tipo": tipo,
            "motivo": motivo,
            "produto_id": produto_id,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
        },
    )

# ====================================================================
# ROTA DE DIAGNÓSTICO DO SISTEMA
# ====================================================================

@app.route("/diagnostico-excel")
@login_required
def diagnostico_excel():
    """Endpoint de diagnóstico para verificar status das bibliotecas Excel"""
    # Apenas admin pode acessar
    if current_user.cargo != "admin":
        flash("Acesso negado! Apenas administradores podem acessar o diagnóstico.", "danger")
        return redirect(url_for("dashboard"))
    
    diagnostico = {
        "excel_disponivel": EXCEL_AVAILABLE,
        "erro_excel": EXCEL_ERROR_MESSAGE,
        "pandas_version": None,
        "openpyxl_version": None,
        "numpy_version": None,
        "python_version": sys.version,
    }
    
    try:
        import pandas as pd_test
        diagnostico["pandas_version"] = pd_test.__version__
    except Exception as e:
        diagnostico["pandas_error"] = str(e)
    
    try:
        import openpyxl as openpyxl_test
        diagnostico["openpyxl_version"] = openpyxl_test.__version__
    except Exception as e:
        diagnostico["openpyxl_error"] = str(e)
    
    try:
        import numpy as np_test
        diagnostico["numpy_version"] = np_test.__version__
    except Exception as e:
        diagnostico["numpy_error"] = str(e)
    
    return jsonify(diagnostico)

# ====================================================================
# FIM DAS ROTAS DE ESTOQUE  # ====================================================================

# ====================================================================
# HANDLERS DE ERRO (layout responsivo/profissional)
# ====================================================================

@app.errorhandler(500)
def handle_500(error):
    """Handler de erro 500 com diagnóstico de conexão"""
    try:
        app.logger.error(f"Erro 500: {error}")
        
        # Diagnóstico de conexão com banco
        db_status = "desconhecido"
        db_error = None
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            db_status = "conectado"
        except Exception as db_err:
            db_status = "erro"
            db_error = str(db_err)
            app.logger.error(f"Erro de conexão com banco: {db_error}")
        
        # Passar informações para o template
        return render_template(
            'errors/500.html',
            db_status=db_status,
            db_error=db_error
        ), 500
    except Exception as e:
        app.logger.error(f"Erro no handler 500: {e}")
        return render_template('errors/500.html'), 500

if __name__ == "__main__":
    # Se estiver rodando no Railway, executar fix do banco antes
    if os.environ.get('RAILWAY_ENVIRONMENT'):
        print("\n🚂 Ambiente Railway detectado - verificando banco de dados...")
        try:
            import subprocess
            import sys
            result = subprocess.run(
                [sys.executable, 'fix_database_railway.py'],
                capture_output=True,
                text=True,
                timeout=30
            )
            if result.returncode == 0:
                print("✅ Banco de dados verificado/corrigido com sucesso")
            else:
                print(f"⚠️ Aviso ao verificar banco: {result.stderr}")
        except Exception as e:
            print(f"⚠️ Erro ao verificar banco: {e}")
    
    print("\n" + "=" * 70)
    print("🚀 SISTEMA DE GESTÃO DE METAS E COMISSÕES - VERSÃO COMPLETA")
    print("=" * 70)
    print("\n✨ Novos Recursos:")
    print("   🔐 Sistema de autenticação (Login/Registro)")
    print("   💾 Banco de dados SQLite/PostgreSQL")
    print("   👥 Gerenciamento de vendedores")
    print("   📊 Gerenciamento de metas")
    print("   🎯 Cálculo automático de comissões")
    print("   ⏰ Backup automático agendado")
    print("\n[OK] Servidor iniciado com sucesso!")
    print("=> Acesse: http://127.0.0.1:5001/login")
    print("[KEY] Pressione CTRL+C para parar o servidor\n")
    print("=" * 70 + "\n")

    app.run(debug=True, port=5001)
